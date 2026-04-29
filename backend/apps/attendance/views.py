from collections import defaultdict
import csv
from datetime import date, datetime
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.db import IntegrityError, models, transaction
from django.db.models import Count, Case, When, IntegerField, Max
from rest_framework import permissions, status
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from config.pagination import ApiPageNumberPagination

from apps.core.models import Class, Section
from apps.students.models import Student

from .models import StudentAttendance, StudentAttendanceBulk
from .serializers import (
    StudentAttendanceSerializer,
    StudentAttendanceStoreRequestSerializer,
    StudentSearchRequestSerializer,
)
from .services import send_present_attendance_notifications


class AttendanceTenantMixin:
    permission_classes = [permissions.IsAuthenticated]

    def _required_permission_code(self):
        class_name = self.__class__.__name__.lower()
        if "import" in class_name or "downloadsample" in class_name:
            return "student_info.student_attendance_import.view"
        return "student_info.student_attendance.view"

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        user = request.user
        if user.is_superuser:
            return
        code = self._required_permission_code()
        if not hasattr(user, "has_permission_code") or not user.has_permission_code(code):
            raise PermissionDenied("You do not have permission to perform this action.")

    def school_filter(self, request):
        return {} if request.user.is_superuser else {"school_id": request.user.school_id}


class StudentAttendanceListCreateAPIView(AttendanceTenantMixin, APIView):
    def get(self, request):
        queryset = StudentAttendance.objects.select_related("student")
        if not request.user.is_superuser:
            if not request.user.school_id:
                return Response([])
            queryset = queryset.filter(school_id=request.user.school_id)

        params = request.query_params
        if params.get("class_id"):
            queryset = queryset.filter(class_id=params["class_id"])
        if params.get("section_id"):
            queryset = queryset.filter(section_id=params["section_id"])
        if params.get("student_id"):
            queryset = queryset.filter(student_id=params["student_id"])
        if params.get("date"):
            queryset = queryset.filter(attendance_date=params["date"])
        if params.get("month"):
            queryset = queryset.filter(attendance_date__month=params["month"])
        if params.get("year"):
            queryset = queryset.filter(attendance_date__year=params["year"])
        if params.get("academic_year_id"):
            queryset = queryset.filter(academic_year_id=params["academic_year_id"])

        ordered_queryset = queryset.order_by("-attendance_date", "student_id")
        paginator = ApiPageNumberPagination()
        page = paginator.paginate_queryset(ordered_queryset, request)
        serializer = StudentAttendanceSerializer(page if page is not None else ordered_queryset, many=True)
        if page is not None:
            return paginator.get_paginated_response(serializer.data)
        return Response(serializer.data)

    def post(self, request):
        serializer = StudentAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            serializer.save(school=request.user.school)
        except IntegrityError:
            return Response(
                {"detail": "Attendance already exists for this student on the selected date."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class StudentAttendanceRetrieveUpdateDeleteAPIView(AttendanceTenantMixin, APIView):
    def get_object(self, request, pk):
        queryset = StudentAttendance.objects.all()
        if not request.user.is_superuser:
            queryset = queryset.filter(school_id=request.user.school_id)
        return get_object_or_404(queryset, pk=pk)

    def get(self, request, pk):
        obj = self.get_object(request, pk)
        return Response(StudentAttendanceSerializer(obj).data)

    def put(self, request, pk):
        obj = self.get_object(request, pk)
        serializer = StudentAttendanceSerializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(school=request.user.school)
        return Response(serializer.data)

    def patch(self, request, pk):
        obj = self.get_object(request, pk)
        serializer = StudentAttendanceSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(school=request.user.school)
        return Response(serializer.data)

    def delete(self, request, pk):
        obj = self.get_object(request, pk)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class StudentAttendanceIndexAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP index(): returns classes used in the criteria form."""

    def get(self, request):
        classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")
        return Response([
            {"id": c.id, "class_name": c.name}
            for c in classes
        ])


class StudentSearchAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentSearch() with enhanced validation."""

    def post(self, request):
        req = StudentSearchRequestSerializer(data=request.data)
        req.is_valid(raise_exception=True)

        class_id = req.validated_data["class"]
        section_id = req.validated_data["section"]
        attendance_date = req.validated_data["attendance_date"]

        # Validate date is not in future
        from datetime import date as date_type
        if attendance_date > date_type.today():
            return Response(
                {"success": False, "message": "Attendance cannot be marked for future dates"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        include_legacy_payload = str(request.data.get("include_legacy_payload", "false")).lower() in {
            "1",
            "true",
            "yes",
        }

        students = list(
            Student.objects.filter(
            current_class_id=class_id,
            is_active=True,
            **self.school_filter(request),
            )
            .filter(
            # Include students assigned to this section OR students with the right class but no section
            models.Q(current_section_id=section_id) | models.Q(current_section_id__isnull=True)
            )
            .order_by("id")
            .values("id", "admission_no", "first_name", "last_name", "roll_no")
        )

        attendance_type = ""

        # Query attendance by student_ids (not class/section) so records saved
        # without class_id/section_id are still found (backward-compatible).
        student_ids = [s["id"] for s in students]
        attendance_rows = {
            row.student_id: row
            for row in StudentAttendance.objects.filter(
                attendance_date=attendance_date,
                student_id__in=student_ids,
                **self.school_filter(request),
            )
        }

        table_students = []
        for student in students:
            sid = student["id"]
            att = attendance_rows.get(sid)
            if attendance_type == "" and att:
                attendance_type = att.attendance_type
            table_students.append(
                {
                    "id": sid,
                    "admission_no": student["admission_no"],
                    "first_name": student["first_name"],
                    "last_name": student["last_name"],
                    "roll_no": student["roll_no"],
                    "attendance_type": att.attendance_type if att else None,
                    "attendance_note": att.notes if att else "",
                    "arrival_time": att.arrival_time.strftime("%H:%M") if att and att.arrival_time else None,
                    "sign_in_time": att.sign_in_time.strftime("%H:%M") if att and att.sign_in_time else None,
                    "sign_out_time": att.sign_out_time.strftime("%H:%M") if att and att.sign_out_time else None,
                    "pickup_time": att.pickup_time.strftime("%H:%M") if att and att.pickup_time else None,
                    "pickup_by": att.pickup_by if att else "",
                    "lunch": att.lunch if att else False,
                }
            )

        response_data = {
            "date": attendance_date,
            "class_id": class_id,
            "section_id": section_id,
            "attendance_type": attendance_type,
            "students": table_students,
        }

        # Optional backward-compatible payload for legacy clients.
        if include_legacy_payload:
            class_info = Class.objects.filter(id=class_id).first()
            section_info = Section.objects.filter(id=section_id).first()
            classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")

            already_assigned_students = []
            new_students = []
            for student in students:
                sid = student["id"]
                att = attendance_rows.get(sid)
                if att:
                    already_assigned_students.append(
                        {
                            "id": att.id,
                            "student_id": att.student_id,
                            "attendance_type": att.attendance_type,
                            "notes": att.notes,
                            "lunch": att.lunch,
                        }
                    )
                else:
                    new_students.append(student)

            response_data.update(
                {
                    "classes": [{"id": c.id, "class_name": c.name} for c in classes],
                    "already_assigned_students": already_assigned_students,
                    "new_students": new_students,
                    "search_info": {
                        "class_name": class_info.name if class_info else "",
                        "section_name": section_info.name if section_info else "",
                        "date": attendance_date,
                    },
                }
            )

        return Response(response_data)


class StudentAttendanceStoreAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceStore() with enhanced validation."""

    @staticmethod
    def _normalize_time(value):
        if value in (None, "", "—"):
            return None
        if hasattr(value, "hour") and hasattr(value, "minute"):
            return value
        if isinstance(value, str):
            text = value.strip()
            for fmt in ["%H:%M", "%H:%M:%S"]:
                try:
                    return datetime.strptime(text, fmt).time()
                except ValueError:
                    continue
        raise ValueError("Invalid time format. Use HH:MM.")

    def post(self, request):
        try:
            req = StudentAttendanceStoreRequestSerializer(data=request.data)
            req.is_valid(raise_exception=True)
            data = req.validated_data

            attendance_map = data.get("attendance") or data.get("attendance_type") or {}
            lunch_map = data.get("lunch") or {}
            arrival_time_map = data.get("arrival_time") or {}
            sign_in_time_map = data.get("sign_in_time") or {}
            sign_out_time_map = data.get("sign_out_time") or {}
            pickup_time_map = data.get("pickup_time") or {}
            pickup_by_map = data.get("pickup_by") or {}
            lock_attendance = data.get("lock_attendance", False)
            note_map_early = data.get("note") or data.get("attendance_note") or {}
            has_time_payload = any([arrival_time_map, sign_in_time_map, sign_out_time_map, pickup_time_map, pickup_by_map])
            has_note_payload = bool(note_map_early)

            if not attendance_map and not lunch_map and not has_time_payload and not has_note_payload:
                return Response(
                    {"success": False, "message": "Please provide attendance, lunch, time or note data"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            attendance_ids = set(data.get("id", []))
            attendance_ids_str = {str(i) for i in attendance_ids}
            lunch_count = len([v for k, v in lunch_map.items() if str(k) in attendance_ids_str])
            status_count = len([v for k, v in attendance_map.items() if str(k) in attendance_ids_str])

            # Require full status coverage only for status submissions.
            if attendance_map and status_count != len(attendance_ids):
                return Response(
                    {"success": False, "message": "Please mark attendance for all students"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Require full lunch coverage only for lunch-only submissions.
            if not attendance_map and lunch_map and lunch_count != len(attendance_ids):
                return Response(
                    {"success": False, "message": "Please mark lunch status for all students"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            school_scope = self.school_filter(request)
            student_scope = Student.objects.filter(id__in=attendance_ids, **school_scope)
            if data.get("class_id"):
                student_scope = student_scope.filter(current_class_id=data.get("class_id"))
            if data.get("section_id"):
                student_scope = student_scope.filter(
                    models.Q(current_section_id=data.get("section_id")) | models.Q(current_section_id__isnull=True)
                )

            students = list(student_scope.select_related("guardian"))
            if not students:
                return Response(
                    {"success": False, "message": "No students found for selected class and section"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            student_map = {s.id: s for s in students}

            note_map = data.get("note") or data.get("attendance_note") or {}

            with transaction.atomic():
                for student_id in data["id"]:
                    if student_id not in student_map:
                        continue

                    existing_lunch = False
                    existing_type = None
                    existing_arrival_time = None
                    existing_sign_in_time = None
                    existing_sign_out_time = None
                    existing_pickup_time = None
                    existing_pickup_by = ""
                    existing_note = ""
                    # Delete existing attendance for same date
                    existing = StudentAttendance.objects.filter(
                        student_id=student_id,
                        attendance_date=data["date"],
                        **school_scope,
                    ).first()
                    if existing and not existing.is_locked:
                        existing_lunch = existing.lunch
                        existing_type = existing.attendance_type
                        existing_arrival_time = existing.arrival_time
                        existing_sign_in_time = existing.sign_in_time
                        existing_sign_out_time = existing.sign_out_time
                        existing_pickup_time = existing.pickup_time
                        existing_pickup_by = existing.pickup_by
                        existing_note = existing.notes
                        existing.delete()
                    elif existing and existing.is_locked:
                        return Response(
                            {"success": False, "message": "Cannot update locked attendance. Only admin can override."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    attendance = StudentAttendance()
                    attendance.student_id = student_id
                    if data.get("mark_holiday"):
                        attendance.attendance_type = "H"
                        attendance.notes = "Holiday"
                    else:
                        raw_type = attendance_map.get(str(student_id))
                        if raw_type is None:
                            raw_type = attendance_map.get(student_id)
                        
                        # Validate attendance type
                        valid_types = ["P", "A", "L", "F", "H"]
                        if raw_type and raw_type not in valid_types:
                            return Response(
                                {"success": False, "message": "Invalid attendance status"},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                        attendance.attendance_type = raw_type or existing_type or "P"

                        raw_note = note_map.get(str(student_id))
                        if raw_note is None:
                            raw_note = note_map.get(student_id)
                        
                        # Keep previously stored reason when note is omitted in incremental updates.
                        note_text = existing_note if raw_note is None else (raw_note or "")
                        if len(note_text) > 250:
                            return Response(
                                {"success": False, "message": "Note cannot exceed 250 characters"},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                        attendance.notes = note_text

                    raw_arrival = arrival_time_map.get(str(student_id), arrival_time_map.get(student_id))
                    raw_sign_in = sign_in_time_map.get(str(student_id), sign_in_time_map.get(student_id))
                    raw_sign_out = sign_out_time_map.get(str(student_id), sign_out_time_map.get(student_id))
                    raw_pickup = pickup_time_map.get(str(student_id), pickup_time_map.get(student_id))
                    raw_pickup_by = pickup_by_map.get(str(student_id), pickup_by_map.get(student_id))

                    try:
                        attendance.arrival_time = self._normalize_time(raw_arrival) if raw_arrival is not None else existing_arrival_time
                        attendance.sign_in_time = self._normalize_time(raw_sign_in) if raw_sign_in is not None else existing_sign_in_time
                        attendance.sign_out_time = self._normalize_time(raw_sign_out) if raw_sign_out is not None else existing_sign_out_time
                        attendance.pickup_time = self._normalize_time(raw_pickup) if raw_pickup is not None else existing_pickup_time
                    except ValueError as e:
                        return Response(
                            {"success": False, "message": str(e)},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    attendance.pickup_by = str(raw_pickup_by).strip() if raw_pickup_by is not None else existing_pickup_by

                    raw_lunch = lunch_map.get(str(student_id))
                    if raw_lunch is None:
                        raw_lunch = lunch_map.get(student_id)
                    attendance.lunch = bool(raw_lunch) if raw_lunch is not None else existing_lunch

                    attendance.attendance_date = data["date"]
                    attendance.school = request.user.school
                    attendance.academic_year_id = data.get("academic_year_id")
                    attendance.class_id = data.get("class_id")
                    attendance.section_id = data.get("section_id")
                    attendance.marked_by = request.user
                    attendance.is_locked = lock_attendance
                    attendance.save()

                    if attendance_map and attendance.attendance_type == "P" and existing_type != "P":
                        student = student_map.get(student_id)
                        if student:
                            send_present_attendance_notifications(student, attendance.attendance_date)

            return Response(
                {"success": True, "message": "Attendance saved successfully"},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            detail = getattr(e, "detail", None)
            if detail is not None:
                if isinstance(detail, dict):
                    first_message = "Validation failed"
                    for value in detail.values():
                        if isinstance(value, (list, tuple)) and value:
                            first_message = str(value[0])
                            break
                        if isinstance(value, str):
                            first_message = value
                            break
                    return Response(
                        {"success": False, "message": first_message, "field_errors": detail},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                return Response(
                    {"success": False, "message": str(detail)},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(
                {"success": False, "message": f"Failed to save attendance: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentAttendanceHolidayAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceHoliday()."""

    def post(self, request):
        class_id = request.data.get("class_id")
        section_id = request.data.get("section_id")
        attendance_date = request.data.get("attendance_date")
        purpose = request.data.get("purpose", "mark")

        students = Student.objects.filter(
            current_class_id=class_id,
            current_section_id=section_id,
            is_active=True,
            **self.school_filter(request),
        )

        if purpose == "mark":
            for student in students:
                existing = StudentAttendance.objects.filter(
                    student_id=student.id,
                    attendance_date=attendance_date,
                    **self.school_filter(request),
                ).first()
                if existing:
                    existing.delete()

                attendance = StudentAttendance(
                    attendance_type="H",
                    notes="Holiday",
                    attendance_date=attendance_date,
                    student_id=student.id,
                    school=request.user.school,
                    academic_year_id=request.data.get("academic_year_id"),
                    class_id=class_id,
                    section_id=section_id,
                )
                attendance.save()

        elif purpose == "unmark":
            for student in students:
                existing = StudentAttendance.objects.filter(
                    student_id=student.id,
                    attendance_date=attendance_date,
                    **self.school_filter(request),
                ).first()
                if existing:
                    existing.delete()

        return Response({"status": "ok"}, status=status.HTTP_200_OK)


class StudentAttendanceMonthlyReportAPIView(AttendanceTenantMixin, APIView):
    def get(self, request):
        class_id = request.query_params.get("class_id")
        section_id = request.query_params.get("section_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")

        if not class_id or not section_id or not month or not year:
            return Response(
                {"detail": "class_id, section_id, month, year are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            class_id_int = int(class_id)
            section_id_int = int(section_id)
            month_int = int(month)
            year_int = int(year)
        except (TypeError, ValueError):
            return Response(
                {"detail": "class_id, section_id, month, year must be numeric."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if month_int < 1 or month_int > 12:
            return Response(
                {"detail": "month must be between 1 and 12."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if year_int < 2000 or year_int > 2100:
            return Response(
                {"detail": "year must be a valid 4-digit year."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        records = (
            StudentAttendance.objects.filter(
                class_id=class_id_int,
                section_id=section_id_int,
                attendance_date__month=month_int,
                attendance_date__year=year_int,
                **self.school_filter(request),
            )
            .select_related("student")
            .order_by("student_id", "attendance_date")
        )

        grouped = defaultdict(lambda: {"P": 0, "A": 0, "L": 0, "F": 0, "H": 0, "name": "", "admission_no": ""})
        for rec in records:
            sid = rec.student_id
            grouped[sid]["name"] = f"{rec.student.first_name} {rec.student.last_name}".strip()
            grouped[sid]["admission_no"] = rec.student.admission_no
            grouped[sid][rec.attendance_type] = grouped[sid].get(rec.attendance_type, 0) + 1

        result = [
            {
                "student_id": sid,
                "admission_no": info["admission_no"],
                "name": info["name"],
                "present": info["P"],
                "absent": info["A"],
                "late": info["L"],
                "half_day": info["F"],
                "holiday": info["H"],
            }
            for sid, info in grouped.items()
        ]
        paginator = ApiPageNumberPagination()
        page = paginator.paginate_queryset(result, request)
        if page is not None:
            return paginator.get_paginated_response(page)
        return Response(result)


class StudentAttendanceImportAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceImport() for criteria/form data."""

    def get(self, request):
        classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")
        return Response(
            {
                "classes": [{"id": c.id, "class_name": c.name} for c in classes],
            }
        )


class StudentAttendanceDownloadSampleAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP downloadStudentAtendanceFile()."""
    
    # Allow any authenticated user to download sample (no special permission needed)
    def _required_permission_code(self):
        return None  # No permission code check for sample download

    def get(self, request):
        try:
            from openpyxl import Workbook
        except Exception:
            return Response({"detail": "openpyxl is required for sample export."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "student_attendance_sheet"
        headers = [
            "admission_no",
            "attendance_date",
            "attendance_type",
            "note",
            "arrival_time",
            "sign_in_time",
            "sign_out_time",
            "pickup_time",
            "pickup_by",
            "lunch",
        ]
        sheet.append(headers)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_attendance_sheet.xlsx"'
        return response


class StudentAttendanceBulkStoreAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceBulkStore()."""

    parser_classes = [MultiPartParser, FormParser]

    def _normalize_date(self, value):
        if value is None or value == "":
            return None
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            try:
                from openpyxl.utils.datetime import from_excel

                converted = from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def post(self, request):
        attendance_date = request.data.get("attendance_date")
        class_id = request.data.get("class")
        section_id = request.data.get("section")
        uploaded_file = request.FILES.get("file")

        # Validation: Required fields
        if not attendance_date or not class_id or not section_id or not uploaded_file:
            return Response(
                {"detail": "attendance_date, file, class and section are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validation: File format
        ext = uploaded_file.name.split(".")[-1].lower() if "." in uploaded_file.name else ""
        if ext not in {"csv", "xlsx", "xls"}:
            return Response(
                {"detail": "The file must be a file of type: xlsx, csv or xls"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validation: File size (5MB max)
        file_size_mb = uploaded_file.size / (1024 * 1024)
        if file_size_mb > 5:
            return Response(
                {"detail": f"File size exceeds 5MB limit (current: {file_size_mb:.2f}MB)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validation: Date format and range
        request_date = self._normalize_date(attendance_date)
        if not request_date:
            return Response({"detail": "Invalid attendance_date."}, status=status.HTTP_400_BAD_REQUEST)

        # Validation: Date not in future
        from datetime import date as date_type
        if request_date > date_type.today():
            return Response(
                {"detail": "Cannot import attendance for future dates."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        school_scope = self.school_filter(request)
        class_qs = Class.objects.filter(id=class_id, **school_scope)
        if not class_qs.exists():
            return Response({"detail": "Invalid class selected."}, status=status.HTTP_400_BAD_REQUEST)
        section_qs = Section.objects.filter(id=section_id, school_class_id=class_id)
        if not request.user.is_superuser:
            section_qs = section_qs.filter(school_class__school_id=request.user.school_id)
        if not section_qs.exists():
            return Response({"detail": "Invalid section selected for class."}, status=status.HTTP_400_BAD_REQUEST)

        imported_rows = []
        errors: list[dict] = []
        row_number = 2  # Start from 2 (header is row 1)

        # Parse file
        if ext == "csv":
            try:
                content = uploaded_file.read().decode("utf-8", errors="ignore")
                reader = csv.DictReader(StringIO(content))
                for row in reader:
                    imported_rows.append((row_number, row))
                    row_number += 1
            except Exception as e:
                return Response(
                    {"detail": f"Failed to parse CSV file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                from openpyxl import load_workbook
            except Exception:
                return Response({"detail": "openpyxl is required for xlsx import."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            try:
                workbook = load_workbook(uploaded_file, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    return Response({"detail": "File is empty."}, status=status.HTTP_400_BAD_REQUEST)

                headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                header_map = {h: i for i, h in enumerate(headers)}

                for idx, row in enumerate(rows[1:], start=2):
                    row_dict = {}
                    for header, col_idx in header_map.items():
                        if col_idx < len(row):
                            row_dict[header] = row[col_idx]
                    imported_rows.append((idx, row_dict))
            except Exception as e:
                return Response(
                    {"detail": f"Failed to parse Excel file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        school_filter = school_scope
        processed_count = 0
        failed_count = 0

        # Validate and process rows
        valid_attendance_types = ["P", "A", "L", "F", "H"]

        def normalize_time_for_import(raw_value):
            if raw_value in (None, "", "—"):
                return None
            if isinstance(raw_value, str):
                text = raw_value.strip()
            else:
                text = str(raw_value).strip()
            for fmt in ["%H:%M", "%H:%M:%S"]:
                try:
                    return datetime.strptime(text, fmt).time()
                except ValueError:
                    continue
            return "invalid"

        for row_num, row_data in imported_rows:
            row_errors = []
            admission_no = str(row_data.get("admission_no") or "").strip()
            attendance_type = str(row_data.get("attendance_type") or "").strip()
            note = str(row_data.get("note") or "").strip()
            arrival_time = normalize_time_for_import(row_data.get("arrival_time"))
            sign_in_time = normalize_time_for_import(row_data.get("sign_in_time"))
            sign_out_time = normalize_time_for_import(row_data.get("sign_out_time"))
            pickup_time = normalize_time_for_import(row_data.get("pickup_time"))
            pickup_by = str(row_data.get("pickup_by") or "").strip()
            lunch_text = str(row_data.get("lunch") or "").strip().lower()
            lunch_value = lunch_text in {"1", "true", "yes", "y", "t"}

            # Validate admission_no
            if not admission_no:
                row_errors.append({"field": "admission_no", "message": "Admission number is required"})

            # Validate attendance_type
            if attendance_type and attendance_type not in valid_attendance_types:
                row_errors.append({
                    "field": "attendance_type",
                    "message": "Invalid attendance type. Use one of P (Present), A (Absent), L (Late), F (Half Day), H (Holiday)."
                })

            # Validate note length
            if len(note) > 250:
                row_errors.append({"field": "note", "message": "Note cannot exceed 250 characters"})

            for field_name, value in [
                ("arrival_time", arrival_time),
                ("sign_in_time", sign_in_time),
                ("sign_out_time", sign_out_time),
                ("pickup_time", pickup_time),
            ]:
                if value == "invalid":
                    row_errors.append({"field": field_name, "message": "Invalid time format. Use HH:MM"})

            if row_errors:
                for err in row_errors:
                    errors.append({
                        "row": row_num,
                        "field": err["field"],
                        "message": err["message"]
                    })
                failed_count += 1
                continue

            # Check student exists
            student = Student.objects.filter(admission_no=admission_no, **school_filter).first()
            if not student:
                errors.append({
                    "row": row_num,
                    "field": "admission_no",
                    "message": f"Admission number '{admission_no}' was not found in this school."
                })
                failed_count += 1
                continue

            # Ensure uploaded admission number belongs to the selected class and section.
            selected_class_id = int(class_id)
            selected_section_id = int(section_id)
            if student.current_class_id != selected_class_id or student.current_section_id != selected_section_id:
                errors.append({
                    "row": row_num,
                    "field": "admission_no",
                    "message": (
                        f"Admission number '{admission_no}' does not belong to the selected "
                        "class/section."
                    ),
                })
                failed_count += 1
                continue

            try:
                # Delete existing attendance for this student on the date
                StudentAttendance.objects.filter(
                    student_id=student.id,
                    attendance_date=request_date,
                    **school_filter,
                ).delete()

                # Create new attendance record
                attendance = StudentAttendance()
                attendance.student_id = student.id
                attendance.attendance_date = request_date
                attendance.attendance_type = attendance_type or "P"
                attendance.notes = note
                attendance.arrival_time = None if arrival_time == "invalid" else arrival_time
                attendance.sign_in_time = None if sign_in_time == "invalid" else sign_in_time
                attendance.sign_out_time = None if sign_out_time == "invalid" else sign_out_time
                attendance.pickup_time = None if pickup_time == "invalid" else pickup_time
                attendance.pickup_by = pickup_by
                attendance.lunch = lunch_value
                attendance.school = request.user.school
                attendance.academic_year_id = request.data.get("academic_year_id")
                attendance.class_id = class_id
                attendance.section_id = section_id
                attendance.save()

                processed_count += 1
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "field": "system",
                    "message": f"Failed to save record: {str(e)[:100]}"
                })
                failed_count += 1

        # Determine response based on results
        if processed_count == 0 and failed_count > 0:
            # All failed
            return Response(
                {
                    "success": False,
                    "message": f"Failed to import any records. {failed_count} errors found.",
                    "data": {
                        "imported": 0,
                        "failed": failed_count,
                        "errors": errors[:50]  # Limit to first 50 errors
                    }
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif failed_count > 0:
            # Partial success
            return Response(
                {
                    "success": True,
                    "message": f"{processed_count} records imported, {failed_count} failed",
                    "data": {
                        "imported": processed_count,
                        "failed": failed_count,
                        "errors": errors[:50]  # Limit to first 50 errors
                    }
                },
                status=status.HTTP_200_OK,
            )
        else:
            # Full success
            return Response(
                {
                    "success": True,
                    "message": f"Successfully imported {processed_count} attendance records",
                    "data": {
                        "imported": processed_count,
                        "failed": 0,
                        "errors": []
                    }
                },
                status=status.HTTP_200_OK,
            )


class ClassAttendanceSummaryAPIView(AttendanceTenantMixin, APIView):
    """Returns per-class attendance breakdown for a given date."""

    def get(self, request):
        from datetime import date as date_type
        date_str = request.query_params.get("date") or str(date_type.today())
        try:
            parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        school_filter = self.school_filter(request)

        # Per-class student totals — query directly from Student model using current_class FK
        student_counts = (
            Student.objects.filter(is_active=True, **school_filter)
            .exclude(current_class=None)
            .values("current_class_id")
            .annotate(count=Count("id"))
        )
        total_by_class: dict[int, int] = {row["current_class_id"]: row["count"] for row in student_counts}

        # Per-class attendance aggregates for the date — only count students
        # who are still active (so deactivated students don't push counts > total).
        # `signed_in` counts present rows that actually have a sign-in time
        # so the percentage reflects who really arrived rather than who was
        # only marked Present in advance.
        active_student_ids = Student.objects.filter(is_active=True, **school_filter).values_list("id", flat=True)
        rows = (
            StudentAttendance.objects
            .filter(attendance_date=parsed_date, student_id__in=list(active_student_ids), **school_filter)
            .values("class_id")
            .annotate(
                present=Count(Case(When(attendance_type="P", then=1), output_field=IntegerField())),
                signed_in=Count(Case(
                    When(attendance_type="P", sign_in_time__isnull=False, then=1),
                    output_field=IntegerField(),
                )),
                absent=Count(Case(When(attendance_type="A", then=1), output_field=IntegerField())),
                late=Count(Case(When(attendance_type="L", then=1), output_field=IntegerField())),
            )
        )

        counts_by_class: dict[int, dict] = {}
        for row in rows:
            cid = row["class_id"]
            if cid is not None:
                counts_by_class[cid] = {
                    "present": row["present"] or 0,
                    "signed_in": row["signed_in"] or 0,
                    "absent": row["absent"] or 0,
                    "late": row["late"] or 0,
                }

        # Build result using union of both sets so all classes are represented
        all_class_ids = set(total_by_class.keys()) | set(counts_by_class.keys())
        result = []
        for cls_id in all_class_ids:
            total = total_by_class.get(cls_id, 0)
            c = counts_by_class.get(cls_id, {"present": 0, "signed_in": 0, "absent": 0, "late": 0})
            # Clamp to total to be defensive against stray duplicates.
            present = min(c["present"], total) if total else c["present"]
            signed_in = min(c["signed_in"], total) if total else c["signed_in"]
            absent = min(c["absent"], total) if total else c["absent"]
            late = min(c["late"], total) if total else c["late"]
            marked = present + absent + late
            attended = signed_in + late
            result.append({
                "class_id": cls_id,
                "total": total,
                "present": present,
                "signed_in": signed_in,
                "absent": absent,
                "late": late,
                "unmarked": max(0, total - marked),
                # pct now reflects students who actually arrived (signed in or late),
                # so a class marked Present in advance without sign-ins shows 0%.
                "pct": round((attended / total) * 100) if total > 0 else 0,
            })

        return Response({"date": date_str, "classes": result})


class StudentAttendanceDailySummaryAPIView(AttendanceTenantMixin, APIView):
    """Returns aggregated attendance counts for a given date (school-wide)."""

    def get(self, request):
        from datetime import date as date_type
        date_str = request.query_params.get("date") or str(date_type.today())
        try:
            parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        school_filter = self.school_filter(request)

        total = Student.objects.filter(is_active=True, **school_filter).count()

        daily_queryset = StudentAttendance.objects.filter(
            attendance_date=parsed_date,
            **school_filter,
        )
        latest_ids = daily_queryset.values("student_id").annotate(latest_id=Max("id")).values("latest_id")

        latest_rows = StudentAttendance.objects.filter(id__in=latest_ids)

        counts = latest_rows.aggregate(
            present=Count(Case(When(attendance_type="P", then=1), output_field=IntegerField())),
            absent=Count(Case(When(attendance_type="A", then=1), output_field=IntegerField())),
            late=Count(Case(When(attendance_type="L", then=1), output_field=IntegerField())),
        )
        absent_with_reason = latest_rows.filter(attendance_type="A").exclude(notes__isnull=True).exclude(notes="").count()

        present = counts.get("present") or 0
        absent = counts.get("absent") or 0
        late = counts.get("late") or 0
        marked = present + absent + late
        unmarked = max(0, total - marked)

        return Response({
            "date": date_str,
            "total_students": total,
            "present": present,
            "absent": absent,
            "absent_with_reason": absent_with_reason,
            "late": late,
            "unmarked": unmarked,
        })


class StudentAttendanceExportAPIView(AttendanceTenantMixin, APIView):
    """Exports attendance rows in CSV/XLSX format.

    XLSX output follows the Attendance Report spec: two sheets (Details + Summary),
    styled header block, color-coded rows, autofilter, freeze panes, group separators,
    and conditional formatting on the summary sheet.
    """

    def get(self, request):
        format_type = (request.query_params.get("format") or "csv").lower()
        class_id = request.query_params.get("class_id")
        section_id = request.query_params.get("section_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        date_value = request.query_params.get("date")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        queryset = StudentAttendance.objects.select_related("student").filter(**self.school_filter(request))
        if class_id:
            queryset = queryset.filter(class_id=class_id)
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        if month:
            queryset = queryset.filter(attendance_date__month=month)
        if year:
            queryset = queryset.filter(attendance_date__year=year)
        if date_value:
            queryset = queryset.filter(attendance_date=date_value)
        if date_from:
            queryset = queryset.filter(attendance_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(attendance_date__lte=date_to)

        if format_type == "xlsx":
            return self._export_xlsx(
                request, queryset,
                class_id=class_id, section_id=section_id,
                month=month, year=year,
                date_value=date_value, date_from=date_from, date_to=date_to,
            )

        # ---- CSV (legacy) ----
        rows = [
            [
                att.student.admission_no,
                f"{att.student.first_name} {att.student.last_name}".strip(),
                att.student.roll_no,
                str(att.attendance_date),
                att.attendance_type,
                att.notes,
                att.arrival_time.strftime("%H:%M") if att.arrival_time else "",
                att.sign_in_time.strftime("%H:%M") if att.sign_in_time else "",
                att.sign_out_time.strftime("%H:%M") if att.sign_out_time else "",
                att.pickup_time.strftime("%H:%M") if att.pickup_time else "",
                att.pickup_by,
                "Yes" if att.lunch else "No",
            ]
            for att in queryset.order_by("attendance_date", "student_id")
        ]
        headers = [
            "admission_no", "student_name", "roll_no", "attendance_date",
            "attendance_type", "note", "arrival_time", "sign_in_time",
            "sign_out_time", "pickup_time", "pickup_by", "lunch",
        ]
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="student_attendance_export.csv"'
        return response

    # ------------------------------------------------------------------
    # XLSX builder (styled, two-sheet, per spec)
    # ------------------------------------------------------------------
    def _export_xlsx(self, request, queryset, *, class_id, section_id,
                     month, year, date_value, date_from, date_to):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import ColorScaleRule
        except Exception:
            return Response(
                {"detail": "openpyxl is required for xlsx export."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # ---- Resolve names ----
        rows_qs = list(queryset.order_by("attendance_date", "class_id", "section_id", "student_id"))

        class_ids = {r.class_id for r in rows_qs if r.class_id}
        section_ids = {r.section_id for r in rows_qs if r.section_id}
        classes_map = {c.id: c.name for c in Class.objects.filter(id__in=class_ids)}
        sections_map = {s.id: s.name for s in Section.objects.filter(id__in=section_ids)}

        # School name
        school_name = "School"
        try:
            from apps.tenancy.models import School
            sid = getattr(request.user, "school_id", None)
            if sid:
                sch = School.objects.filter(id=sid).first()
                if sch and getattr(sch, "name", None):
                    school_name = sch.name
        except Exception:
            pass

        # ---- Status helpers ----
        STATUS_MAP = {"P": "Present", "A": "Absent", "L": "Late", "F": "Half Day", "H": "Holiday"}

        def status_label(att):
            base = STATUS_MAP.get(att.attendance_type, att.attendance_type or "")
            # Derive "Early Sign-Out" for present rows where stay < 60 min
            if base == "Present" and att.sign_in_time and att.sign_out_time:
                try:
                    base_dt = datetime.combine(att.attendance_date, att.sign_in_time)
                    out_dt = datetime.combine(att.attendance_date, att.sign_out_time)
                    if (out_dt - base_dt).total_seconds() < 3600:
                        return "Early Sign-Out"
                except Exception:
                    pass
            return base

        # ---- Styles ----
        navy = "1A237E"
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        subtitle_font = Font(name="Calibri", size=12, bold=True)
        meta_font = Font(name="Calibri", size=10, italic=True, color="424242")
        navy_fill = PatternFill("solid", fgColor=navy)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CFD8DC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        STATUS_STYLE = {
            "Present":        {"bg": "E8F5E9", "fg": "1B5E20"},
            "Absent":         {"bg": "FFEBEE", "fg": "B71C1C"},
            "Late":           {"bg": "FFF8E1", "fg": "E65100"},
            "Early Sign-Out": {"bg": "E3F2FD", "fg": "0D47A1"},
            "Half Day":       {"bg": "F3E5F5", "fg": "4A148C"},
            "Holiday":        {"bg": "ECEFF1", "fg": "37474F"},
        }
        group_fill = PatternFill("solid", fgColor="ECEFF1")
        group_font = Font(name="Calibri", size=10, bold=True, color="263238")

        # ---- Determine report metadata ----
        def fmt_month(d):
            return d.strftime("%B %Y") if d else ""

        if date_from and date_to:
            try:
                df = datetime.strptime(date_from, "%Y-%m-%d").date()
                dt = datetime.strptime(date_to, "%Y-%m-%d").date()
                if df.month == dt.month and df.year == dt.year:
                    title_period = fmt_month(df)
                else:
                    title_period = f"{df.strftime('%d %b %Y')} – {dt.strftime('%d %b %Y')}"
                file_period = f"{df.strftime('%d%b')}_{dt.strftime('%d%b_%Y')}"
            except Exception:
                title_period = f"{date_from} – {date_to}"
                file_period = f"{date_from}_{date_to}"
        elif month and year:
            try:
                d = date(int(year), int(month), 1)
                title_period = fmt_month(d)
                file_period = d.strftime("%B_%Y")
            except Exception:
                title_period = f"{month}/{year}"
                file_period = f"{month}_{year}"
        elif date_value:
            try:
                d = datetime.strptime(date_value, "%Y-%m-%d").date()
                title_period = d.strftime("%d %B %Y")
                file_period = d.strftime("%d%b_%Y")
            except Exception:
                title_period = date_value
                file_period = date_value
        elif rows_qs:
            d = rows_qs[0].attendance_date
            title_period = fmt_month(d)
            file_period = d.strftime("%B_%Y")
        else:
            today = date.today()
            title_period = fmt_month(today)
            file_period = today.strftime("%B_%Y")

        # Filter description
        if class_id and section_id:
            cls_name = classes_map.get(int(class_id), f"[ID: {class_id}]") if str(class_id).isdigit() else class_id
            sec_name = sections_map.get(int(section_id), f"[ID: {section_id}]") if str(section_id).isdigit() else section_id
            filter_label = f"Filter: Class: {cls_name} — Section: {sec_name}"
            file_filter = f"{cls_name}_{sec_name}".replace(" ", "")
        elif class_id:
            cls_name = classes_map.get(int(class_id), f"[ID: {class_id}]") if str(class_id).isdigit() else class_id
            filter_label = f"Filter: Class: {cls_name}"
            file_filter = str(cls_name).replace(" ", "")
        elif section_id:
            sec_name = sections_map.get(int(section_id), f"[ID: {section_id}]") if str(section_id).isdigit() else section_id
            filter_label = f"Filter: Section: {sec_name}"
            file_filter = f"Section{sec_name}".replace(" ", "")
        else:
            filter_label = "Filter: All Classes"
            file_filter = "All_Classes"

        # Academic year — derive from earliest year in data (Apr–Mar). Fall back to current year.
        if rows_qs:
            sample_year = rows_qs[0].attendance_date.year
            sample_month = rows_qs[0].attendance_date.month
        else:
            sample_year = date.today().year
            sample_month = date.today().month
        ay_start = sample_year if sample_month >= 4 else sample_year - 1
        academic_year = f"{ay_start}–{ay_start + 1}"

        # ---- Build workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Details"

        columns = [
            ("S.No", 6),
            ("Date", 14),
            ("Day", 12),
            ("Academic Year", 14),
            ("Month", 12),
            ("Class", 16),
            ("Section", 14),
            ("Student ID", 12),
            ("Student Name", 24),
            ("Status", 16),
            ("Sign-In Time", 13),
            ("Sign-Out Time", 13),
            ("Remarks / Reason", 32),
        ]
        last_col = len(columns)
        last_col_letter = get_column_letter(last_col)

        # ---- Header block (rows 1-4) ----
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(row=1, column=1, value=school_name)
        c.font = title_font
        c.fill = navy_fill
        c.alignment = center
        ws.row_dimensions[1].height = 32

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        c = ws.cell(row=2, column=1, value=f"Attendance Report — {title_period}")
        c.font = subtitle_font
        c.alignment = center
        ws.row_dimensions[2].height = 22

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        c = ws.cell(row=3, column=1, value=f"Academic Year: {academic_year}")
        c.font = meta_font
        c.alignment = center

        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
        c = ws.cell(row=4, column=1, value=filter_label)
        c.font = meta_font
        c.alignment = center

        # Row 5 blank separator
        header_row_idx = 6

        # Column widths
        for i, (_, w) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(w, 10)

        # ---- Column header row ----
        for i, (h, _) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row_idx, column=i, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[header_row_idx].height = 28

        # ---- Build data rows, grouped by class/section when multi-class ----
        # Pre-sort: Date desc → Class → Section → Student Name
        def sort_key(att):
            name = (
                f"{att.student.first_name} {att.student.last_name}".strip().lower()
                if att.student else ""
            )
            return (
                -att.attendance_date.toordinal(),
                classes_map.get(att.class_id or 0, ""),
                sections_map.get(att.section_id or 0, ""),
                name,
            )
        rows_sorted = sorted(rows_qs, key=sort_key)

        unique_classes = {r.class_id for r in rows_sorted}
        multi_class = len(unique_classes) > 1

        current_row = header_row_idx + 1
        sno = 0
        last_group = None
        # Group buckets for separator headers
        from itertools import groupby
        def group_keyer(att):
            return (att.class_id, att.section_id)

        for group_key, group_iter in groupby(rows_sorted, key=group_keyer):
            group_list = list(group_iter)
            if multi_class:
                cls_n = classes_map.get(group_key[0] or 0, f"[ID: {group_key[0]}]")
                sec_n = sections_map.get(group_key[1] or 0, f"[ID: {group_key[1]}]")
                pres = sum(1 for a in group_list if a.attendance_type == "P")
                absn = sum(1 for a in group_list if a.attendance_type == "A")
                latn = sum(1 for a in group_list if a.attendance_type == "L")
                tot = len({a.student_id for a in group_list})
                group_label = (
                    f"Class: {cls_n} — Section: {sec_n}  |  Total Students: {tot}  |  "
                    f"Present: {pres}  |  Absent: {absn}  |  Late: {latn}"
                )
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
                gc = ws.cell(row=current_row, column=1, value=group_label)
                gc.fill = group_fill
                gc.font = group_font
                gc.alignment = left
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            for att in group_list:
                sno += 1
                full_name = (
                    f"{att.student.first_name} {att.student.last_name}".strip()
                    if att.student else f"[ID: {att.student_id}]"
                )
                if not full_name:
                    full_name = f"[ID: {att.student_id}]"
                status_text = status_label(att)
                style = STATUS_STYLE.get(status_text)

                cls_n = classes_map.get(att.class_id or 0, f"[ID: {att.class_id}]" if att.class_id else "")
                sec_n = sections_map.get(att.section_id or 0, f"[ID: {att.section_id}]" if att.section_id else "")

                values = [
                    sno,
                    att.attendance_date,
                    att.attendance_date.strftime("%A"),
                    academic_year,
                    att.attendance_date.strftime("%B"),
                    cls_n,
                    sec_n,
                    att.student.admission_no if att.student else att.student_id,
                    full_name,
                    status_text,
                    att.sign_in_time.strftime("%H:%M") if att.sign_in_time else "",
                    att.sign_out_time.strftime("%H:%M") if att.sign_out_time else "",
                    att.notes or "",
                ]
                for i, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=i, value=val)
                    cell.border = border
                    if i == 2:  # Date column
                        cell.number_format = "DD MMM YYYY"
                        cell.alignment = center
                    elif i in (1, 3, 4, 5, 7, 10, 11, 12):
                        cell.alignment = center
                    elif i == 8:
                        cell.alignment = right
                    else:
                        cell.alignment = left
                    if style:
                        cell.fill = PatternFill("solid", fgColor=style["bg"])
                        cell.font = Font(name="Calibri", size=10, color=style["fg"])
                current_row += 1

        # AutoFilter + freeze panes
        if current_row > header_row_idx + 1:
            ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{current_row - 1}"
        ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=3)  # freeze header + S.No, Date

        # ---- Sheet 2: Summary ----
        ws2 = wb.create_sheet("Summary")
        sum_cols = [
            ("Class", 18), ("Section", 14), ("Total Students", 14),
            ("Present", 10), ("Absent", 10), ("Late", 10),
            ("Early Sign-Out", 14), ("Attendance %", 14),
        ]
        for i, (_, w) in enumerate(sum_cols, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # Title row
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sum_cols))
        tc = ws2.cell(row=1, column=1, value=f"{school_name} — Attendance Summary ({title_period})")
        tc.font = title_font
        tc.fill = navy_fill
        tc.alignment = center
        ws2.row_dimensions[1].height = 30

        # Header row
        for i, (h, _) in enumerate(sum_cols, start=1):
            cell = ws2.cell(row=3, column=i, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = border
        ws2.row_dimensions[3].height = 26

        # Aggregate
        from collections import defaultdict as _dd
        bucket = _dd(lambda: {"students": set(), "P": 0, "A": 0, "L": 0, "ESO": 0})
        for att in rows_sorted:
            key = (att.class_id, att.section_id)
            b = bucket[key]
            b["students"].add(att.student_id)
            stext = status_label(att)
            if stext == "Early Sign-Out":
                b["ESO"] += 1
                b["P"] += 1  # still counted as present for attendance %
            elif att.attendance_type == "P":
                b["P"] += 1
            elif att.attendance_type == "A":
                b["A"] += 1
            elif att.attendance_type == "L":
                b["L"] += 1

        sum_row = 4
        sorted_keys = sorted(
            bucket.keys(),
            key=lambda k: (classes_map.get(k[0] or 0, ""), sections_map.get(k[1] or 0, "")),
        )
        tot_students = tot_p = tot_a = tot_l = tot_eso = 0
        for k in sorted_keys:
            b = bucket[k]
            cls_n = classes_map.get(k[0] or 0, f"[ID: {k[0]}]" if k[0] else "—")
            sec_n = sections_map.get(k[1] or 0, f"[ID: {k[1]}]" if k[1] else "—")
            total = len(b["students"])
            pct = round(((b["P"] + b["L"]) / total) * 100, 1) if total else 0.0
            row_vals = [cls_n, sec_n, total, b["P"], b["A"], b["L"], b["ESO"], pct / 100.0]
            for i, v in enumerate(row_vals, start=1):
                cell = ws2.cell(row=sum_row, column=i, value=v)
                cell.border = border
                if i == 8:
                    cell.number_format = "0.0%"
                    cell.alignment = right
                elif i in (3, 4, 5, 6, 7):
                    cell.alignment = right
                else:
                    cell.alignment = left
            tot_students += total
            tot_p += b["P"]; tot_a += b["A"]; tot_l += b["L"]; tot_eso += b["ESO"]
            sum_row += 1

        # Totals row
        tot_pct = round(((tot_p + tot_l) / tot_students) * 100, 1) if tot_students else 0.0
        totals = ["TOTAL", "", tot_students, tot_p, tot_a, tot_l, tot_eso, tot_pct / 100.0]
        for i, v in enumerate(totals, start=1):
            cell = ws2.cell(row=sum_row, column=i, value=v)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.border = border
            if i == 8:
                cell.number_format = "0.0%"
                cell.alignment = right
            elif i in (3, 4, 5, 6, 7):
                cell.alignment = right
            else:
                cell.alignment = left

        # Conditional 3-color scale on Attendance % column (H4..H{sum_row-1})
        if sum_row > 4:
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="FFCDD2",
                mid_type="num", mid_value=0.85, mid_color="FFF9C4",
                end_type="num", end_value=1, end_color="C8E6C9",
            )
            ws2.conditional_formatting.add(f"H4:H{sum_row - 1}", rule)
            ws2.auto_filter.ref = f"A3:H{sum_row - 1}"
        ws2.freeze_panes = "A4"

        # ---- Filename ----
        filename = f"Attendance_{file_filter}_{file_period}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class StudentAttendanceReportInsightsAPIView(AttendanceTenantMixin, APIView):
    """Weekly and reason insights used by richer attendance reports."""

    def get(self, request):
        class_id = request.query_params.get("class_id")
        section_id = request.query_params.get("section_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        if not month or not year:
            return Response({"detail": "month and year are required."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = StudentAttendance.objects.filter(
            attendance_date__month=month,
            attendance_date__year=year,
            **self.school_filter(request),
        )
        if class_id:
            queryset = queryset.filter(class_id=class_id)
        if section_id:
            queryset = queryset.filter(section_id=section_id)

        weekly = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0})
        for att in queryset:
            week = ((att.attendance_date.day - 1) // 7) + 1
            if att.attendance_type == "P":
                weekly[week]["present"] += 1
            elif att.attendance_type == "A":
                weekly[week]["absent"] += 1
            elif att.attendance_type == "L":
                weekly[week]["late"] += 1

        reason_rows = queryset.filter(attendance_type="A").exclude(notes="").values("notes").annotate(count=Count("id")).order_by("-count")[:10]
        late_rows = queryset.filter(attendance_type="L").exclude(notes="").values("notes").annotate(count=Count("id")).order_by("-count")[:10]

        weekly_result = []
        for week in sorted(weekly.keys()):
            data = weekly[week]
            total = data["present"] + data["absent"] + data["late"]
            weekly_result.append(
                {
                    "week": week,
                    "present": data["present"],
                    "absent": data["absent"],
                    "late": data["late"],
                    "present_pct": round((data["present"] / total) * 100, 2) if total else 0,
                }
            )

        return Response(
            {
                "weekly": weekly_result,
                "top_absent_reasons": [{"reason": r["notes"], "count": r["count"]} for r in reason_rows],
                "top_late_reasons": [{"reason": r["notes"], "count": r["count"]} for r in late_rows],
            }
        )
