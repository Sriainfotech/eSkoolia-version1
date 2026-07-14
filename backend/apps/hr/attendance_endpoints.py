from io import BytesIO, StringIO
import csv
from datetime import date, datetime
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import BaseRenderer
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.exceptions import PermissionDenied
from .models import Staff, StaffAttendance, Department


class StaffAttendancePermissionMixin:
    """Authenticated + `human_resource.staff_attendance.view` permission code.

    Mirrors the gate the staff-attendance ViewSet (SchoolScopedModelViewSet) and
    the student attendance views enforce, so these file-export/import endpoints
    can't be reached by an authenticated user who lacks HR attendance access.
    """

    permission_classes = [permissions.IsAuthenticated]
    required_permission_code = "human_resource.staff_attendance.view"

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        user = request.user
        if user.is_superuser:
            return
        code = self.required_permission_code
        if not code:
            return
        if not hasattr(user, "has_permission_code") or not user.has_permission_code(code):
            raise PermissionDenied("You do not have permission to perform this action.")


class StaffAttendanceDownloadSampleAPIView(StaffAttendancePermissionMixin, APIView):
    def get(self, request):
        try:
            from openpyxl import Workbook
        except Exception:
            return Response({"detail": "openpyxl is required for sample export."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "staff_attendance_sheet"
        headers = [
            "staff_no",
            "attendance_date",
            "attendance_type",
            "note",
            "arrival_time",
            "sign_in_time",
            "sign_out_time",
            "lunch",
        ]
        sheet.append(headers)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="staff_attendance_sheet.xlsx"'
        return response

class StaffAttendanceExportAPIView(StaffAttendancePermissionMixin, APIView):
    """Exports staff attendance in CSV / XLSX format.

    Mirrors the student attendance export: a styled workbook with a Details
    sheet (grouped by department when more than one is present) and a Summary
    sheet, plus a plain CSV fallback.
    """

    # Force a single passthrough renderer + ignore the `format` URL kwarg/query
    # param. Without this, a `?format=excel` (or `?format=xlsx`) from the client
    # makes DRF's content negotiation look for a renderer whose format is
    # "excel"/"xlsx", find none, and raise Http404 — that was the 404 users hit.
    class _PassthroughRenderer(BaseRenderer):
        media_type = "*/*"
        format = "*"
        charset = None
        render_style = "binary"

        def render(self, data, accepted_media_type=None, renderer_context=None):
            return data

    renderer_classes = [_PassthroughRenderer]

    def perform_content_negotiation(self, request, force=False):
        renderer = self.renderer_classes[0]()
        return (renderer, renderer.media_type)

    def _school_filter(self, request):
        return {"school_id": request.user.school_id}

    def get(self, request):
        params = request.query_params
        # Accept both `fmt` (preferred, matches student export) and legacy
        # `format`. "excel"/"xlsx"/"xls" → styled workbook, anything else → CSV.
        fmt_raw = (params.get("fmt") or params.get("format") or "xlsx").lower()
        format_type = "xlsx" if fmt_raw in ("xlsx", "excel", "xls") else "csv"

        date_value = params.get("date")
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        month = params.get("month")
        year = params.get("year")
        department = params.get("department")
        attendance_type = params.get("attendance_type")

        qs = StaffAttendance.objects.select_related(
            "staff", "staff__department", "staff__designation"
        ).filter(**self._school_filter(request))
        if date_value:
            qs = qs.filter(attendance_date=date_value)
        if date_from:
            qs = qs.filter(attendance_date__gte=date_from)
        if date_to:
            qs = qs.filter(attendance_date__lte=date_to)
        if month:
            qs = qs.filter(attendance_date__month=month)
        if year:
            qs = qs.filter(attendance_date__year=year)
        if department and str(department).lower() != "all":
            # The HR attendance UI sends the department NAME in its <select>,
            # but a numeric id is also accepted (e.g. from API callers).
            if str(department).isdigit():
                qs = qs.filter(staff__department_id=department)
            else:
                qs = qs.filter(staff__department__name=department)
        if attendance_type and str(attendance_type).lower() != "all":
            qs = qs.filter(attendance_type=attendance_type)

        if format_type == "xlsx":
            return self._export_xlsx(
                request, qs,
                date_value=date_value, date_from=date_from, date_to=date_to,
                month=month, year=year, department=department,
            )
        return self._export_csv(qs, date_value)

    # ------------------------------------------------------------------
    # CSV (plain, legacy)
    # ------------------------------------------------------------------
    def _export_csv(self, qs, date_value):
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "staff_no", "first_name", "last_name", "department", "designation",
            "attendance_date", "attendance_type", "arrival_time",
            "sign_in_time", "sign_out_time", "lunch", "note",
        ])
        for att in qs.order_by("attendance_date", "staff_id"):
            staff = att.staff
            writer.writerow([
                staff.staff_no if staff else "",
                staff.first_name if staff else "",
                staff.last_name if staff else "",
                staff.department.name if staff and staff.department else "",
                staff.designation.name if staff and staff.designation else "",
                att.attendance_date,
                att.attendance_type,
                att.arrival_time.strftime("%H:%M") if att.arrival_time else "",
                att.sign_in_time.strftime("%H:%M") if att.sign_in_time else "",
                att.sign_out_time.strftime("%H:%M") if att.sign_out_time else "",
                "Yes" if att.lunch else "No",
                att.note or "",
            ])
        # Prepend UTF-8 BOM so Excel detects the encoding and shows non-ASCII
        # names/notes correctly instead of mojibake on Windows.
        body = "﻿" + output.getvalue()
        response = HttpResponse(body.encode("utf-8"), content_type="text/csv; charset=utf-8")
        suffix = date_value or "export"
        response["Content-Disposition"] = f'attachment; filename="staff_attendance_{suffix}.csv"'
        return response

    # ------------------------------------------------------------------
    # XLSX builder (styled, two-sheet) — modelled on the student export
    # ------------------------------------------------------------------
    def _export_xlsx(self, request, queryset, *, date_value, date_from,
                     date_to, month, year, department):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import ColorScaleRule
        except Exception:
            return Response(
                {"detail": "openpyxl is required for xlsx export."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        rows_qs = list(queryset.order_by("attendance_date", "staff_id"))

        # ---- School name ----
        school_name = "School"
        try:
            from apps.tenancy.models import School
            sid = getattr(request.user, "school_id", None)
            if sid:
                sch = School.objects.filter(id=sid).first()
                if sch and getattr(sch, "name", None):
                    school_name = sch.name
        except Exception:
            pass

        # ---- Status helpers ----
        STATUS_MAP = {"P": "Present", "A": "Absent", "L": "Leave", "F": "Half Day", "H": "Holiday"}

        def status_label(att):
            return STATUS_MAP.get(att.attendance_type, att.attendance_type or "")

        # ---- Styles ----
        navy = "1A237E"
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        subtitle_font = Font(name="Calibri", size=12, bold=True)
        meta_font = Font(name="Calibri", size=10, italic=True, color="424242")
        navy_fill = PatternFill("solid", fgColor=navy)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CFD8DC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        STATUS_STYLE = {
            "Present":  {"bg": "E8F5E9", "fg": "1B5E20"},
            "Absent":   {"bg": "FFEBEE", "fg": "B71C1C"},
            "Leave":    {"bg": "FFF8E1", "fg": "E65100"},
            "Half Day": {"bg": "F3E5F5", "fg": "4A148C"},
            "Holiday":  {"bg": "ECEFF1", "fg": "37474F"},
        }
        group_fill = PatternFill("solid", fgColor="ECEFF1")
        group_font = Font(name="Calibri", size=10, bold=True, color="263238")

        # ---- Report period metadata ----
        def fmt_month(d):
            return d.strftime("%B %Y") if d else ""

        if date_from and date_to:
            try:
                df = datetime.strptime(date_from, "%Y-%m-%d").date()
                dt = datetime.strptime(date_to, "%Y-%m-%d").date()
                if df.month == dt.month and df.year == dt.year:
                    title_period = fmt_month(df)
                else:
                    title_period = f"{df.strftime('%d %b %Y')} – {dt.strftime('%d %b %Y')}"
                file_period = f"{df.strftime('%d%b')}_{dt.strftime('%d%b_%Y')}"
            except Exception:
                title_period = f"{date_from} – {date_to}"
                file_period = f"{date_from}_{date_to}"
        elif month and year:
            try:
                d = date(int(year), int(month), 1)
                title_period = fmt_month(d)
                file_period = d.strftime("%B_%Y")
            except Exception:
                title_period = f"{month}/{year}"
                file_period = f"{month}_{year}"
        elif date_value:
            try:
                d = datetime.strptime(date_value, "%Y-%m-%d").date()
                title_period = d.strftime("%d %B %Y")
                file_period = d.strftime("%d%b_%Y")
            except Exception:
                title_period = date_value
                file_period = date_value
        elif rows_qs:
            d = rows_qs[0].attendance_date
            title_period = fmt_month(d)
            file_period = d.strftime("%B_%Y")
        else:
            today = date.today()
            title_period = fmt_month(today)
            file_period = today.strftime("%B_%Y")

        # ---- Filter description ----
        dept_name = None
        if department and str(department).lower() != "all":
            if str(department).isdigit():
                # Scope the lookup to the caller's school so another tenant's
                # department name can't leak into the label/filename.
                d = Department.objects.filter(id=department, **self._school_filter(request)).first()
                dept_name = d.name if d else f"[ID: {department}]"
            else:
                dept_name = str(department)
        if dept_name:
            filter_label = f"Filter: Department: {dept_name}"
            file_filter = str(dept_name).replace(" ", "")
        else:
            filter_label = "Filter: All Departments"
            file_filter = "All_Departments"

        # ---- Build workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Details"

        columns = [
            ("S.No", 6),
            ("Date", 14),
            ("Day", 12),
            ("Department", 20),
            ("Designation", 20),
            ("Staff ID", 12),
            ("Staff Name", 24),
            ("Status", 14),
            ("Sign-In Time", 13),
            ("Sign-Out Time", 13),
            ("Lunch", 9),
            ("Remarks / Reason", 32),
        ]
        last_col = len(columns)
        last_col_letter = get_column_letter(last_col)

        # ---- Header block (rows 1-4) ----
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(row=1, column=1, value=school_name)
        c.font = title_font
        c.fill = navy_fill
        c.alignment = center
        ws.row_dimensions[1].height = 32

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        c = ws.cell(row=2, column=1, value=f"Staff Attendance Report — {title_period}")
        c.font = subtitle_font
        c.alignment = center
        ws.row_dimensions[2].height = 22

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        c = ws.cell(row=3, column=1, value=f"Generated: {timezone.localtime().strftime('%d %b %Y %H:%M')}")
        c.font = meta_font
        c.alignment = center

        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
        c = ws.cell(row=4, column=1, value=filter_label)
        c.font = meta_font
        c.alignment = center

        header_row_idx = 6
        for i, (_, w) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(w, 9)

        for i, (h, _) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row_idx, column=i, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[header_row_idx].height = 28

        # ---- Sort: Department → Staff name → Date ascending ----
        def dept_of(att):
            return att.staff.department.name if att.staff and att.staff.department else "—"

        def name_of(att):
            if not att.staff:
                return ""
            return f"{att.staff.first_name} {att.staff.last_name}".strip()

        def sort_key(att):
            return (dept_of(att).lower(), name_of(att).lower(), att.attendance_date.toordinal())
        rows_sorted = sorted(rows_qs, key=sort_key)

        multi_dept = len({dept_of(a) for a in rows_sorted}) > 1

        from itertools import groupby
        current_row = header_row_idx + 1
        sno = 0

        for dept_key, group_iter in groupby(rows_sorted, key=dept_of):
            group_list = list(group_iter)
            if multi_dept:
                distinct_staff = len({a.staff_id for a in group_list})
                days = len({a.attendance_date for a in group_list})
                pres = sum(1 for a in group_list if a.attendance_type == "P")
                absn = sum(1 for a in group_list if a.attendance_type == "A")
                leav = sum(1 for a in group_list if a.attendance_type == "L")
                group_label = (
                    f"  {dept_key}  |  {distinct_staff} staff  |  {days} day(s)  |  "
                    f"Present: {pres}  Absent: {absn}  Leave: {leav}"
                )
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
                gc = ws.cell(row=current_row, column=1, value=group_label)
                gc.fill = group_fill
                gc.font = group_font
                gc.alignment = left
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            for att in group_list:
                sno += 1
                staff = att.staff
                full_name = name_of(att) or (f"[ID: {att.staff_id}]")
                status_text = status_label(att)
                style = STATUS_STYLE.get(status_text)
                desig = staff.designation.name if staff and staff.designation else ""
                dept_n = dept_of(att)

                values = [
                    sno,
                    att.attendance_date,
                    att.attendance_date.strftime("%A"),
                    dept_n,
                    desig,
                    staff.staff_no if staff else att.staff_id,
                    full_name,
                    status_text,
                    att.sign_in_time.strftime("%H:%M") if att.sign_in_time else "",
                    att.sign_out_time.strftime("%H:%M") if att.sign_out_time else "",
                    "Yes" if att.lunch else "No",
                    att.note or "",
                ]
                for i, val in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=i, value=val)
                    cell.border = border
                    if i == 2:  # Date
                        cell.number_format = "DD MMM YYYY"
                        cell.alignment = center
                    elif i in (1, 3, 8, 9, 10, 11):
                        cell.alignment = center
                    elif i == 6:
                        cell.alignment = right
                    else:
                        cell.alignment = left
                    if style:
                        cell.fill = PatternFill("solid", fgColor=style["bg"])
                        cell.font = Font(name="Calibri", size=10, color=style["fg"])
                current_row += 1

        if current_row > header_row_idx + 1:
            ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{current_row - 1}"
        ws.freeze_panes = f"A{header_row_idx + 1}"

        # ---- Sheet 2: Summary (by department) ----
        ws2 = wb.create_sheet("Summary")
        sum_cols = [
            ("Department", 22), ("Total Staff", 12), ("Present", 10),
            ("Absent", 10), ("Leave", 10), ("Half Day", 10),
            ("Holiday", 10), ("Attendance %", 14),
        ]
        for i, (_, w) in enumerate(sum_cols, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sum_cols))
        tc = ws2.cell(row=1, column=1, value=f"{school_name} — Staff Attendance Summary ({title_period})")
        tc.font = title_font
        tc.fill = navy_fill
        tc.alignment = center
        ws2.row_dimensions[1].height = 30

        for i, (h, _) in enumerate(sum_cols, start=1):
            cell = ws2.cell(row=3, column=i, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = border
        ws2.row_dimensions[3].height = 26

        from collections import defaultdict as _dd
        bucket = _dd(lambda: {"staff": set(), "P": 0, "A": 0, "L": 0, "F": 0, "H": 0})
        for att in rows_sorted:
            b = bucket[dept_of(att)]
            b["staff"].add(att.staff_id)
            if att.attendance_type in b:
                b[att.attendance_type] += 1

        sum_row = 4
        tot_staff = tot_p = tot_a = tot_l = tot_f = tot_h = 0
        for dept_n in sorted(bucket.keys(), key=lambda k: k.lower()):
            b = bucket[dept_n]
            total = len(b["staff"])
            # Working days = Present + Absent + Half Day. Holiday and approved
            # Leave are excused, so they're excluded from the denominator (a
            # staff member on leave is not counted as "absent"). Half-day counts
            # as half present.
            worked = b["P"] + b["A"] + b["F"]
            pct = ((b["P"] + 0.5 * b["F"]) / worked) if worked else 0.0
            row_vals = [dept_n, total, b["P"], b["A"], b["L"], b["F"], b["H"], pct]
            for i, v in enumerate(row_vals, start=1):
                cell = ws2.cell(row=sum_row, column=i, value=v)
                cell.border = border
                if i == 8:
                    cell.number_format = "0.0%"
                    cell.alignment = right
                elif i in (2, 3, 4, 5, 6, 7):
                    cell.alignment = right
                else:
                    cell.alignment = left
            tot_staff += total
            tot_p += b["P"]; tot_a += b["A"]; tot_l += b["L"]; tot_f += b["F"]; tot_h += b["H"]
            sum_row += 1

        tot_worked = tot_p + tot_a + tot_f
        tot_pct = ((tot_p + 0.5 * tot_f) / tot_worked) if tot_worked else 0.0
        totals = ["TOTAL", tot_staff, tot_p, tot_a, tot_l, tot_f, tot_h, tot_pct]
        for i, v in enumerate(totals, start=1):
            cell = ws2.cell(row=sum_row, column=i, value=v)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.border = border
            if i == 8:
                cell.number_format = "0.0%"
                cell.alignment = right
            elif i in (2, 3, 4, 5, 6, 7):
                cell.alignment = right
            else:
                cell.alignment = left

        if sum_row > 4:
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="FFCDD2",
                mid_type="num", mid_value=0.85, mid_color="FFF9C4",
                end_type="num", end_value=1, end_color="C8E6C9",
            )
            ws2.conditional_formatting.add(f"H4:H{sum_row - 1}", rule)
            ws2.auto_filter.ref = f"A3:H{sum_row - 1}"
        ws2.freeze_panes = "A4"

        filename = f"Staff_Attendance_{file_filter}_{file_period}.xlsx"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

class StaffAttendanceImportBulkAPIView(StaffAttendancePermissionMixin, APIView):
    parser_classes = [MultiPartParser, FormParser]

    def _normalize_date(self, value):
        if value is None or value == "":
            return None
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            try:
                from openpyxl.utils.datetime import from_excel
                converted = from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def post(self, request):
        attendance_date = request.data.get("attendance_date")
        uploaded_file = request.FILES.get("file")

        if not attendance_date or not uploaded_file:
            return Response(
                {"detail": "attendance_date and file are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        request_date = self._normalize_date(attendance_date)
        if not request_date:
            return Response({"detail": "Invalid attendance_date."}, status=status.HTTP_400_BAD_REQUEST)

        school = request.user.school or getattr(request, "school", None)
        
        imported_rows = []
        ext = uploaded_file.name.split(".")[-1].lower() if "." in uploaded_file.name else ""
        if ext == "csv":
            try:
                content = uploaded_file.read().decode("utf-8", errors="ignore")
                from io import StringIO
                reader = csv.DictReader(StringIO(content))
                for row in reader:
                    imported_rows.append(row)
            except Exception as e:
                return Response({"detail": f"Failed to parse CSV: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded_file, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                if rows:
                    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                    for row in rows[1:]:
                        row_dict = {}
                        for i, h in enumerate(headers):
                            if i < len(row):
                                row_dict[h] = row[i]
                        imported_rows.append(row_dict)
            except Exception as e:
                return Response({"detail": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        valid_types = ["P", "A", "L", "F", "H"]
        processed = 0
        failed = 0
        errors = []

        for row_num, row in enumerate(imported_rows, start=2):
            staff_no = str(row.get("staff_no") or "").strip()
            if not staff_no:
                errors.append({"row": row_num, "message": "staff_no is required"})
                failed += 1
                continue

            staff = Staff.objects.filter(staff_no=staff_no, school=school).first()
            if not staff:
                errors.append({"row": row_num, "message": f"Staff with staff_no {staff_no} not found"})
                failed += 1
                continue

            att_type = str(row.get("attendance_type") or "").strip()
            if att_type and att_type not in valid_types:
                errors.append({"row": row_num, "message": f"Invalid attendance_type {att_type}"})
                failed += 1
                continue
            
            note = str(row.get("note") or "").strip()
            lunch = str(row.get("lunch") or "").strip().lower() in ["1", "true", "yes", "y", "t"]
            
            try:
                StaffAttendance.objects.update_or_create(
                    school=school,
                    staff=staff,
                    attendance_date=request_date,
                    defaults={
                        "attendance_type": att_type or "P",
                        "note": note,
                        "lunch": lunch
                    }
                )
                processed += 1
            except Exception as e:
                errors.append({"row": row_num, "message": str(e)})
                failed += 1

        return Response({
            "success": failed == 0,
            "message": f"Imported {processed}, failed {failed}",
            "data": {"imported": processed, "failed": failed, "errors": errors[:50]}
        })
