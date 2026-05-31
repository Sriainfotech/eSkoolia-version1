"""Super Admin Console API views.

All endpoints are public-schema only and protected by IsSuperAdmin.
Tenant users must receive 403 responses.
"""

from __future__ import annotations

import csv
import io
import json
import os
import secrets
import string
from collections import OrderedDict
from datetime import timedelta
from decimal import Decimal
from uuid import uuid4

import yaml
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from django.db.models import Count, IntegerField, OuterRef, Q, Subquery, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import parsers, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from django_tenants.utils import schema_context

from apps.access_control.permission_classes import IsSuperAdmin
from apps.hr.models import Staff
from apps.students.models import Student
from apps.tenancy.audit import log_audit
from apps.tenancy.models import (
    Domain,
    School,
    SchoolTenant,
    SubscriptionPlan,
    SuperAdminInvoice,
    SuperAdminInvoicePayment,
    SuperAdminPolicy,
    TenantAuditLog,
)
from apps.tenancy.super_admin.utils import build_invoice_number, current_month_range, previous_month_range, to_money

from .serializers import (
    AuditEventSerializer,
    BillingMrrSerializer,
    DashboardDataSerializer,
    InvoiceCreateSerializer,
    InvoicePaymentCreateSerializer,
    InvoicePaymentSerializer,
    InvoiceSerializer,
    InvoiceUpdateSerializer,
    POLICY_GROUP_METADATA,
    PolicyGroupSerializer,
    PolicySettingsSerializer,
    PolicySerializer,
    ProvisionSchoolRequestSerializer,
    ProvisionSchoolResponseSerializer,
    SchoolTenantDetailSerializer,
    SchoolTenantListSerializer,
    SchoolTenantUpdateSerializer,
    SubscriptionPlanCreateSerializer,
    SubscriptionPlanSerializer,
    SubscriptionPlanUpdateSerializer,
)


DEFAULT_POLICY_ROWS = [
    {
        "key": "password.min_length",
        "category": "security",
        "description": "Minimum password length for super-admin and platform accounts.",
        "value": 10,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": False,
        "default_value": 10,
        "updated_by": "system",
    },
    {
        "key": "session.timeout_minutes",
        "category": "security",
        "description": "Session timeout before re-authentication is required.",
        "value": 30,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": False,
        "default_value": 30,
        "updated_by": "system",
    },
    {
        "key": "mfa.required",
        "category": "security",
        "description": "Require MFA for super-admin accounts.",
        "value": True,
        "value_type": "boolean",
        "is_toggle": True,
        "is_overridable": False,
        "default_value": True,
        "updated_by": "system",
    },
    {
        "key": "tenant.public_schema_only",
        "category": "data_isolation",
        "description": "Super-admin APIs are restricted to the public schema.",
        "value": True,
        "value_type": "boolean",
        "is_toggle": True,
        "is_overridable": False,
        "default_value": True,
        "updated_by": "system",
    },
    {
        "key": "audit.retention_days",
        "category": "data_isolation",
        "description": "Audit log retention window in days.",
        "value": 365,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": False,
        "default_value": 365,
        "updated_by": "system",
    },
    {
        "key": "gst.rate_percent",
        "category": "billing",
        "description": "Default GST rate used for billing calculations.",
        "value": 18,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": True,
        "default_value": 18,
        "updated_by": "system",
    },
    {
        "key": "invoice.payment_terms_days",
        "category": "billing",
        "description": "Payment due window for issued invoices.",
        "value": 15,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": True,
        "default_value": 15,
        "updated_by": "system",
    },
    {
        "key": "backup.retention_days",
        "category": "system",
        "description": "Daily backup retention window.",
        "value": 30,
        "value_type": "number",
        "is_toggle": False,
        "is_overridable": True,
        "default_value": 30,
        "updated_by": "system",
    },
    {
        "key": "multi_tenancy.enabled",
        "category": "system",
        "description": "Controls whether schema-based multi-tenancy is enabled.",
        "value": False,
        "value_type": "boolean",
        "is_toggle": True,
        "is_overridable": False,
        "default_value": False,
        "updated_by": "system",
    },
]


class SuperAdminPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 200


class SuperAdminBaseAPIView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def _public_queryset(self, model):
        return model.objects.using("default").all()

    def _client_ip(self, request):
        forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
        if forwarded:
            return forwarded.split(",")[0].strip()
        return request.META.get("REMOTE_ADDR", "")

    def _paginate(self, request, queryset, serializer_class):
        paginator = SuperAdminPagination()
        page = paginator.paginate_queryset(queryset, request, view=self)
        serializer = serializer_class(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def _normalize_school_payload(self, payload):
        if not isinstance(payload, dict):
            return payload

        normalized = dict(payload)
        aliases = {
            "udiseCode": "udise_code",
            "shortCode": "short_code",
            "subdomainUrl": "subdomain_url",
            "shardRegion": "shard_region",
            "storageRegion": "storage_region",
            "backupRetention": "backup_retention",
            "ssoMethod": "sso_method",
        }
        for alias, canonical in aliases.items():
            if alias in normalized and canonical not in normalized:
                normalized[canonical] = normalized[alias]
        return normalized


def _safe_float(value) -> float:
    if value is None:
        return 0.0
    return float(Decimal(str(value)))


def _invoice_grand_total(invoice: SuperAdminInvoice) -> float:
    tax_breakdown = invoice.tax_breakdown or {}
    if isinstance(tax_breakdown, dict) and tax_breakdown.get("grand_total") is not None:
        return _safe_float(tax_breakdown.get("grand_total"))

    subtotal = _safe_float(tax_breakdown.get("subtotal"))
    total_tax = _safe_float(tax_breakdown.get("total_tax"))
    return subtotal + total_tax


def _invoice_tax_total(invoice: SuperAdminInvoice) -> float:
    tax_breakdown = invoice.tax_breakdown or {}
    if isinstance(tax_breakdown, dict) and tax_breakdown.get("total_tax") is not None:
        return _safe_float(tax_breakdown.get("total_tax"))

    subtotal = _safe_float(tax_breakdown.get("subtotal"))
    grand_total = _invoice_grand_total(invoice)
    return max(grand_total - subtotal, 0.0)


def _money_words(value: float) -> str:
    return f"{value:.2f} INR"


def _schema_name_for(subdomain_url: str) -> str:
    cleaned = "".join(character.lower() if character.isalnum() else "_" for character in subdomain_url)
    cleaned = cleaned.strip("_") or "school"
    return f"school_{cleaned}"[:63]


def _gen_admin_password(length: int = 14) -> str:
    """Generate a cryptographically secure password meeting basic complexity."""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    while True:
        pwd = "".join(secrets.choice(alphabet) for _ in range(length))
        if (any(c.islower() for c in pwd)
                and any(c.isupper() for c in pwd)
                and any(c.isdigit() for c in pwd)
                and any(c in "!@#$%" for c in pwd)):
            return pwd


def _safe_school_code(subdomain: str, length: int = 32) -> str:
    """Derive a unique-safe school code from the subdomain."""
    code = subdomain.upper().replace("-", "_").replace(".", "_")
    # Keep only alphanumeric + underscore
    code = "".join(c if c.isalnum() or c == "_" else "_" for c in code)
    return code[:length]


def _ensure_default_policies():
    for row in DEFAULT_POLICY_ROWS:
        SuperAdminPolicy.objects.using("default").update_or_create(
            key=row["key"],
            defaults=row,
        )


def _policy_groups():
    _ensure_default_policies()
    policies = list(SuperAdminPolicy.objects.using("default").all().order_by("category", "key"))
    grouped = []
    for category, metadata in POLICY_GROUP_METADATA.items():
        grouped.append(
            {
                "category": category,
                "label": metadata["label"],
                "description": metadata["description"],
                "policies": [policy for policy in policies if policy.category == category],
            }
        )
    return grouped


def normalize_board(value: str) -> str:
    """Fix #12 – normalize board values: uppercase + replace spaces with underscores.

    Collapses variants like "SSC AP" and "SSC_AP" to a single consistent key.
    """
    if not value:
        return "OTHER"
    return value.strip().upper().replace(" ", "_")


class DashboardKPIView(SuperAdminBaseAPIView):
    def get(self, request):
        tenants = self._public_queryset(SchoolTenant)
        invoices = self._public_queryset(SuperAdminInvoice).exclude(status="cancelled")

        total_schools = tenants.count()
        active_schools = tenants.exclude(status__in=["suspended", "archived", "pending"]).count()
        active_students   = Student.objects.using("default").filter(status="active").count()
        inactive_students = Student.objects.using("default").filter(status="inactive").count()
        total_students    = active_students + inactive_students
        total_staff = Staff.objects.using("default").count()

        current_start, current_end = current_month_range()
        previous_start, previous_end = previous_month_range()

        current_month_invoices = invoices.filter(invoice_date__gte=current_start, invoice_date__lt=current_end)
        previous_month_invoices = invoices.filter(invoice_date__gte=previous_start, invoice_date__lt=previous_end)

        current_mrr = sum(_invoice_grand_total(invoice) for invoice in current_month_invoices)
        previous_mrr = sum(_invoice_grand_total(invoice) for invoice in previous_month_invoices)
        mrr_trend = 0.0
        if previous_mrr:
            mrr_trend = round(((current_mrr - previous_mrr) / previous_mrr) * 100, 2)

        # Fix #12 – normalise board names and merge variants (e.g. "SSC AP" / "SSC_AP")
        _raw_board_rows = list(tenants.values("board").annotate(count=Count("id")))
        _board_counts: dict = {}
        for row in _raw_board_rows:
            key = normalize_board(row.get("board") or "")
            _board_counts[key] = _board_counts.get(key, 0) + (row.get("count") or 0)
        board_breakdown = []
        for board_key, board_count in sorted(_board_counts.items()):
            board_breakdown.append(
                {
                    "board": board_key,
                    "count": board_count,
                    "percent": round((board_count / total_schools) * 100, 2) if total_schools else 0,
                }
            )

        overdue_invoices = invoices.filter(status="overdue").count()
        blocked_tenants = tenants.filter(status__in=["suspended", "archived"]).count()

        # Fix #5 – prefetch all school names in a single query to eliminate N+1
        audit_logs = list(self._public_queryset(TenantAuditLog).order_by("-created_at")[:10])
        _tenant_ids = [log.tenant_id for log in audit_logs if log.tenant_id]
        _school_name_map = dict(
            self._public_queryset(SchoolTenant)
            .filter(tenant_id__in=_tenant_ids)
            .values_list("tenant_id", "name")
        )
        recent_events = []
        for log in audit_logs:
            details = log.details if isinstance(log.details, dict) else {}
            recent_events.append(
                {
                    "id": str(log.id),
                    "timestamp": log.created_at,
                    "actor": log.actor_username or "System",
                    "action": log.action,
                    "detail": details.get("message") or log.action,
                    "severity": "error" if log.status == "failed" else ("warning" if log.status == "partial" else "info"),
                    "tenantId": log.tenant_id,
                    "schoolName": details.get("school_name") or _school_name_map.get(log.tenant_id),
                }
            )

        _STATE_NAMES = {
            '36': 'Telangana', '37': 'Andhra Pradesh', '29': 'Karnataka',
            '33': 'Tamil Nadu', '27': 'Maharashtra', '07': 'Delhi',
            '24': 'Gujarat', '32': 'Kerala', '19': 'West Bengal', '09': 'Uttar Pradesh',
        }
        _PLAN_PRICING = {
            'starter': 4500, 'standard': 9000, 'premium': 19500,
            'enterprise': 34500, 'trial': 0, 'custom': 0,
        }

        def normalize_state(value: str) -> str:
            """Fix #10 – accept both GST numeric codes and full state names stored in SchoolTenant.state."""
            if not value:
                return "Unknown"
            # If already a known state name, use it directly
            if value in _STATE_NAMES.values():
                return value
            # Otherwise treat as a GST state code and look it up
            return _STATE_NAMES.get(value.strip(), value)

        # Fix #7 – compute MoM new-student enrollment trend
        current_month_students = Student.objects.using("default").filter(
            created_at__gte=current_start, created_at__lt=current_end
        ).count()
        previous_month_students = Student.objects.using("default").filter(
            created_at__gte=previous_start, created_at__lt=previous_end
        ).count()
        if previous_month_students:
            students_trend = round(
                ((current_month_students - previous_month_students) / previous_month_students) * 100, 1
            )
        else:
            students_trend = 0.0

        # Fix #6 – compute actual MRR per plan from current-month invoices
        # Fall back to _PLAN_PRICING estimate only for plans that have no invoices yet
        actual_mrr_by_plan: dict = {}
        for inv in current_month_invoices.select_related("tenant"):
            plan_key = (inv.tenant.plan if inv.tenant else "trial").lower()
            actual_mrr_by_plan[plan_key] = actual_mrr_by_plan.get(plan_key, 0.0) + _invoice_grand_total(inv)

        # Build a subquery for live student counts per school
        _student_sq = (
            Student.objects.filter(school_id=OuterRef("pk"))
            .values("school_id")
            .annotate(cnt=Count("id"))
            .values("cnt")
        )
        tenants_with_students = tenants.annotate(
            live_student_count=Coalesce(Subquery(_student_sq, output_field=IntegerField()), 0)
        )

        state_breakdown = []
        for row in tenants_with_students.values("state").annotate(
            count=Count("id"), students=Sum("live_student_count")
        ).order_by("-count"):
            code = (row.get("state") or "").strip()
            state_breakdown.append({
                "state": normalize_state(code),  # Fix #10
                "code": code,
                "count": row.get("count") or 0,
                "students": row.get("students") or 0,
            })

        plan_breakdown = []
        for row in tenants_with_students.values("plan").annotate(
            count=Count("id"), students=Sum("live_student_count")
        ).order_by("-count"):
            plan_key = (row.get("plan") or "trial").lower()
            plan_count = row.get("count") or 0
            # Fix #6 – prefer actual invoice revenue; fall back to plan pricing for new installs with no invoices
            if plan_key in actual_mrr_by_plan:
                mrr_val = actual_mrr_by_plan[plan_key]
            else:
                mrr_val = float(plan_count * _PLAN_PRICING.get(plan_key, 0))
            plan_breakdown.append({
                "plan": plan_key.capitalize(),
                "count": plan_count,
                "mrr": mrr_val,
                "students": row.get("students") or 0,
            })

        payload = {
            "totalSchools": total_schools,
            "activeSchools": active_schools,
            "totalStudents": total_students,
            "activeStudents": active_students,
            "inactiveStudents": inactive_students,
            "totalStaff": total_staff,
            "mrr": {
                "current": to_money(current_mrr),
                "previous": to_money(previous_mrr),
                "trend": mrr_trend,
            },
            "alertCount": overdue_invoices + blocked_tenants,
            "overdueCount": overdue_invoices,
            "blockedCount": blocked_tenants,
            "boardBreakdown": board_breakdown,
            "trends": {
                "students": students_trend,  # Fix #7 – real MoM enrollment change
                "mrr": mrr_trend,
            },
            "recentEvents": recent_events,
            "stateBreakdown": state_breakdown,
            "planBreakdown": plan_breakdown,
        }
        return Response(DashboardDataSerializer(payload).data)

    def _school_name_for_tenant(self, tenant_id):
        if not tenant_id:
            return None
        tenant = self._public_queryset(SchoolTenant).filter(tenant_id=tenant_id).only("name").first()
        return tenant.name if tenant else None


class SchoolTenantListView(SuperAdminBaseAPIView):
    ORDERING_ALIASES = {
        "created_at": "provisioned_at",
        "updated_at": "last_activity_at",
    }

    def get_queryset(self):
        student_sq = (
            Student.objects.filter(school_id=OuterRef("pk"))
            .values("school_id")
            .annotate(cnt=Count("id"))
            .values("cnt")
        )
        active_student_sq = (
            Student.objects.filter(school_id=OuterRef("pk"), status="active")
            .values("school_id")
            .annotate(cnt=Count("id"))
            .values("cnt")
        )
        staff_sq = (
            Staff.objects.filter(school_id=OuterRef("pk"))
            .values("school_id")
            .annotate(cnt=Count("id"))
            .values("cnt")
        )
        queryset = (
            self._public_queryset(SchoolTenant)
            .annotate(
                live_student_count=Coalesce(Subquery(student_sq, output_field=IntegerField()), 0),
                live_active_student_count=Coalesce(Subquery(active_student_sq, output_field=IntegerField()), 0),
                live_staff_count=Coalesce(Subquery(staff_sq, output_field=IntegerField()), 0),
            )
            .order_by("-provisioned_at", "name")
        )

        search = self.request.query_params.get("search", "").strip()
        if search:
            # Always filter by school name.
            # Only extend to internal identifiers (tenant_id, subdomain,
            # gstin) when the query is long enough to be intentional
            # (>=5 chars), avoiding false matches on short queries like
            # "ab" hitting random hex characters inside tenant IDs.
            name_q = Q(name__icontains=search)
            if len(search) >= 5:
                queryset = queryset.filter(
                    name_q
                    | Q(tenant_id__icontains=search)
                    | Q(subdomain_url__icontains=search)
                    | Q(gstin__icontains=search)
                )
            else:
                queryset = queryset.filter(name_q)

        _VALID_STATUS_PARAMS = {"active", "trial", "suspended", "archived"}
        status_value = self.request.query_params.get("status")
        if status_value and status_value in _VALID_STATUS_PARAMS:
            if status_value == "active":
                # "active" = schools with status explicitly set to "active".
                queryset = queryset.filter(status="active")
            elif status_value == "trial":
                # "trial" maps to plan field — schools on the trial plan that
                # are not yet suspended or archived.
                queryset = queryset.filter(plan="trial").exclude(status__in=["archived", "suspended"])
            elif status_value == "suspended":
                queryset = queryset.filter(status="suspended")
            elif status_value == "archived":
                queryset = queryset.filter(status="archived")

        for field in ["plan", "board", "region", "state"]:
            value = self.request.query_params.get(field)
            if value:
                queryset = queryset.filter(**{field: value})

        ordering = self.request.query_params.get("ordering")
        if ordering:
            is_desc = ordering.startswith("-")
            key = ordering.lstrip("-")
            normalized = self.ORDERING_ALIASES.get(key, key)
            allowed = {"name", "status", "plan", "provisioned_at", "student_count", "staff_count", "last_activity_at"}
            if normalized in allowed:
                queryset = queryset.order_by(f"-{normalized}" if is_desc else normalized)

        # ── Health-flag filter ────────────────────────────────────────────────
        health_flag = self.request.query_params.get("health_flag")
        if health_flag:
            today = timezone.now().date()
            if health_flag == "billing_overdue":
                overdue_ids = (
                    SuperAdminInvoice.objects.using("default")
                    .filter(
                        Q(status="overdue")
                        | Q(due_date__lt=today, status__in=["draft", "sent"])
                    )
                    .exclude(tenant__isnull=True)
                    .values_list("tenant_id", flat=True)
                )
                queryset = queryset.filter(pk__in=list(overdue_ids))
            elif health_flag == "trial_ending":
                cutoff_start = today - timedelta(days=37)
                cutoff_end = today - timedelta(days=23)
                queryset = queryset.filter(
                    plan="trial",
                    provisioned_at__date__gte=cutoff_start,
                    provisioned_at__date__lte=cutoff_end,
                )
            elif health_flag == "gstin_missing":
                queryset = queryset.filter(Q(gstin="") | Q(gstin__isnull=True))
            elif health_flag == "storage_80":
                # No per-tenant storage tracking yet — return empty
                queryset = queryset.none()

        return queryset

    def _health_flags_counts(self):
        """Return health-flag counts over ALL non-archived tenants."""
        today = timezone.now().date()
        base = self._public_queryset(SchoolTenant).exclude(status="archived")

        overdue_ids = (
            SuperAdminInvoice.objects.using("default")
            .filter(
                Q(status="overdue")
                | Q(due_date__lt=today, status__in=["draft", "sent"])
            )
            .exclude(tenant__isnull=True)
            .values_list("tenant_id", flat=True)
        )
        billing_overdue = base.filter(pk__in=list(overdue_ids)).count()

        cutoff_start = today - timedelta(days=37)
        cutoff_end = today - timedelta(days=23)
        trial_ending = base.filter(
            plan="trial",
            provisioned_at__date__gte=cutoff_start,
            provisioned_at__date__lte=cutoff_end,
        ).count()

        gstin_missing = base.filter(Q(gstin="") | Q(gstin__isnull=True)).count()

        return {
            "billing_overdue": billing_overdue,
            "storage_80": 0,          # placeholder — no storage field yet
            "trial_ending": trial_ending,
            "gstin_missing": gstin_missing,
        }

    def _status_counts(self):
        """Return per-tab counts computed from the DB (no pagination), so tab
        badges always reflect the real totals regardless of current page."""
        base = self._public_queryset(SchoolTenant)
        return {
            "all":       base.count(),
            "active":    base.filter(status="active").count(),
            "trial":     base.filter(plan="trial").exclude(status__in=["archived", "suspended"]).count(),
            "suspended": base.filter(status="suspended").count(),
            "archived":  base.filter(status="archived").count(),
        }

    def get(self, request):
        resp = self._paginate(request, self.get_queryset(), SchoolTenantListSerializer)
        resp.data["health_flags_counts"] = self._health_flags_counts()
        resp.data["status_counts"] = self._status_counts()
        return resp


class SchoolTenantExportXlsxView(SchoolTenantListView):
    """Export all schools matching the current filters as an Excel (.xlsx) file."""

    def get(self, request):
        import io
        from datetime import datetime

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from django.http import HttpResponse

        qs = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schools"

        headers = [
            "School Name", "Tenant ID", "State", "Board",
            "Plan", "Status", "GSTIN", "Students", "Staff", "Provisioned At",
        ]

        header_fill = PatternFill(start_color="5B4FCF", end_color="5B4FCF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        col_widths = [32, 18, 18, 10, 10, 12, 18, 10, 8, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        row_align = Alignment(vertical="center")
        for row_idx, school in enumerate(qs.iterator(chunk_size=500), start=2):
            provisioned = (
                school.provisioned_at.strftime("%d/%m/%Y %H:%M")
                if school.provisioned_at else "-"
            )
            row_data = [
                school.name or "-",
                school.tenant_id or "-",
                school.state or "-",
                school.board or "-",
                school.plan or "-",
                school.status or "-",
                school.gstin or "-",
                getattr(school, "live_student_count", school.student_count or 0),
                getattr(school, "live_staff_count", school.staff_count or 0),
                provisioned,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = row_align

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"schools-export-{timestamp}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class SchoolTenantProvisionView(SuperAdminBaseAPIView):
    def post(self, request):
        serializer = ProvisionSchoolRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        subdomain = data["subdomain_url"].strip().lower()

        if self._public_queryset(SchoolTenant).filter(subdomain_url=subdomain).exists():
            return Response({"detail": "Subdomain already exists."}, status=status.HTTP_400_BAD_REQUEST)

        User = get_user_model()

        # Resolve admin credentials (use provided or auto-generate)
        admin_username = (data.get("admin_username") or "").strip() or f"{subdomain.replace('-', '_').replace('.', '_')}_admin"
        if User.objects.filter(username=admin_username).exists():
            admin_username = f"{admin_username[:24]}_{uuid4().hex[:4]}"
        admin_password = (data.get("admin_password") or "").strip() or _gen_admin_password()

        try:
            with transaction.atomic():
                # 1. Create SchoolTenant (provisioning record)
                tenant = self._public_queryset(SchoolTenant).create(
                    tenant_id=f"TNT_{uuid4().hex[:8].upper()}",
                    name=data["name"],
                    short_code=data.get("short_code") or data["name"][:10].upper(),
                    subdomain_url=subdomain,
                    schema_name=_schema_name_for(subdomain),
                    shard_region=data.get("shard_region") or "default",
                    storage_region=data.get("storage_region") or "default",
                    backup_retention=data.get("backup_retention") or 30,
                    sso_method=data.get("sso_method") or "native",
                    api_access=False,
                    plan=data["plan"],
                    status="onboarding",
                    provisioned_at=timezone.now(),
                    board=data["board"],
                    state=data["state"],
                    region=data.get("shard_region") or "default",
                    seats=data.get("seats") or 0,
                    student_count=0,
                    staff_count=0,
                    gstin=data.get("gstin") or "",
                    pan=data.get("pan") or "",
                    udise_code=data.get("udise_code") or "",
                    brand_color=data.get("brand_color") or "",
                    principal_name=data.get("principal_name") or "",
                    principal_email=data.get("principal_email") or "",
                    principal_phone=data.get("principal_phone") or "",
                    campus_address=data.get("campus_address") or "",
                    city=data.get("city") or "",
                    pin_code=data.get("pin_code") or "",
                    affiliation_number=data.get("affiliation_number") or "",
                )

                # 2. Create ERP School record (used for login + data isolation)
                school_code = _safe_school_code(subdomain)
                if School.objects.filter(code=school_code).exists():
                    school_code = f"{school_code[:28]}_{uuid4().hex[:3].upper()}"
                erp_school = School.objects.create(
                    name=data["name"],
                    code=school_code,
                    subdomain=subdomain,
                    is_active=True,
                )

                # 3a. Create Domain record so school_info_view can resolve this subdomain
                Domain.objects.get_or_create(
                    domain=f"{subdomain}.eskoolia.com",
                    defaults={"tenant": tenant, "is_primary": True},
                )

                # 3. Create school admin user
                admin_user = User.objects.create_user(
                    username=admin_username,
                    email=f"admin@{subdomain}.eskoolia.com",
                    password=admin_password,
                    first_name="School",
                    last_name="Admin",
                )
                admin_user.school = erp_school
                admin_user.is_school_admin = True
                admin_user.save(update_fields=["school", "is_school_admin"])

        except Exception as exc:
            return Response(
                {"detail": f"Provisioning failed: {exc}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        log_audit(
            action="school.provision",
            tenant_id=tenant.tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "school_name": tenant.name,
                "subdomain": tenant.subdomain_url,
                "plan": tenant.plan,
                "board": tenant.board,
                "erp_school_id": erp_school.id,
                "admin_username": admin_username,
            },
        )

        response = ProvisionSchoolResponseSerializer(
            {
                "tenant_id": tenant.tenant_id,
                "status": tenant.status,
                "message": "School provisioned successfully",
                "school_id": erp_school.id,
                "admin_username": admin_username,
                "admin_password": admin_password,
            }
        )
        return Response(response.data, status=status.HTTP_201_CREATED)


class SchoolTenantDetailView(SuperAdminBaseAPIView):
    def get_object(self, tenant_id):
        try:
            return self._public_queryset(SchoolTenant).get(tenant_id=tenant_id)
        except SchoolTenant.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound(f"School '{tenant_id}' not found.")

    def get(self, request, tenant_id):
        tenant = self.get_object(tenant_id)
        return Response(SchoolTenantDetailSerializer(tenant).data)

    def patch(self, request, tenant_id):
        tenant = self.get_object(tenant_id)
        serializer = SchoolTenantUpdateSerializer(tenant, data=self._normalize_school_payload(request.data), partial=True)
        serializer.is_valid(raise_exception=True)
        updated = serializer.save()

        log_audit(
            action="school.update",
            tenant_id=updated.tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "school_name": updated.name,
                "affected_fields": list(serializer.validated_data.keys()),
            },
        )

        return Response(SchoolTenantDetailSerializer(updated).data)

    def delete(self, request, tenant_id):
        tenant = self.get_object(tenant_id)
        if tenant.status == "archived":
            return Response({"detail": "School is already archived."}, status=status.HTTP_400_BAD_REQUEST)
        tenant.status = "archived"
        tenant.api_access = False
        tenant.save(update_fields=["status", "api_access"])

        log_audit(
            action="school.archive",
            tenant_id=tenant.tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"school_name": tenant.name},
        )

        return Response(status=status.HTTP_204_NO_CONTENT)


class SchoolImpersonateView(SuperAdminBaseAPIView):
    """Mint a short-lived impersonation JWT and return a tenant handoff URL."""

    def post(self, request, tenant_id: str):
        try:
            tenant = self._public_queryset(SchoolTenant).get(tenant_id=tenant_id)
        except SchoolTenant.DoesNotExist:
            return Response({"detail": "Tenant not found."}, status=status.HTTP_404_NOT_FOUND)

        User = get_user_model()
        target_username = request.data.get("username")
        target = None

        # Look up user AND generate tokens inside the tenant schema so the JWT
        # user-id resolves correctly when the token is later validated on the
        # tenant's own domain (e.g. mastermind.eskoolia.local:8000).
        with schema_context(tenant.schema_name):
            if target_username:
                target = User.objects.filter(username=target_username, is_active=True).first()
            if target is None:
                target = (
                    User.objects.filter(is_active=True)
                    .order_by("-is_school_admin", "-is_superuser", "id")
                    .first()
                )
            if target is None:
                return Response(
                    {"detail": "No active staff user found for this tenant."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            refresh = RefreshToken.for_user(target)
            refresh["impersonated_by"] = request.user.username
            refresh["tenant_id"] = tenant.tenant_id
            access = refresh.access_token
            access["impersonated_by"] = request.user.username
            access["tenant_id"] = tenant.tenant_id

        scheme = "https" if request.is_secure() else "http"
        raw_host = request.get_host()
        host = raw_host.split(":")[0]
        port_part = (":" + raw_host.split(":")[1]) if ":" in raw_host else ""
        target_subdomain = tenant.subdomain_url or ""
        target_host = f"{target_subdomain}.{host}{port_part}" if target_subdomain else f"{host}{port_part}"
        handoff_url = f"{scheme}://{target_host}/login?impersonate=1&token={str(access)}"

        log_audit(
            action="auth.impersonate",
            tenant_id=tenant.tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"target_username": target.username},
        )

        return Response({
            "tenant_id": tenant.tenant_id,
            "username": target.username,
            "access": str(access),
            "refresh": str(refresh),
            "handoff_url": handoff_url,
            "expires_in": int(access.lifetime.total_seconds()),
        })


class SchoolResetAdminPasswordView(SuperAdminBaseAPIView):
    """
    POST /api/super-admin/schools/{tenant_id}/reset-admin-password/

    Generates a new secure password for the school's admin account,
    applies it immediately, and returns it ONCE to the Super Admin.
    The password is never stored — only shown in this single response.
    Works for all schools, including those provisioned before credential
    vault storage was available.
    """

    def post(self, request, tenant_id: str):
        tenant = get_object_or_404(
            self._public_queryset(SchoolTenant), tenant_id=tenant_id
        )

        User = get_user_model()

        # Resolve the ERP School record linked to this tenant via subdomain
        erp_school = School.objects.filter(subdomain=tenant.subdomain_url).first()
        if not erp_school:
            return Response(
                {"detail": "No ERP school record found for this tenant."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Find the primary school admin user
        admin_user = (
            User.objects.filter(school=erp_school, is_school_admin=True, is_active=True)
            .order_by("id")
            .first()
        )
        if not admin_user:
            return Response(
                {"detail": "No active school admin user found for this school."},
                status=status.HTTP_404_NOT_FOUND,
            )

        new_password = _gen_admin_password()
        admin_user.set_password(new_password)
        admin_user.save(update_fields=["password"])

        log_audit(
            action="school.admin_password_reset",
            tenant_id=tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "school_name": tenant.name,
                "target_username": admin_user.username,
                "reset_by": request.user.username,
            },
        )

        return Response({
            "admin_username": admin_user.username,
            "admin_password": new_password,
            "message": (
                "Password reset successfully. "
                "Share these credentials with the school. "
                "This password will not be shown again."
            ),
        })


class SchoolLogoUploadView(SuperAdminBaseAPIView):
    """Upload or replace a school's logo image."""

    parser_classes = [parsers.MultiPartParser, parsers.FormParser]
    _ALLOWED_MIME = {"image/png", "image/jpeg", "image/webp", "image/gif", "image/svg+xml"}
    _MAX_BYTES = 5 * 1024 * 1024  # 5 MB

    def post(self, request, tenant_id: str):
        school = get_object_or_404(
            self._public_queryset(SchoolTenant), tenant_id=tenant_id
        )
        logo = request.FILES.get("logo")
        if not logo:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if logo.content_type not in self._ALLOWED_MIME:
            return Response(
                {"error": "Only PNG, JPEG, WEBP, GIF, or SVG images are allowed."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if logo.size > self._MAX_BYTES:
            return Response(
                {"error": "File size exceeds 5 MB limit."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext = os.path.splitext(logo.name)[1].lower() or ".png"
        upload_dir = os.path.join(settings.MEDIA_ROOT, "school_logos")
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"{tenant_id}{ext}"
        file_path = os.path.join(upload_dir, filename)
        with open(file_path, "wb+") as dest:
            for chunk in logo.chunks():
                dest.write(chunk)

        logo_url = f"{settings.MEDIA_URL}school_logos/{filename}"
        school.logo_url = logo_url
        school.save(update_fields=["logo_url"])

        log_audit(
            action="school.logo_upload",
            tenant_id=school.tenant_id,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"logo_url": logo_url},
        )
        return Response({"logo_url": logo_url}, status=status.HTTP_200_OK)


class AuditLogListView(SuperAdminBaseAPIView):
    def get_queryset(self):
        queryset = self._public_queryset(TenantAuditLog).order_by("-created_at")

        actor = self.request.query_params.get("actor")
        if actor:
            queryset = queryset.filter(Q(actor_username__icontains=actor) | Q(actor_user_id__icontains=actor))

        action = self.request.query_params.get("action")
        if action:
            queryset = queryset.filter(action=action)

        tenant_id = self.request.query_params.get("tenant_id")
        if tenant_id:
            queryset = queryset.filter(tenant_id=tenant_id)

        severity = self.request.query_params.get("severity")
        if severity == "critical":
            queryset = queryset.filter(status="failed")
        elif severity == "warning":
            queryset = queryset.filter(status="partial")
        elif severity == "info":
            queryset = queryset.filter(status="success")

        date_from = self.request.query_params.get("date_from")
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(actor_username__icontains=search)
                | Q(action__icontains=search)
                | Q(tenant_id__icontains=search)
                | Q(schema_name__icontains=search)
            )

        return queryset

    def get(self, request):
        return self._paginate(request, self.get_queryset(), AuditEventSerializer)


class AuditLogExportView(SuperAdminBaseAPIView):
    def get(self, request):
        queryset = AuditLogListView()
        queryset.request = request
        logs = queryset.get_queryset()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "id",
            "timestamp",
            "actor",
            "actor_ip",
            "action",
            "detail",
            "severity",
            "tenant_id",
            "school_name",
            "status",
            "error_message",
        ])

        serializer = AuditEventSerializer(logs, many=True)
        for row in serializer.data:
            writer.writerow([
                row["id"],
                row["timestamp"],
                row["actor"],
                row.get("actor_ip") or "",
                row["action"],
                row["detail"],
                row["severity"],
                row.get("tenant_id") or "",
                row.get("school_name") or "",
                row.get("status") or "",
                row.get("error_message") or "",
            ])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="audit-log.csv"'
        return response


class BillingInvoiceListCreateView(SuperAdminBaseAPIView):
    def get_queryset(self):
        queryset = self._public_queryset(SuperAdminInvoice).order_by("-invoice_date", "-created_at")

        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        school_name = self.request.query_params.get("school_name")
        if school_name:
            queryset = queryset.filter(school_name__icontains=school_name)

        tenant_id = self.request.query_params.get("tenant_id")
        if tenant_id:
            queryset = queryset.filter(tenant__tenant_id=tenant_id)

        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search)
                | Q(school_name__icontains=search)
                | Q(buyer_name__icontains=search)
                | Q(tenant__tenant_id__icontains=search)
            )

        date_from = self.request.query_params.get("date_from")
        if date_from:
            queryset = queryset.filter(invoice_date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            queryset = queryset.filter(invoice_date__lte=date_to)

        return queryset

    def get(self, request):
        return self._paginate(request, self.get_queryset(), InvoiceSerializer)

    def post(self, request):
        # Allow callers to explicitly override the duplicate guard when they
        # have confirmed the intent (e.g. correction invoice, multi-license).
        force = str(request.data.get("force", "")).lower() in ("true", "1", "yes")

        serializer = InvoiceCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        tenant = None
        if data.get("tenant_id"):
            tenant = self._public_queryset(SchoolTenant).filter(tenant_id=data["tenant_id"]).first()

        # ── Duplicate invoice guard ─────────────────────────────────────────
        # For a given school, billing month, and plan (matched by first line
        # item description), only one active invoice should exist.  Return
        # 409 so the UI can surface the conflict and ask for confirmation
        # before re-submitting with force=true.
        if tenant and not force:
            inv_year = data["invoice_date"].year
            inv_month = data["invoice_date"].month
            first_desc = (
                data["line_items"][0].get("description", "").strip()
                if data["line_items"]
                else ""
            )
            existing_qs = (
                self._public_queryset(SuperAdminInvoice)
                .filter(tenant=tenant, invoice_date__year=inv_year, invoice_date__month=inv_month)
                .exclude(status="cancelled")
            )
            conflict = None
            for inv in existing_qs:
                inv_desc = (inv.line_items[0].get("description", "").strip() if inv.line_items else "")
                if not first_desc or inv_desc == first_desc:
                    conflict = inv
                    break
            if conflict:
                return Response(
                    {
                        "code": "duplicate_invoice",
                        "detail": (
                            f"Invoice {conflict.invoice_number} already exists for "
                            f"{conflict.school_name} for "
                            f"{data['invoice_date'].strftime('%B %Y')}. "
                            "Void/cancel the existing invoice first, or pass force=true to override."
                        ),
                        "existing_invoice": {
                            "id": str(conflict.id),
                            "invoice_number": conflict.invoice_number,
                            "status": conflict.status,
                        },
                    },
                    status=status.HTTP_409_CONFLICT,
                )
        # ── End duplicate guard ─────────────────────────────────────────────

        line_items = data["line_items"]
        subtotal = sum(_safe_float(item.get("amount")) for item in line_items)
        tax_breakdown = data.get("tax_breakdown") or {}
        if not tax_breakdown:
            tax_total = sum(_safe_float(item.get("gst_amount")) for item in line_items)
            grand_total = subtotal + tax_total
            tax_breakdown = {
                "subtotal": to_money(subtotal),
                "igst": to_money(tax_total),
                "cgst": 0,
                "sgst": 0,
                "total_tax": to_money(tax_total),
                "grand_total": to_money(grand_total),
                "amount_in_words": _money_words(grand_total),
            }

        try:
            invoice = self._public_queryset(SuperAdminInvoice).create(
                invoice_number=data.get("invoice_number") or build_invoice_number(),
                tenant=tenant,
                school_name=data["school_name"],
                invoice_date=data["invoice_date"],
                due_date=data["due_date"],
                status=data.get("status") or "draft",
                seller_name=data["seller_name"],
                seller_gstin=data.get("seller_gstin") or "",
                seller_state=data["seller_state"],
                buyer_name=data["buyer_name"],
                buyer_gstin=data.get("buyer_gstin") or "",
                buyer_state=data["buyer_state"],
                line_items=line_items,
                tax_breakdown=tax_breakdown,
                notes=data.get("notes") or "",
                terms_conditions=data.get("terms_conditions") or "",
                reverse_charge=data.get("reverse_charge") or False,
            )
        except IntegrityError:
            return Response(
                {"error": {"code": "duplicate_invoice", "message": "Invoice number already exists."}},
                status=status.HTTP_409_CONFLICT,
            )

        log_audit(
            action="invoice.generated",
            tenant_id=tenant.tenant_id if tenant else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"school_name": invoice.school_name, "invoice_number": invoice.invoice_number, "status": invoice.status},
        )

        return Response(InvoiceSerializer(invoice).data, status=status.HTTP_201_CREATED)


class BillingMRRView(SuperAdminBaseAPIView):
    def get(self, request):
        invoices = self._public_queryset(SuperAdminInvoice).exclude(status="cancelled")
        current_start, current_end = current_month_range()
        previous_start, previous_end = previous_month_range()

        current_month = invoices.filter(invoice_date__gte=current_start, invoice_date__lt=current_end)
        previous_month = invoices.filter(invoice_date__gte=previous_start, invoice_date__lt=previous_end)

        current_mrr = sum(_invoice_grand_total(invoice) for invoice in current_month)
        previous_mrr = sum(_invoice_grand_total(invoice) for invoice in previous_month)

        gst_collected = sum(_invoice_tax_total(invoice) for invoice in current_month if invoice.status == "paid")
        # Outstanding = sum of remaining due on every non-cancelled invoice.
        # Includes partially_paid invoices (only the unpaid remainder counts).
        outstanding_amount = float(
            invoices.filter(status__in=["draft", "sent", "partially_paid", "overdue"])
            .aggregate(total=Coalesce(Sum("due_amount"), Decimal("0")))["total"]
        )
        at_risk_amount = float(
            invoices.filter(status="overdue")
            .aggregate(total=Coalesce(Sum("due_amount"), Decimal("0")))["total"]
        )

        trend_percent = 0.0
        if previous_mrr:
            trend_percent = round(((current_mrr - previous_mrr) / previous_mrr) * 100, 2)

        payload = {
            "current_mrr": to_money(current_mrr),
            "previous_mrr": to_money(previous_mrr),
            "gst_collected": to_money(gst_collected),
            "outstanding_amount": to_money(outstanding_amount),
            "at_risk_amount": to_money(at_risk_amount),
            "trend_percent": trend_percent,
        }
        return Response(BillingMrrSerializer(payload).data)


class BillingPlansView(SuperAdminBaseAPIView):
    """Subscription plans catalog for the Super Admin billing screen.

    Plans are persisted in `tenancy.SubscriptionPlan`. GST 18% under SAC
    998313 (Education software) is applied at invoice time, not stored
    on the plan.
    """

    CATALOG_META = {
        "gst_percent": 18,
        "sac_code": "998313",
        "sac_description": "Education software",
        "currency": "INR",
    }

    def get(self, request):
        plans = SubscriptionPlan.objects.filter(is_active=True).order_by(
            "sort_order", "price_inr", "name"
        )
        return Response({
            "plans": SubscriptionPlanSerializer(plans, many=True).data,
            **self.CATALOG_META,
        })

    def post(self, request):
        serializer = SubscriptionPlanCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        plan = SubscriptionPlan.objects.create(
            code=data["code"],
            name=data["name"],
            description=data.get("description", ""),
            price_inr=data["price_inr"],
            billing_cycle=data.get("billing_cycle", "monthly"),
            popular=data.get("popular", False),
            features=data.get("features", []),
            sort_order=data.get("sort_order", 0),
            is_active=data.get("is_active", True),
            sac_code=data.get("sac_code", "998313"),
        )

        log_audit(
            action="plan.created",
            actor_user=request.user if request.user.is_authenticated else None,
            status="success",
            details={
                "code": plan.code,
                "name": plan.name,
                "price_inr": str(plan.price_inr),
                "billing_cycle": plan.billing_cycle,
            },
        )

        return Response(
            SubscriptionPlanSerializer(plan).data,
            status=status.HTTP_201_CREATED,
        )


class BillingPlanDetailView(SuperAdminBaseAPIView):
    """Update or delete a single subscription plan by code."""

    def _get(self, code):
        try:
            return SubscriptionPlan.objects.get(code=code)
        except SubscriptionPlan.DoesNotExist:
            return None

    def patch(self, request, code):
        plan = self._get(code)
        if plan is None:
            return Response({"detail": "Plan not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = SubscriptionPlanUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        for field, value in serializer.validated_data.items():
            setattr(plan, field, value)
        plan.save()

        log_audit(
            action="plan.updated",
            actor_user=request.user if request.user.is_authenticated else None,
            status="success",
            details={"code": plan.code, "fields": list(serializer.validated_data.keys())},
        )
        return Response(SubscriptionPlanSerializer(plan).data)

    def delete(self, request, code):
        plan = self._get(code)
        if plan is None:
            return Response({"detail": "Plan not found."}, status=status.HTTP_404_NOT_FOUND)

        # Soft-delete: mark inactive — plan disappears from catalog,
        # historical invoices and school associations remain intact.
        plan.is_active = False
        plan.save(update_fields=["is_active", "updated_at"])

        log_audit(
            action="plan.archived",
            actor_user=request.user if request.user.is_authenticated else None,
            status="success",
            details={"code": plan.code, "name": plan.name, "price_inr": str(plan.price_inr)},
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class BillingGSTR1ExportView(SuperAdminBaseAPIView):
    def get(self, request):
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        queryset = BillingInvoiceListCreateView()
        queryset.request = request
        invoices = queryset.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GSTR-1 Report"

        headers = [
            "Invoice Number",
            "School Name",
            "Tenant ID",
            "Buyer GSTIN",
            "Place of Supply",
            "Invoice Date",
            "Due Date",
            "Status",
            "Subtotal (₹)",
            "IGST (₹)",
            "CGST (₹)",
            "SGST (₹)",
            "Total Tax (₹)",
            "Grand Total (₹)",
        ]

        header_fill = PatternFill(start_color="5B4FCF", end_color="5B4FCF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws.row_dimensions[1].height = 30

        col_widths = [24, 28, 18, 20, 18, 14, 14, 10, 16, 14, 14, 14, 14, 16]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        row_align = Alignment(vertical="center")
        for row_idx, invoice in enumerate(invoices, start=2):
            tax = invoice.tax_breakdown or {}
            ws.cell(row=row_idx, column=1,  value=invoice.invoice_number).alignment = row_align
            ws.cell(row=row_idx, column=2,  value=invoice.school_name or "-").alignment = row_align
            ws.cell(row=row_idx, column=3,  value=(invoice.tenant.tenant_id if invoice.tenant_id and invoice.tenant else "-")).alignment = row_align
            ws.cell(row=row_idx, column=4,  value=invoice.buyer_gstin or "Unregistered").alignment = row_align
            ws.cell(row=row_idx, column=5,  value=invoice.buyer_state or "-").alignment = row_align
            ws.cell(row=row_idx, column=6,  value=str(invoice.invoice_date) if invoice.invoice_date else "-").alignment = row_align
            ws.cell(row=row_idx, column=7,  value=str(invoice.due_date) if invoice.due_date else "-").alignment = row_align
            ws.cell(row=row_idx, column=8,  value=invoice.status or "-").alignment = row_align
            ws.cell(row=row_idx, column=9,  value=float(tax.get("subtotal", 0))).alignment = row_align
            ws.cell(row=row_idx, column=10, value=float(tax.get("igst", 0))).alignment = row_align
            ws.cell(row=row_idx, column=11, value=float(tax.get("cgst", 0))).alignment = row_align
            ws.cell(row=row_idx, column=12, value=float(tax.get("sgst", 0))).alignment = row_align
            ws.cell(row=row_idx, column=13, value=float(tax.get("total_tax", 0))).alignment = row_align
            ws.cell(row=row_idx, column=14, value=float(tax.get("grand_total", _invoice_grand_total(invoice)))).alignment = row_align

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"gstr1-report-{timestamp}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PoliciesView(SuperAdminBaseAPIView):
    def get(self, request):
        return Response(PolicyGroupSerializer(_policy_groups(), many=True).data)

    def patch(self, request):
        updates = request.data
        if isinstance(updates, dict) and "updates" in updates and isinstance(updates["updates"], dict):
            updates = updates["updates"]

        if not isinstance(updates, dict) or not updates:
            return Response({"detail": "No policy updates supplied."}, status=status.HTTP_400_BAD_REQUEST)

        _ensure_default_policies()
        changed = []

        for key, value in updates.items():
            seed = next((row for row in DEFAULT_POLICY_ROWS if row["key"] == key), None)
            category = seed["category"] if seed else "system"
            description = seed["description"] if seed else "Updated platform policy."
            value_type = seed["value_type"] if seed else ("boolean" if isinstance(value, bool) else ("number" if isinstance(value, (int, float)) else "string"))
            default_value = seed["default_value"] if seed else value
            is_toggle = seed["is_toggle"] if seed else isinstance(value, bool)
            is_overridable = seed["is_overridable"] if seed else True

            policy, _ = self._public_queryset(SuperAdminPolicy).update_or_create(
                key=key,
                defaults={
                    "category": category,
                    "description": description,
                    "value": value,
                    "value_type": value_type,
                    "is_toggle": is_toggle,
                    "is_overridable": is_overridable,
                    "default_value": default_value,
                    "updated_by": getattr(request.user, "username", "system"),
                },
            )
            changed.append(policy)

        log_audit(
            action="policy.updated",
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"updates": updates, "count": len(changed)},
        )

        return Response(PolicyGroupSerializer(_policy_groups(), many=True).data)


class PolicySettingsView(SuperAdminBaseAPIView):
    def get(self, request):
        payload = {
            "system": {
                "name": "eSkoolia Platform",
                "version": "1.0.0",
                "environment": "development" if settings.DEBUG else "production",
                "multi_tenancy_enabled": settings.MULTI_TENANCY_ENABLED,
                "timezone": getattr(settings, "TIME_ZONE", "UTC"),
            },
            "notification": {
                "email_enabled": True,
                "sms_enabled": False,
                "push_enabled": True,
                "webhook_enabled": True,
            },
            "integrations": {
                "google_sso": True,
                "microsoft_sso": True,
                "saml_enabled": False,
                "ldap_enabled": False,
            },
            "storage": {
                "provider": "s3",
                "bucket": "eskoolia-prod",
                "cdn_enabled": True,
                "max_file_size_mb": 100,
            },
            "api": {
                "rate_limiting": True,
                "api_versioning": "v1",
                "api_documentation_url": "/api/docs/",
                "swagger_ui_enabled": True,
            },
        }
        return Response(PolicySettingsSerializer(payload).data)


class PoliciesExportView(SuperAdminBaseAPIView):
    def get(self, request):
        export_format = (request.query_params.get("format") or "json").lower()
        data = PolicyGroupSerializer(_policy_groups(), many=True).data

        if export_format in {"yaml", "yml"}:
            response = HttpResponse(yaml.safe_dump(data, sort_keys=False, default_flow_style=False), content_type="application/yaml")
            response["Content-Disposition"] = 'attachment; filename="policies.yaml"'
            return response

        response = HttpResponse(json.dumps(data, indent=2, default=str), content_type="application/json")
        response["Content-Disposition"] = 'attachment; filename="policies.json"'
        return response


class BillingInvoiceDetailView(SuperAdminBaseAPIView):
    def _get_invoice(self, invoice_id: str) -> SuperAdminInvoice:
        return get_object_or_404(
            self._public_queryset(SuperAdminInvoice), id=invoice_id
        )

    def get(self, request, invoice_id: str):
        invoice = self._get_invoice(invoice_id)
        return Response(InvoiceSerializer(invoice).data)

    def patch(self, request, invoice_id: str):
        invoice = self._get_invoice(invoice_id)
        if invoice.status in ("paid", "cancelled"):
            return Response(
                {"detail": f"Invoice is '{invoice.status}' and cannot be edited. "
                           "To make corrections, cancel the invoice and re-issue a new one."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = InvoiceUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if not data:
            return Response(InvoiceSerializer(invoice).data)

        update_fields = []
        previous_status = invoice.status
        for field, value in data.items():
            setattr(invoice, field, value)
            update_fields.append(field)
        update_fields.append("updated_at")
        invoice.save(update_fields=update_fields)

        log_audit(
            action="invoice.updated",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "invoice_number": invoice.invoice_number,
                "previous_status": previous_status,
                "changed_fields": list(data.keys()),
            },
        )
        return Response(InvoiceSerializer(invoice).data)

    def delete(self, request, invoice_id: str):
        """Cancel (soft-delete) an invoice. Hard delete is never allowed."""
        invoice = self._get_invoice(invoice_id)
        previous_status = invoice.status
        invoice.status = "cancelled"
        invoice.save(update_fields=["status", "updated_at"])
        log_audit(
            action="invoice.cancelled",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "invoice_number": invoice.invoice_number,
                "previous_status": previous_status,
            },
        )
        return Response(InvoiceSerializer(invoice).data)


class BillingInvoiceMarkPaidView(SuperAdminBaseAPIView):
    """Shortcut to settle the full outstanding balance via a single payment
    entry in the ledger. Prefer POST .../payments/ for partial amounts.
    """

    def post(self, request, invoice_id: str):
        invoice = get_object_or_404(
            self._public_queryset(SuperAdminInvoice), id=invoice_id
        )
        if invoice.status == "cancelled":
            return Response(
                {"detail": "Cancelled invoices cannot be marked paid."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        previous_status = invoice.status
        with transaction.atomic():
            invoice = (
                self._public_queryset(SuperAdminInvoice)
                .select_for_update()
                .get(pk=invoice.pk)
            )
            outstanding = invoice.grand_total() - (invoice.paid_amount or 0)
            if outstanding > 0:
                SuperAdminInvoicePayment.objects.create(
                    invoice=invoice,
                    amount=outstanding,
                    paid_on=timezone.localdate(),
                    method=request.data.get("method") or "bank_transfer",
                    reference_no=request.data.get("reference_no") or "",
                    received_by=request.user if request.user.is_authenticated else None,
                    notes=request.data.get("notes") or "Full settlement via Mark as paid",
                )
            invoice.status = "paid"  # force final state even if grand_total is 0
            invoice.save(update_fields=["status", "updated_at"])
            invoice.recalculate()

        log_audit(
            action="invoice.mark_paid",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "invoice_number": invoice.invoice_number,
                "previous_status": previous_status,
            },
        )
        return Response(InvoiceSerializer(invoice).data)


class BillingInvoicePaymentsView(SuperAdminBaseAPIView):
    """GET   list payments for an invoice
    POST  record a new payment (supports partial amounts).
    """

    def _get_invoice(self, invoice_id: str) -> SuperAdminInvoice:
        return get_object_or_404(
            self._public_queryset(SuperAdminInvoice), id=invoice_id
        )

    def get(self, request, invoice_id: str):
        invoice = self._get_invoice(invoice_id)
        return Response(
            InvoicePaymentSerializer(invoice.payments.all(), many=True).data
        )

    def post(self, request, invoice_id: str):
        serializer = InvoicePaymentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        with transaction.atomic():
            invoice = (
                self._public_queryset(SuperAdminInvoice)
                .select_for_update()
                .get(id=invoice_id)
            )
            if invoice.status == "cancelled":
                return Response(
                    {"detail": "Cannot record payments against a cancelled invoice."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            grand = invoice.grand_total()
            outstanding = grand - (invoice.paid_amount or Decimal("0"))
            amount = Decimal(str(data["amount"]))
            if amount > outstanding + Decimal("0.005"):
                return Response(
                    {
                        "detail": (
                            f"Amount {amount} exceeds outstanding balance "
                            f"{outstanding}. Adjust the amount or issue a credit note."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            payment = SuperAdminInvoicePayment.objects.create(
                invoice=invoice,
                amount=amount,
                paid_on=data.get("paid_on") or timezone.localdate(),
                method=data.get("method") or "bank_transfer",
                reference_no=data.get("reference_no") or "",
                notes=data.get("notes") or "",
                received_by=request.user if request.user.is_authenticated else None,
            )
            invoice.recalculate()

        log_audit(
            action="invoice.payment_recorded",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "invoice_number": invoice.invoice_number,
                "amount": str(amount),
                "method": payment.method,
                "new_status": invoice.status,
                "remaining_due": str(invoice.due_amount),
            },
        )
        return Response(
            {
                "payment": InvoicePaymentSerializer(payment).data,
                "invoice": InvoiceSerializer(invoice).data,
            },
            status=status.HTTP_201_CREATED,
        )


class BillingInvoicePaymentDetailView(SuperAdminBaseAPIView):
    """DELETE a previously recorded payment (e.g. cheque bounced, wrong entry).
    Reverses ledger entry and recalculates invoice totals.
    """

    def delete(self, request, invoice_id: str, payment_id: str):
        with transaction.atomic():
            invoice = get_object_or_404(
                self._public_queryset(SuperAdminInvoice).select_for_update(),
                id=invoice_id,
            )
            payment = get_object_or_404(
                invoice.payments.select_for_update(), id=payment_id
            )
            payment_summary = {
                "amount": str(payment.amount),
                "method": payment.method,
                "reference_no": payment.reference_no,
            }
            payment.delete()
            invoice.recalculate()

        log_audit(
            action="invoice.payment_reversed",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "invoice_number": invoice.invoice_number,
                "reversed_payment": payment_summary,
                "new_status": invoice.status,
                "remaining_due": str(invoice.due_amount),
            },
        )
        return Response(InvoiceSerializer(invoice).data)


class BillingInvoiceReminderView(SuperAdminBaseAPIView):
    def post(self, request, invoice_id: str):
        invoice = get_object_or_404(
            self._public_queryset(SuperAdminInvoice), id=invoice_id
        )

        # No mailer wired in public schema; bump draft → sent and record audit.
        if invoice.status == "draft":
            invoice.status = "sent"
            invoice.save(update_fields=["status", "updated_at"])

        log_audit(
            action="invoice.reminder_sent",
            tenant_id=invoice.tenant.tenant_id if invoice.tenant_id else None,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={"invoice_number": invoice.invoice_number, "channel": "email"},
        )
        return Response(
            {
                "invoice_number": invoice.invoice_number,
                "status": invoice.status,
                "reminder_recorded": True,
            }
        )


# ---------------------------------------------------------------------------
# LLM Access Management (Tasks 4 & 5)
# ---------------------------------------------------------------------------

class SchoolLLMListView(SuperAdminBaseAPIView):
    """GET /api/v1/super-admin/llm/schools/
    List all School records with their LLM access status.
    Supports ?llm_enabled=true|false filter.
    """

    def get(self, request):
        qs = School.objects.select_related("llm_enabled_by").order_by("name")

        llm_param = request.query_params.get("llm_enabled")
        if llm_param is not None:
            qs = qs.filter(llm_enabled=(llm_param.lower() == "true"))

        # Build name → tenant_id mapping so the frontend can key by tenant_id
        tenant_by_name = dict(
            self._public_queryset(SchoolTenant).values_list("name", "tenant_id")
        )

        data = [
            {
                "id": s.id,
                "name": s.name,
                "code": s.code,
                "tenant_id": tenant_by_name.get(s.name, ""),
                "llm_enabled": s.llm_enabled,
                "llm_enabled_at": s.llm_enabled_at,
                "llm_enabled_by": s.llm_enabled_by.username if s.llm_enabled_by_id else None,
                "is_active": s.is_active,
            }
            for s in qs
        ]
        return Response({"count": len(data), "results": data})


class ToggleSchoolLLMView(SuperAdminBaseAPIView):
    """POST /api/v1/super-admin/llm/schools/<school_id>/
    Enable or disable LLM access for a school.
    Body: {"enabled": true|false}
    """

    def post(self, request, school_id: int):
        enabled = bool(request.data.get("enabled"))
        school = get_object_or_404(School, pk=school_id)

        school.llm_enabled = enabled
        school.llm_enabled_at = timezone.now() if enabled else None
        school.llm_enabled_by = request.user if enabled else None
        school.save(update_fields=["llm_enabled", "llm_enabled_at", "llm_enabled_by"])

        log_audit(
            action="llm.access_enabled" if enabled else "llm.access_disabled",
            tenant_id=school.code,
            status="success",
            actor_user=request.user,
            actor_ip=self._client_ip(request),
            details={
                "school_id": school.id,
                "school_name": school.name,
                "llm_enabled": enabled,
            },
        )

        return Response(
            {
                "school_id": school.id,
                "school_name": school.name,
                "llm_enabled": school.llm_enabled,
                "llm_enabled_at": school.llm_enabled_at,
                "llm_enabled_by": request.user.username if enabled else None,
            },
            status=status.HTTP_200_OK,
        )
