from django.db import IntegrityError, transaction
import logging
import csv
import io
import os
import re
from uuid import uuid4
from datetime import datetime
from django.conf import settings
from django.core.cache import cache
from django.core.files.storage import default_storage
from django.utils import timezone
from django.db.models.deletion import ProtectedError
from django.db.models import Count, Q
import requests
from rest_framework import permissions, status, viewsets
from config.pagination import ApiPageNumberPagination
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response
from .forms import StudentValidationModelForm
from .models import (
    Guardian,
    PromotionAuditLog,
    PromotionBatch,
    PromotionRecord,
    Student,
    StudentCategory,
    StudentDocument,
    StudentGroup,
    StudentMultiClassRecord,
    StudentRecordAudit,
    StudentSubjectAssignment,
    StudentPromotionHistory,
    StudentTransferHistory,
)
from .serializers import (
    GuardianSerializer,
    PromotionAiRequestSerializer,
    PromotionAuditLogSerializer,
    PromotionBatchCreateSerializer,
    PromotionBatchSerializer,
    PromotionBulkUpdateSerializer,
    PromotionRecordSerializer,
    PromotionRecordUpdateSerializer,
    StudentCategorySerializer,
    StudentDocumentSerializer,
    StudentGroupSerializer,
    StudentMultiClassBulkSaveSerializer,
    StudentMultiClassRecordSerializer,
    StudentRecordAuditSerializer,
    StudentSubjectAssignmentRequestSerializer,
    StudentSubjectAssignmentSerializer,
    StudentPromoteRequestSerializer,
    StudentPromotionHistorySerializer,
    PincodeLookupQuerySerializer,
    StudentListSerializer,
    StudentSerializer,
    StudentTransferHistorySerializer,
)


logger = logging.getLogger(__name__)


class TenantScopedModelViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = ApiPageNumberPagination
    model = None
    permission_codes = {}

    def get_required_permission_code(self):
        action = getattr(self, "action", None)
        if action and action in self.permission_codes:
            return self.permission_codes[action]
        return self.permission_codes.get("*")

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        code = self.get_required_permission_code()
        if not code:
            return
        user = request.user
        if user.is_superuser:
            return
        if not hasattr(user, "has_permission_code") or not user.has_permission_code(code):
            raise PermissionDenied("You do not have permission to perform this action.")

    def get_queryset(self):
        user = self.request.user
        qs = self.model.objects.all()
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(school_id=user.school_id)
        return qs.none()

    def perform_create(self, serializer):
        school = self.request.user.school
        if not school and getattr(self.request, "school", None):
            school = self.request.school
        if school:
            serializer.save(school=school)
            return
        # Superuser without school can provide school explicitly in payload
        serializer.save()

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            raise ValidationError({"detail": "Duplicate value violates a uniqueness constraint."})

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            raise ValidationError({"detail": "Duplicate value violates a uniqueness constraint."})

    def partial_update(self, request, *args, **kwargs):
        try:
            return super().partial_update(request, *args, **kwargs)
        except IntegrityError:
            raise ValidationError({"detail": "Duplicate value violates a uniqueness constraint."})


class StudentCategoryViewSet(TenantScopedModelViewSet):
    model = StudentCategory
    serializer_class = StudentCategorySerializer
    pagination_class = ApiPageNumberPagination
    permission_codes = {"*": "student_info.student_category.view"}

    def _build_validation_error_response(self, field_errors=None, message="Validation failed"):
        return Response(
            {
                "success": False,
                "error_code": "VALIDATION_ERROR",
                "message": message,
                "field_errors": field_errors or {},
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _normalize_field_errors(self, serializer_errors):
        normalized = {}
        for field, errors in serializer_errors.items():
            if isinstance(errors, (list, tuple)):
                normalized[field] = [str(error) for error in errors]
            else:
                normalized[field] = [str(errors)]
        return normalized

    def _first_error_message(self, field_errors):
        for errors in field_errors.values():
            if isinstance(errors, list) and errors:
                return str(errors[0])
        return "Validation failed"

    def get_queryset(self):
        qs = super().get_queryset().order_by("name")
        params = self.request.query_params
        search = str(params.get("search") or params.get("q") or "").strip()
        status_filter = str(params.get("status") or "").strip().lower()
        name_filter = str(params.get("name") or "").strip()

        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(code__icontains=search)
                | Q(description__icontains=search)
            )
        if name_filter:
            qs = qs.filter(name__icontains=name_filter)
        if status_filter in {"active", "inactive"}:
            qs = qs.filter(status=status_filter)
        return qs

    @action(detail=False, methods=["get"], url_path="check-name")
    def check_name(self, request):
        name = str(request.query_params.get("name") or "").strip()
        if not name:
            return Response({"success": False, "message": "Category name is required."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(name__iexact=name)
        if getattr(self, "action", None) == "check_name" and self.request.query_params.get("exclude_id"):
            queryset = queryset.exclude(id=self.request.query_params.get("exclude_id"))
        exists = queryset.exists()
        return Response({"success": True, "exists": exists, "message": "Category exists." if exists else "Category available."})

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            field_errors = self._normalize_field_errors(serializer.errors)
            return self._build_validation_error_response(field_errors, self._first_error_message(field_errors))

        try:
            self.perform_create(serializer)
        except IntegrityError:
            return self._build_validation_error_response(
                {"name": ["Category name already exists."]},
                "Category name already exists.",
            )

        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "success": True,
                "message": "Student category created successfully.",
                "data": serializer.data,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            field_errors = self._normalize_field_errors(serializer.errors)
            return self._build_validation_error_response(field_errors, self._first_error_message(field_errors))

        try:
            self.perform_update(serializer)
        except IntegrityError:
            return self._build_validation_error_response(
                {"name": ["Category name already exists."]},
                "Category name already exists.",
            )

        return Response(
            {
                "success": True,
                "message": "Student category updated successfully.",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=["patch"], url_path="bulk-status")
    def bulk_status(self, request):
        ids = request.data.get("ids") or []
        status_value = str(request.data.get("status") or "").strip().lower()
        if status_value not in {"active", "inactive"}:
            return self._build_validation_error_response({}, "Status must be either active or inactive.")
        if not isinstance(ids, list) or not ids:
            return self._build_validation_error_response({}, "Select at least one category.")

        queryset = self.get_queryset().filter(id__in=ids)
        already_count = queryset.filter(status=status_value).count()
        update_queryset = queryset.exclude(status=status_value)
        updated = update_queryset.update(status=status_value)

        if updated == 0:
            message = f"Selected categories are already {status_value}."
        elif already_count == 0:
            noun = "category" if updated == 1 else "categories"
            message = f"{updated} {noun} updated successfully."
        else:
            updated_noun = "category" if updated == 1 else "categories"
            already_noun = "category is" if already_count == 1 else "categories are"
            message = (
                f"{updated} {updated_noun} updated successfully. "
                f"{already_count} {already_noun} already {status_value}."
            )

        return Response({"success": True, "message": message})

    @action(detail=False, methods=["delete"], url_path="bulk-delete")
    def bulk_delete(self, request):
        ids = request.data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return self._build_validation_error_response({}, "Select at least one category.")

        queryset = self.get_queryset().filter(id__in=ids)
        blocked = queryset.filter(students__isnull=False).distinct()
        if blocked.exists():
            return self._build_validation_error_response({}, "One or more selected categories are assigned to students and cannot be deleted.")

        count = queryset.count()
        queryset.delete()
        return Response({"success": True, "message": f"{count} categories deleted successfully."}, status=status.HTTP_200_OK)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.students.exists():
            return self._build_validation_error_response(
                {},
                "Cannot delete category as it is assigned to students",
            )

        self.perform_destroy(instance)
        return Response(
            {
                "success": True,
                "message": "Student category deleted successfully.",
            },
            status=status.HTTP_200_OK,
        )


class StudentGroupViewSet(TenantScopedModelViewSet):
    model = StudentGroup
    serializer_class = StudentGroupSerializer
    pagination_class = ApiPageNumberPagination
    permission_codes = {"*": "student_info.student_group.view"}

    def _school_id(self):
        user = self.request.user
        return None if user.is_superuser else getattr(user, "school_id", None)

    def get_queryset(self):
        user = self.request.user
        qs = StudentGroup.objects.all()
        search = str(self.request.query_params.get("search") or "").strip()
        sort_by = str(self.request.query_params.get("sort_by") or "name").strip().lower()

        if user.is_superuser:
            qs = qs.annotate(students_count=Count("students", distinct=True))
        elif user.school_id:
            qs = qs.filter(school_id=user.school_id).annotate(students_count=Count("students", distinct=True))
        else:
            return qs.none()

        if search:
            qs = qs.filter(name__icontains=search)

        if sort_by == "count":
            return qs.order_by("-students_count", "name")

        if sort_by == "name":
            return qs.order_by("name")

        return qs.order_by("name")

    def perform_create(self, serializer):
        user = self.request.user
        school_id = getattr(user, "school_id", None)
        if not school_id:
            raw_school_id = self.request.data.get("school")
            if raw_school_id not in (None, ""):
                try:
                    school_id = int(raw_school_id)
                except (TypeError, ValueError):
                    raise ValidationError({"school": "School must be a valid numeric id."})

        if not school_id:
            raise ValidationError({"school": "School is required to create a student group."})

        serializer.save(school_id=school_id)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Unassign all students from this group (SET_NULL via FK cascade is automatic,
        # but we do it explicitly to be safe and return meaningful data)
        instance.students.update(student_group=None)
        self.perform_destroy(instance)
        return Response({"success": True, "message": "Group deleted successfully."}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        school_id = self._school_id()
        qs = Student.objects.filter(is_deleted=False, is_active=True)
        if school_id:
            qs = qs.filter(school_id=school_id)
        groups_qs = StudentGroup.objects.all()
        if school_id:
            groups_qs = groups_qs.filter(school_id=school_id)

        total = qs.count()
        assigned = qs.filter(student_group__isnull=False).count()
        unassigned = total - assigned
        house_count = groups_qs.filter(type="HOUSE").count()
        club_count = groups_qs.filter(type="CLUB").count()
        return Response({
            "totalStudents": total,
            "assigned": assigned,
            "unassigned": unassigned,
            "houseCount": house_count,
            "clubCount": club_count,
        })

    @action(detail=False, methods=["get"], url_path="students")
    def list_students(self, request):
        school_id = self._school_id()
        qs = Student.objects.filter(is_deleted=False, is_active=True)
        if school_id:
            qs = qs.filter(school_id=school_id)
        qs = qs.select_related("current_class", "current_section", "student_group")

        # Filters
        class_name = request.query_params.get("class")
        section_name = request.query_params.get("section")
        group_id = request.query_params.get("groupId")
        query_status = request.query_params.get("status")

        if class_name:
            qs = qs.filter(current_class__name=class_name)
        if section_name:
            qs = qs.filter(current_section__name=section_name)
        if group_id:
            ids = [gid.strip() for gid in group_id.split(",") if gid.strip().isdigit()]
            if ids:
                qs = qs.filter(student_group_id__in=ids)
        if query_status == "unassigned":
            qs = qs.filter(student_group__isnull=True)

        qs = qs.order_by(
            "current_class__numeric_order", "current_section__name",
            "first_name", "last_name"
        )

        result = []
        for st in qs:
            result.append({
                "id": st.id,
                "name": f"{st.first_name} {st.last_name}".strip(),
                "admissionNo": st.admission_no,
                "class": st.current_class.name if st.current_class else "",
                "section": st.current_section.name if st.current_section else "",
                "classIndex": st.current_class.numeric_order if st.current_class else 999,
                "currentGroupId": st.student_group_id,
                "gender": st.gender,
            })
        return Response(result)

    @action(detail=False, methods=["post"], url_path="assign")
    def assign(self, request):
        school_id = self._school_id()
        student_id = request.data.get("studentId")
        group_id = request.data.get("groupId")

        if not student_id:
            return Response({"error": "studentId is required"}, status=400)

        qs = Student.objects.filter(id=student_id)
        if school_id:
            qs = qs.filter(school_id=school_id)
        student = qs.first()
        if not student:
            return Response({"error": "Student not found"}, status=404)

        if group_id is None:
            student.student_group = None
        else:
            group_qs = StudentGroup.objects.filter(id=group_id)
            if school_id:
                group_qs = group_qs.filter(school_id=school_id)
            group = group_qs.first()
            if not group:
                return Response({"error": "Group not found"}, status=404)
            student.student_group = group

        student.save(update_fields=["student_group"])
        return Response({"studentId": student.id, "groupId": student.student_group_id})

    @action(detail=False, methods=["post"], url_path="bulk-assign")
    def bulk_assign(self, request):
        school_id = self._school_id()
        student_ids = request.data.get("studentIds", [])
        group_id = request.data.get("groupId")

        if not isinstance(student_ids, list) or not student_ids:
            return Response({"error": "studentIds must be a non-empty list"}, status=400)

        qs = Student.objects.filter(id__in=student_ids)
        if school_id:
            qs = qs.filter(school_id=school_id)

        if group_id is None:
            assigned = qs.update(student_group=None)
        else:
            group_qs = StudentGroup.objects.filter(id=group_id)
            if school_id:
                group_qs = group_qs.filter(school_id=school_id)
            group = group_qs.first()
            if not group:
                return Response({"error": "Group not found"}, status=404)
            assigned = qs.update(student_group=group)

        return Response({"assigned": assigned})

    @action(detail=False, methods=["get"], url_path="sortwell-preview")
    def sortwell_preview(self, request):
        school_id = self._school_id()
        scope = request.query_params.get("scope", "unassigned")

        houses = list(StudentGroup.objects.filter(
            **{"school_id": school_id} if school_id else {},
            type="HOUSE"
        ).order_by("id"))

        if not houses:
            return Response({"houses": [], "total": 0})

        qs = Student.objects.filter(is_deleted=False, is_active=True)
        if school_id:
            qs = qs.filter(school_id=school_id)
        if scope == "unassigned":
            qs = qs.filter(student_group__isnull=True)

        total = qs.count()
        n = len(houses)
        base = total // n
        remainder = total % n

        preview = []
        for i, house in enumerate(houses):
            count = base + (1 if i < remainder else 0)
            preview.append({
                "groupId": house.id,
                "groupName": house.name,
                "emoji": house.emoji,
                "color": house.color,
                "bgColor": house.bg_color,
                "count": count,
            })

        return Response({"houses": preview, "total": total, "houseCount": n})

    @action(detail=False, methods=["post"], url_path="sortwell")
    def sortwell(self, request):
        import random as rnd
        school_id = self._school_id()
        method = request.data.get("method", "random")
        scope = request.data.get("scope", "unassigned")

        houses = list(StudentGroup.objects.filter(
            **{"school_id": school_id} if school_id else {},
            type="HOUSE"
        ).order_by("id"))

        if not houses:
            return Response({"error": "No house-type groups found"}, status=400)

        with transaction.atomic():
            if scope == "all":
                # Clear all house-type assignments
                qs_clear = Student.objects.filter(student_group__type="HOUSE")
                if school_id:
                    qs_clear = qs_clear.filter(school_id=school_id)
                qs_clear.update(student_group=None)

            qs = Student.objects.filter(is_deleted=False, is_active=True)
            if school_id:
                qs = qs.filter(school_id=school_id)
            if scope == "unassigned":
                qs = qs.filter(student_group__isnull=True)

            students = list(qs.select_related("current_class"))

            if method == "random":
                rnd.shuffle(students)
            elif method == "alpha":
                students.sort(key=lambda s: f"{s.first_name} {s.last_name}".strip().lower())
            elif method == "classwise":
                students.sort(key=lambda s: (
                    s.current_class.numeric_order if s.current_class else 999,
                    f"{s.first_name} {s.last_name}".strip().lower()
                ))
            elif method == "gender":
                males = [s for s in students if s.gender == "male"]
                females = [s for s in students if s.gender != "male"]
                interleaved = []
                for idx in range(max(len(males), len(females))):
                    if idx < len(males):
                        interleaved.append(males[idx])
                    if idx < len(females):
                        interleaved.append(females[idx])
                students = interleaved

            distribution = {h.id: [] for h in houses}
            for idx, student in enumerate(students):
                house = houses[idx % len(houses)]
                distribution[house.id].append(student.id)

            total = 0
            for house_id, sids in distribution.items():
                if sids:
                    Student.objects.filter(id__in=sids).update(student_group_id=house_id)
                    total += len(sids)

        dist_result = [
            {"groupId": h.id, "groupName": h.name, "count": len(distribution[h.id])}
            for h in houses
        ]
        return Response({"assigned": total, "distribution": dist_result})

    @action(detail=True, methods=["post"], url_path="assign-students")
    def assign_students(self, request, *args, **kwargs):
        group = self.get_object()
        raw_ids = request.data.get("student_ids")
        if not isinstance(raw_ids, list) or not raw_ids:
            return Response(
                {"success": False, "message": "student_ids is required.", "field_errors": {"student_ids": "Select at least one student."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        valid_ids = []
        for value in raw_ids:
            try:
                valid_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        if not valid_ids:
            return Response(
                {"success": False, "message": "No valid student ids provided.", "field_errors": {"student_ids": "Invalid student ids."}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        students_qs = Student.objects.filter(id__in=valid_ids)
        if not request.user.is_superuser:
            students_qs = students_qs.filter(school_id=request.user.school_id)

        updated = students_qs.update(student_group=group)
        return Response(
            {
                "success": True,
                "message": f"{updated} student(s) assigned to group successfully.",
                "updated": updated,
            },
            status=status.HTTP_200_OK,
        )


class GuardianViewSet(TenantScopedModelViewSet):
    model = Guardian
    serializer_class = GuardianSerializer
    pagination_class = ApiPageNumberPagination
    permission_codes = {"*": "student_info.add_student.view"}


class StudentViewSet(TenantScopedModelViewSet):
    model = Student
    serializer_class = StudentSerializer
    pagination_class = ApiPageNumberPagination
    permission_codes = {
        "list": "student_info.student_list.view",
        "retrieve": "student_info.student_list.view",
        "create": "student_info.add_student.view",
        "update": "student_info.add_student.view",
        "partial_update": "student_info.add_student.view",
        "check_admission_no": "student_info.add_student.view",
        "upload_photo": "student_info.add_student.view",
        "pincode_details": "student_info.add_student.view",
        "destroy": "student_info.delete_student_record.view",
        "soft_delete": "student_info.delete_student_record.view",
        "restore": "student_info.delete_student_record.view",
        "permanent_delete": "student_info.delete_student_record.view",
        "promote": "student_info.student_promote.view",
        "auto_assign_classes": "student_info.student_list.view",
    }

    def get_serializer_class(self):
        if getattr(self, "action", None) == "list":
            return StudentListSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        user = self.request.user
        action = getattr(self, "action", None)

        if action == "list":
            qs = Student.objects.select_related("current_class", "current_section", "guardian", "category", "student_group").only(
                "id",
                "school_id",
                "academic_year_id",
                "admission_no",
                "roll_no",
                "first_name",
                "last_name",
                "date_of_birth",
                "gender",
                "phone",
                "category_id",
                "student_group_id",
                "guardian_id",
                "current_class_id",
                "current_section_id",
                "status",
                "is_disabled",
                "is_active",
                "is_deleted",
                "created_at",
            )
        else:
            qs = Student.objects.select_related(
                "school",
                "academic_year",
                "category",
                "student_group",
                "guardian",
                "current_class",
                "current_section",
                "admission_inquiry",
            )

        if action in {"retrieve", "create", "update", "partial_update"}:
            qs = qs.prefetch_related("documents")

        search = (self.request.query_params.get("search") or "").strip()
        class_id = self.request.query_params.get("class") or self.request.query_params.get("current_class")
        section_id = self.request.query_params.get("section") or self.request.query_params.get("current_section")
        academic_year_id = self.request.query_params.get("academic_year")
        is_active_param = (self.request.query_params.get("is_active") or "").strip().lower()
        include_deleted = (self.request.query_params.get("include_deleted") or "").strip().lower() in {"1", "true", "yes"}
        deleted_only = (self.request.query_params.get("deleted_only") or "").strip().lower() in {"1", "true", "yes"}
        unassigned_only = (self.request.query_params.get("unassigned") or "").strip().lower() in {"1", "true", "yes"}

        if deleted_only:
            qs = qs.filter(is_deleted=True)
        elif not include_deleted:
            qs = qs.filter(is_deleted=False)

        if search and not re.fullmatch(r"[A-Za-z0-9 _\-./]+", search):
            raise ValidationError({"search": "Please enter valid search text"})

        if class_id and not str(class_id).isdigit():
            raise ValidationError({"current_class": "Please select a valid class."})
        if section_id and not str(section_id).isdigit():
            raise ValidationError({"current_section": "Please select a valid section."})
        if academic_year_id and not str(academic_year_id).isdigit():
            raise ValidationError({"academic_year": "Please select a valid academic year."})

        from apps.core.models import AcademicYear, Class, Section

        if class_id:
            class_qs = Class.objects.filter(id=int(class_id))
            if not user.is_superuser:
                class_qs = class_qs.filter(school_id=user.school_id)
            if not class_qs.exists():
                raise ValidationError({"current_class": "Selected class is not available."})

        if section_id:
            section_qs = Section.objects.filter(id=int(section_id))
            if class_id:
                section_qs = section_qs.filter(school_class_id=int(class_id))
            if not user.is_superuser:
                section_qs = section_qs.filter(school_class__school_id=user.school_id)
            if not section_qs.exists():
                raise ValidationError({"current_section": "Selected section is not available."})

        if academic_year_id:
            year_qs = AcademicYear.objects.filter(id=int(academic_year_id))
            if not user.is_superuser:
                year_qs = year_qs.filter(school_id=user.school_id)
            if not year_qs.exists():
                raise ValidationError({"academic_year": "Selected academic year is not available."})

        if class_id:
            qs = qs.filter(current_class_id=class_id)
        if section_id:
            qs = qs.filter(current_section_id=section_id)
        if unassigned_only:
            qs = qs.filter(current_class_id__isnull=True, current_section_id__isnull=True)
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)
        if is_active_param in {"1", "true", "yes"}:
            qs = qs.filter(is_active=True)
        elif is_active_param in {"0", "false", "no"}:
            qs = qs.filter(is_active=False)
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(admission_no__icontains=search)
                | Q(roll_no__icontains=search)
            )

        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(school_id=user.school_id)
        return qs.none()

    def _field_alias(self, field):
        aliases = {
            "current_class": "class",
            "current_section": "section",
            "date_of_birth": "dob",
        }
        return aliases.get(field, field)

    def _normalize_pincode_office(self, office):
        office = office or {}
        return {
            "name": str(office.get("Name") or "").strip(),
            "branch_type": str(office.get("BranchType") or "").strip(),
            "delivery_status": str(office.get("DeliveryStatus") or "").strip(),
            "district": str(office.get("District") or "").strip(),
            "state": str(office.get("State") or "").strip(),
            "region": str(office.get("Region") or "").strip(),
            "division": str(office.get("Division") or "").strip(),
            "circle": str(office.get("Circle") or "").strip(),
            "taluk": str(office.get("Taluk") or "").strip(),
            "block": str(office.get("Block") or "").strip(),
            "country": str(office.get("Country") or "").strip(),
            "pincode": str(office.get("Pincode") or "").strip(),
        }

    def _city_from_post_office_name(self, office):
        raw_name = str((office or {}).get("name") or "").strip()
        if not raw_name:
            return ""

        # Remove common India Post branch suffixes so UI gets clean city-like values.
        cleaned = re.sub(r"\s*[-(]?\s*(?:S\.?O|B\.?O|H\.?O|G\.?P\.?O|SO|BO|HO|GPO)\s*[).-]*\s*$", "", raw_name, flags=re.IGNORECASE)
        return cleaned.strip()

    def _fetch_pincode_details(self, pincode):
        cache_key = f"student_pincode_details:v2:{pincode}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "application/json",
        }
        response = requests.get(f"https://api.postalpincode.in/pincode/{pincode}", timeout=8, headers=headers)
        response.raise_for_status()
        payload = response.json()

        if not isinstance(payload, list) or not payload:
            raise ValidationError({"pincode": "Invalid PIN Code"})

        entry = payload[0] or {}
        if str(entry.get("Status") or "").strip().lower() != "success":
            raise ValidationError({"pincode": "Invalid PIN Code"})

        offices = entry.get("PostOffice") or []
        normalized_offices = [self._normalize_pincode_office(office) for office in offices if office]
        normalized_offices = [office for office in normalized_offices if office.get("name")]

        if not normalized_offices:
            raise ValidationError({"pincode": "Invalid PIN Code"})

        selected_office = next((office for office in normalized_offices if office.get("district")), normalized_offices[0])
        district = str(selected_office.get("district") or "").strip()
        state = str(selected_office.get("state") or "").strip()

        if not district or not state:
            raise ValidationError({"pincode": "Invalid PIN Code"})

        city_options = []
        for office in normalized_offices:
            city_candidate = self._city_from_post_office_name(office)
            if city_candidate:
                city_options.append(city_candidate)

        # District fallback only when post-office names are unavailable.
        if not city_options and district:
            city_options = [district]

        city_options = sorted(set(city_options), key=lambda value: value.lower())
        city = city_options[0] if city_options else ""

        result = {
            "pincode": pincode,
            "state": state,
            "district": district,
            "city": city,
            "city_options": city_options,
            "selected_post_office": selected_office,
            "post_offices": normalized_offices,
            "multiple_post_offices": len(normalized_offices) > 1,
        }
        cache.set(cache_key, result, 24 * 60 * 60)
        return result

    def _normalize_field_errors(self, serializer_errors):
        normalized = {}
        for field, errors in serializer_errors.items():
            key = self._field_alias(str(field))
            if isinstance(errors, (list, tuple)):
                normalized[key] = str(errors[0]) if errors else "Validation failed"
            else:
                normalized[key] = str(errors)
        return normalized

    def _validation_response(self, serializer_errors=None, message="Validation failed"):
        return Response(
            {
                "success": False,
                "message": message,
                "field_errors": self._normalize_field_errors(serializer_errors or {}),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _normalize_import_row(self, raw_row):
        key_aliases = {
            "admission no": "admission_no",
            "admission_no": "admission_no",
            "admission number": "admission_no",
            "admission_number": "admission_no",
            "roll no": "roll_no",
            "roll_no": "roll_no",
            "roll number": "roll_no",
            "roll_number": "roll_no",
            "first name": "first_name",
            "first_name": "first_name",
            "last name": "last_name",
            "last_name": "last_name",
            "dob": "date_of_birth",
            "date of birth": "date_of_birth",
            "date_of_birth": "date_of_birth",
            "academic year": "academic_year",
            "academic_year": "academic_year",
            "class": "current_class",
            "current class": "current_class",
            "current_class": "current_class",
            "section": "current_section",
            "current section": "current_section",
            "current_section": "current_section",
            "custom gender": "custom_gender",
            "custom_gender": "custom_gender",
            "blood group": "blood_group",
            "blood_group": "blood_group",
            "address": "address_line",
            "address line": "address_line",
            "address_line": "address_line",
            "pin": "pincode",
            "zip": "pincode",
            "postal code": "pincode",
            "postal_code": "pincode",
            "student category": "category",
            "student_category": "category",
        }

        cleaned = {}
        for key, value in (raw_row or {}).items():
            key_text = str(key or "").strip()
            if not key_text:
                continue
            normalized_key = key_aliases.get(key_text.lower(), key_text.replace(" ", "_").lower())

            value_text = "" if value is None else str(value).strip()
            if value_text == "":
                cleaned[normalized_key] = ""
                continue

            if normalized_key == "date_of_birth":
                parsed = None
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        parsed = datetime.strptime(value_text, fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                cleaned[normalized_key] = parsed or value_text
                continue

            if normalized_key in {"academic_year", "current_class", "current_section", "category", "guardian"}:
                cleaned[normalized_key] = int(value_text) if value_text.isdigit() else value_text
                continue

            if normalized_key in {"is_active", "is_disabled"}:
                lowered = value_text.lower()
                if lowered in {"true", "1", "yes", "y"}:
                    cleaned[normalized_key] = True
                elif lowered in {"false", "0", "no", "n"}:
                    cleaned[normalized_key] = False
                else:
                    cleaned[normalized_key] = value_text
                continue

            cleaned[normalized_key] = value_text

        return cleaned

    def _duplicate_warning(self, validated_data):
        first_name = (validated_data.get("first_name") or "").strip()
        last_name = (validated_data.get("last_name") or "").strip()
        dob = validated_data.get("date_of_birth")
        phone = (validated_data.get("phone") or "").strip()
        if not (first_name and dob and phone):
            return None

        queryset = Student.objects.filter(
            school_id=self.request.user.school_id,
            first_name__iexact=first_name,
            last_name__iexact=last_name,
            date_of_birth=dob,
            phone=phone,
        )
        if queryset.exists():
            return "Possible duplicate found: same name, DOB and phone already exists"
        return None

    def _build_model_form_data(self, data):
        if hasattr(data, "dict"):
            raw_data = data.dict()
        else:
            raw_data = dict(data or {})

        normalized = {}
        for key, value in raw_data.items():
            if isinstance(value, (list, tuple)):
                normalized[str(key)] = value[-1] if value else ""
            else:
                normalized[str(key)] = value

        if "class" in normalized and "current_class" not in normalized:
            normalized["current_class"] = normalized.get("class")
        if "section" in normalized and "current_section" not in normalized:
            normalized["current_section"] = normalized.get("section")
        if "dob" in normalized and "date_of_birth" not in normalized:
            normalized["date_of_birth"] = normalized.get("dob")

        return normalized

    def _validate_with_model_form(self, data, instance=None, partial=False):
        form = StudentValidationModelForm(
            data=self._build_model_form_data(data),
            instance=instance,
            partial=partial,
            user=self.request.user,
        )
        if form.is_valid():
            return None
        return self._validation_response(form.errors)

    def create(self, request, *args, **kwargs):
        form_error_response = self._validate_with_model_form(request.data)
        if form_error_response:
            return form_error_response

        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return self._validation_response(serializer.errors)

        warning = self._duplicate_warning(serializer.validated_data)

        try:
            self.perform_create(serializer)
        except IntegrityError:
            return self._validation_response({"admission_no": ["Admission number already exists"]})

        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "success": True,
                "message": "Student added successfully",
                "warning": warning,
                "data": serializer.data,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        form_error_response = self._validate_with_model_form(request.data, instance=instance, partial=partial)
        if form_error_response:
            return form_error_response

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            return self._validation_response(serializer.errors)

        try:
            self.perform_update(serializer)
        except IntegrityError:
            return self._validation_response({"admission_no": ["Admission number already exists"]})

        return Response(
            {
                "success": True,
                "message": "Student updated successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="summary", permission_classes=[permissions.IsAuthenticated])
    def summary(self, request):
        """KPI summary for the Student List page.

        Honours search/current_class/current_section/academic_year filters.
        Always returns totals across is_active and deleted statuses so each
        KPI card reflects an absolute number (not limited by the currently
        selected status pill).
        """
        from datetime import timedelta
        from django.utils import timezone

        user = request.user
        qs = Student.objects.all()
        if not user.is_superuser:
            if not user.school_id:
                return Response(
                    {
                        "success": True,
                        "message": "Summary retrieved successfully",
                        "data": {
                            "total_count": 0,
                            "active_count": 0,
                            "inactive_count": 0,
                            "archived_count": 0,
                            "new_count": 0,
                            "docs_pending_count": 0,
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            qs = qs.filter(school_id=user.school_id)

        search = (request.query_params.get("search") or "").strip()
        class_id = request.query_params.get("class") or request.query_params.get("current_class")
        section_id = request.query_params.get("section") or request.query_params.get("current_section")
        academic_year_id = request.query_params.get("academic_year")

        if class_id and str(class_id).isdigit():
            qs = qs.filter(current_class_id=int(class_id))
        if section_id and str(section_id).isdigit():
            qs = qs.filter(current_section_id=int(section_id))
        if academic_year_id and str(academic_year_id).isdigit():
            qs = qs.filter(academic_year_id=int(academic_year_id))
        if search and re.fullmatch(r"[A-Za-z0-9 _\-./]+", search):
            qs = qs.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(admission_no__icontains=search)
                | Q(roll_no__icontains=search)
            )

        total = qs.count()
        archived = qs.filter(is_deleted=True).count()
        live = qs.filter(is_deleted=False)
        active = live.filter(is_active=True, is_disabled=False).count()
        inactive = live.filter(is_active=False).count()
        docs_pending = live.filter(is_disabled=True).count()
        cutoff = timezone.now() - timedelta(days=30)
        new_count = live.filter(created_at__gte=cutoff).count()

        return Response(
            {
                "success": True,
                "message": "Summary retrieved successfully",
                "data": {
                    "total_count": total,
                    "active_count": active,
                    "inactive_count": inactive,
                    "archived_count": archived,
                    "new_count": new_count,
                    "docs_pending_count": docs_pending,
                },
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="check-admission-no")
    def check_admission_no(self, request):
        admission_no = str(request.query_params.get("admission_no") or "").strip()
        exclude_id = str(request.query_params.get("exclude_id") or "").strip()

        if not admission_no:
            return Response(
                {
                    "success": False,
                    "message": "Admission number is required.",
                    "field_errors": {"admission_no": "Admission number is required."},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not re.fullmatch(r"[A-Za-z0-9]+", admission_no):
            return Response(
                {
                    "success": False,
                    "message": "Admission number should contain only letters and numbers.",
                    "field_errors": {"admission_no": "Admission number should contain only letters and numbers."},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(admission_no__iexact=admission_no)
        if exclude_id.isdigit():
            queryset = queryset.exclude(id=int(exclude_id))

        exists = queryset.exists()
        return Response(
            {
                "success": True,
                "exists": exists,
                "message": "Admission number already exists." if exists else "Admission number is available.",
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="pincode-details")
    def pincode_details(self, request):
        serializer = PincodeLookupQuerySerializer(data=request.query_params)
        if not serializer.is_valid():
            field_errors = {}
            for field, errors in serializer.errors.items():
                field_errors[field] = [str(error) for error in (errors or [])] if isinstance(errors, (list, tuple)) else [str(errors)]
            message = field_errors.get("pincode", ["Pincode must be exactly 6 digits."])[0]
            return Response(
                {
                    "success": False,
                    "message": message,
                    "field_errors": field_errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        pincode = serializer.validated_data["pincode"]

        try:
            data = self._fetch_pincode_details(pincode)
        except ValidationError as exc:
            message = "Invalid PIN Code"
            detail = exc.detail
            if isinstance(detail, dict):
                pincode_message = detail.get("pincode")
                if isinstance(pincode_message, list) and pincode_message:
                    message = str(pincode_message[0])
                elif isinstance(pincode_message, str):
                    message = pincode_message
            return Response(
                {
                    "success": False,
                    "message": message,
                    "field_errors": {"pincode": message},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except requests.RequestException:
            return Response(
                {
                    "success": False,
                    "message": "Unable to fetch pincode details right now.",
                    "field_errors": {"pincode": "Unable to fetch pincode details right now."},
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        return Response(
            {
                "success": True,
                "message": "Pincode details fetched successfully.",
                "data": data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["post"], url_path="upload-photo", parser_classes=[MultiPartParser, FormParser])
    def upload_photo(self, request):
        image = request.FILES.get("photo")
        if not image:
            return Response(
                {
                    "success": False,
                    "message": "Photo file is required.",
                    "field_errors": {"photo": "Photo file is required."},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowed_types = {"image/jpeg", "image/png"}
        if image.content_type not in allowed_types:
            return Response(
                {
                    "success": False,
                    "message": "Only JPEG and PNG files are allowed.",
                    "field_errors": {"photo": "Only JPEG and PNG files are allowed."},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        max_size_bytes = 2 * 1024 * 1024
        if image.size > max_size_bytes:
            return Response(
                {
                    "success": False,
                    "message": "Photo size must be 2MB or less.",
                    "field_errors": {"photo": "Photo size must be 2MB or less."},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext = os.path.splitext(image.name or "")[1].lower()
        if ext not in {".jpg", ".jpeg", ".png"}:
            ext = ".jpg" if image.content_type == "image/jpeg" else ".png"

        relative_path = f"student_photos/{uuid4().hex}{ext}"
        saved_path = default_storage.save(relative_path, image)
        normalized = str(saved_path).replace("\\", "/")
        media_prefix = settings.MEDIA_URL if settings.MEDIA_URL.endswith("/") else f"{settings.MEDIA_URL}/"
        photo_url = request.build_absolute_uri(f"{media_prefix}{normalized}")

        return Response(
            {
                "success": True,
                "message": "Photo uploaded successfully.",
                "data": {"photo": photo_url},
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["post"], url_path="promote")
    def promote(self, request):
        serializer = StudentPromoteRequestSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        school_id = request.user.school_id
        if not school_id and not request.user.is_superuser:
            raise ValidationError("User school context is required.")

        target_class_id = data["to_class"]
        target_section_id = data.get("to_section")
        target_year_id = data.get("to_academic_year")

        from apps.core.models import AcademicYear, Class, Section

        # Validate target class/section/year belongs to same school for tenant safety
        class_qs = Class.objects.filter(id=target_class_id)
        if not request.user.is_superuser:
            class_qs = class_qs.filter(school_id=school_id)
        target_class = class_qs.first()
        if not target_class:
            raise ValidationError("Target class not found in your school.")

        if target_section_id:
            section_qs = Section.objects.filter(id=target_section_id, school_class_id=target_class_id)
            if not section_qs.exists():
                raise ValidationError("Target section not found under target class.")

        if target_year_id:
            year_qs = AcademicYear.objects.filter(id=target_year_id)
            if not request.user.is_superuser:
                year_qs = year_qs.filter(school_id=school_id)
            if not year_qs.exists():
                raise ValidationError("Target academic year not found in your school.")

        students_qs = Student.objects.filter(id__in=data["student_ids"])
        if not request.user.is_superuser:
            students_qs = students_qs.filter(school_id=school_id)
        students = list(students_qs)

        if not students:
            raise ValidationError("No students found for promotion.")

        # Track promoted count and errors for partial success
        promoted_count = 0
        failed_count = 0
        errors_list = []

        # Process each student individually to support partial success
        for st in students:
            try:
                # Validate: prevent promoting to same class
                if st.current_class_id == target_class_id:
                    failed_count += 1
                    errors_list.append({
                        "student_id": st.id,
                        "admission_no": st.admission_no,
                        "error": f"Cannot promote to the same class ({target_class.name})"
                    })
                    continue

                # Create promotion history
                StudentPromotionHistory.objects.create(
                    student=st,
                    from_class=st.current_class,
                    from_section=st.current_section,
                    to_class_id=target_class_id,
                    to_section_id=target_section_id,
                    to_academic_year_id=target_year_id,
                    note=data.get("note", ""),
                    promoted_by=request.user,
                )

                # Update student's current class and section
                st.current_class_id = target_class_id
                st.current_section_id = target_section_id
                st.save(update_fields=["current_class", "current_section", "updated_at"])
                promoted_count += 1

            except Exception as e:
                failed_count += 1
                errors_list.append({
                    "student_id": st.id,
                    "admission_no": st.admission_no,
                    "error": str(e)
                })
                continue

        # Return comprehensive response
        response_data = {
            "promoted": promoted_count,
            "failed": failed_count,
            "total": promoted_count + failed_count,
            "success": failed_count == 0
        }

        if errors_list:
            # Limit errors to first 50 for response
            response_data["errors"] = errors_list[:50]

        # Determine status code based on promotion success
        if promoted_count == 0:
            # Complete failure
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
        elif failed_count > 0:
            # Partial success
            return Response(response_data, status=status.HTTP_200_OK)
        else:
            # Complete success
            return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="auto-assign-classes")
    def auto_assign_classes(self, request):
        """
        Auto-assign current_class and current_section to students who are missing either.
        Uses date_of_birth to pick the appropriate grade where available, otherwise
        distributes round-robin across classes.
        Pass ?dry_run=true to preview without saving.
        """
        from datetime import date as date_type
        import re as _re
        from django.db import models as db_models
        from apps.core.models import Class, Section

        school_id = request.user.school_id
        if not school_id and not request.user.is_superuser:
            return Response({"error": "School context required."}, status=status.HTTP_403_FORBIDDEN)

        dry_run = str(request.data.get("dry_run", "false")).lower() in ("1", "true", "yes")
        ref_year = int(request.data.get("year", date_type.today().year))

        unassigned_qs = Student.objects.filter(
            is_deleted=False,
            **self.school_filter(request),
        ).filter(
            db_models.Q(current_class__isnull=True) | db_models.Q(current_section__isnull=True)
        )
        students = list(unassigned_qs.select_related("current_class").order_by("id"))

        if not students:
            return Response({"assigned": 0, "message": "All students already have class and section assigned."})

        # Load classes with their first section, ordered correctly
        cls_qs = (
            Class.objects.filter(**self.school_filter(request))
            .prefetch_related("sections")
            .order_by("numeric_order", "name")
        )
        school_classes = []
        for c in cls_qs:
            secs = list(c.sections.order_by("name"))
            if secs:
                m = _re.search(r"\d+", c.name)
                school_classes.append({
                    "cls": c,
                    "sections": secs,
                    "grade": int(m.group()) if m else 0,
                })

        if not school_classes:
            return Response({"error": "No classes with sections found."}, status=status.HTTP_400_BAD_REQUEST)

        def grade_from_dob(dob):
            age = ref_year - dob.year - (1 if (dob.month, dob.day) > (4, 1) else 0)
            return max(0, min(age - 4, 12))

        results = []
        to_update = []

        for idx, student in enumerate(students):
            chosen = None

            if student.date_of_birth and not student.current_class_id:
                grade = grade_from_dob(student.date_of_birth)
                if grade <= 0:
                    for entry in school_classes:
                        if entry["cls"].name.upper() in ("LKG", "UKG", "NURSERY"):
                            chosen = entry
                            break
                if not chosen:
                    best, best_diff = None, 9999
                    for entry in school_classes:
                        diff = abs(entry["grade"] - grade)
                        if diff < best_diff:
                            best_diff = diff
                            best = entry
                    chosen = best

            if not chosen:
                chosen = school_classes[idx % len(school_classes)]

            # Determine section: if student already has a class, use a section from it
            if student.current_class_id:
                existing_secs = list(Section.objects.filter(school_class_id=student.current_class_id).order_by("name"))
                target_section = existing_secs[0] if existing_secs else chosen["sections"][0]
                target_class = student.current_class
            else:
                target_class = chosen["cls"]
                target_section = chosen["sections"][0]

            results.append({
                "student_id": student.id,
                "admission_no": student.admission_no,
                "name": f"{student.first_name} {student.last_name}".strip(),
                "assigned_class": target_class.name,
                "assigned_section": target_section.name,
            })

            if not dry_run:
                update_fields = {}
                if not student.current_class_id:
                    update_fields["current_class_id"] = target_class.id
                if not student.current_section_id:
                    update_fields["current_section_id"] = target_section.id
                if update_fields:
                    to_update.append((student.pk, update_fields))

        if not dry_run:
            with transaction.atomic():
                for pk, fields in to_update:
                    Student.objects.filter(pk=pk).update(**fields)

        return Response({
            "assigned": len(results),
            "dry_run": dry_run,
            "students": results[:100],  # cap response at 100 rows
        })

    def destroy(self, request, *args, **kwargs):
        return self.soft_delete(request, *args, **kwargs)

    def _linked_record_exists(self, student):
        from apps.attendance.models import StudentAttendance, SubjectAttendance
        from apps.exams.models import ExamAttendanceChild, ExamMark, ExamMarkRegister, OnlineExamTake
        from apps.fees.models import FeesAssignment, FeesPayment

        return (
            StudentAttendance.objects.filter(student=student).exists()
            or SubjectAttendance.objects.filter(student=student).exists()
            or ExamMark.objects.filter(student=student).exists()
            or ExamAttendanceChild.objects.filter(student=student).exists()
            or ExamMarkRegister.objects.filter(student=student).exists()
            or OnlineExamTake.objects.filter(student=student).exists()
            or FeesAssignment.objects.filter(student=student).exists()
            or FeesPayment.objects.filter(student=student).exists()
        )

    def _audit_log(self, *, action, student, request, note="", metadata=None):
        try:
            school = getattr(student, "school", None)
            if not school:
                return
            StudentRecordAudit.objects.create(
                school=school,
                student=student,
                action=action,
                performed_by=request.user,
                note=note,
                metadata=metadata or {},
            )
        except Exception:
            # Audit failures should not block primary record operations.
            logger.exception("Unable to write student audit log")

    def _get_student_for_record_action(self, request, pk):
        qs = Student.objects.select_related("school")
        if not request.user.is_superuser:
            qs = qs.filter(school_id=request.user.school_id)
        return qs.filter(pk=pk).first()

    @action(detail=True, methods=["post"], url_path="soft-delete")
    def soft_delete(self, request, *args, **kwargs):
        student = self._get_student_for_record_action(request, kwargs.get("pk"))
        if not student:
            return Response(
                {"success": False, "message": "Student record not found", "field_errors": {}},
                status=status.HTTP_404_NOT_FOUND,
            )
        if student.is_deleted:
            return Response(
                {"success": False, "message": "Student record already deleted", "field_errors": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if self._linked_record_exists(student):
            return Response(
                {"success": False, "message": "Cannot delete student. Linked records exist", "field_errors": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        student.is_deleted = True
        student.is_active = False
        student.is_disabled = True
        student.status = "deleted"
        student.deleted_at = timezone.now()
        student.deleted_by = request.user
        student.save(update_fields=["is_deleted", "is_active", "is_disabled", "status", "deleted_at", "deleted_by", "updated_at"])

        self._audit_log(
            action=StudentRecordAudit.ACTION_SOFT_DELETE,
            student=student,
            request=request,
            note="Student soft-deleted",
            metadata={"student_id": student.id, "admission_no": student.admission_no},
        )

        return Response({"success": True, "message": "Student deleted successfully"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, *args, **kwargs):
        student = self._get_student_for_record_action(request, kwargs.get("pk"))
        if not student:
            return Response(
                {"success": False, "message": "Student record not found", "field_errors": {}},
                status=status.HTTP_404_NOT_FOUND,
            )
        if not student.is_deleted:
            return Response(
                {"success": False, "message": "Student is not deleted", "field_errors": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            with transaction.atomic():
                student.is_deleted = False
                student.is_disabled = False
                student.is_active = True
                student.status = "active"
                student.deleted_at = None
                student.deleted_by = None
                student.save(update_fields=["is_deleted", "is_disabled", "is_active", "status", "deleted_at", "deleted_by", "updated_at"])

                self._audit_log(
                    action=StudentRecordAudit.ACTION_RESTORE,
                    student=student,
                    request=request,
                    note="Student restored",
                    metadata={"student_id": student.id, "admission_no": student.admission_no},
                )
        except IntegrityError:
            return Response(
                {"success": False, "message": "Unable to restore student due to data dependency.", "field_errors": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            logger.exception("Restore student failed")
            return Response(
                {"success": False, "message": "Unable to restore student at this time.", "field_errors": {}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"success": True, "message": "Student restored successfully"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["delete"], url_path="permanent-delete")
    def permanent_delete(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response(
                {"success": False, "message": "You do not have permission to delete student records", "field_errors": {}},
                status=status.HTTP_403_FORBIDDEN,
            )

        student = self._get_student_for_record_action(request, kwargs.get("pk"))
        if not student:
            return Response(
                {"success": False, "message": "Student record not found", "field_errors": {}},
                status=status.HTTP_404_NOT_FOUND,
            )
        if self._linked_record_exists(student):
            return Response(
                {
                    "success": False,
                    "message": "Unable to permanently delete student due to linked records",
                    "field_errors": {},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            with transaction.atomic():
                self._audit_log(
                    action=StudentRecordAudit.ACTION_PERMANENT_DELETE,
                    student=student,
                    request=request,
                    note="Student permanently deleted",
                    metadata={"student_id": student.id, "admission_no": student.admission_no},
                )
                student.delete()
        except ProtectedError:
            return Response(
                {
                    "success": False,
                    "message": "Unable to permanently delete student due to linked records",
                    "field_errors": {},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except IntegrityError:
            return Response(
                {
                    "success": False,
                    "message": "Unable to permanently delete student due to data dependency.",
                    "field_errors": {},
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            return Response(
                {
                    "success": False,
                    "message": "Unable to permanently delete student at this time.",
                    "field_errors": {},
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"success": True, "message": "Student permanently deleted successfully"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="subject-assignment-stats")
    def subject_assignment_stats(self, request):
        """Return KPI counts for the Multi Subject Assignment page."""
        user = request.user
        base_qs = Student.objects.filter(is_deleted=False)
        if not user.is_superuser:
            base_qs = base_qs.filter(school_id=user.school_id)

        enrolled = base_qs.count()
        # Each student gets annotated with how many *optional* subject
        # assignments they have.  4 = at least one per category (L2, L3,
        # sport, art) → "Assigned".  1-3 → "Partial".  0 → "Pending".
        annotated = base_qs.annotate(
            opt_count=Count(
                "subject_assignments",
                filter=Q(subject_assignments__is_optional=True),
                distinct=True,
            )
        )
        pending = annotated.filter(opt_count=0).count()
        assigned = annotated.filter(opt_count__gte=4).count()
        partial = max(0, enrolled - pending - assigned)

        return Response(
            {
                "enrolled": enrolled,
                "assigned": assigned,
                "partial": partial,
                "pending": pending,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="class-section-tree")
    def class_section_tree(self, request):
        """Return classes → sections → students tree with subject assignment data."""
        from apps.core.models import Class, Section

        user = request.user
        school_id = None if user.is_superuser else user.school_id

        # Fetch all classes for this school
        cls_qs = Class.objects.all()
        if school_id:
            cls_qs = cls_qs.filter(school_id=school_id)
        cls_qs = cls_qs.prefetch_related("sections")

        # Prefetch all students with their optional subject assignments
        student_qs = Student.objects.filter(
            is_deleted=False,
            current_class__isnull=False,
        ).select_related("current_class", "current_section").prefetch_related(
            "subject_assignments__subject"
        )
        if school_id:
            student_qs = student_qs.filter(school_id=school_id)

        # Build a lookup: {(class_id, section_id): [student, ...]}
        from collections import defaultdict
        section_map = defaultdict(list)
        for st in student_qs:
            key = (st.current_class_id, st.current_section_id)
            section_map[key].append(st)

        result = []
        for cls in cls_qs:
            sections_out = []
            for sec in cls.sections.all():
                students_here = section_map.get((cls.id, sec.id), [])
                students_out = []
                for st in students_here:
                    opt_subs = [
                        a.subject.name
                        for a in st.subject_assignments.all()
                        if a.is_optional
                    ]
                    count = len(opt_subs)
                    sstatus = "done" if count >= 4 else ("partial" if count > 0 else "empty")
                    # Positionally map to L2/L3/SP/AR for display
                    students_out.append({
                        "id": st.id,
                        "name": f"{st.first_name} {st.last_name}".strip(),
                        "admNo": st.admission_no,
                        "rollNo": st.roll_no or "",
                        "lang2": opt_subs[0] if len(opt_subs) > 0 else "",
                        "lang3": opt_subs[1] if len(opt_subs) > 1 else "",
                        "sport": opt_subs[2] if len(opt_subs) > 2 else "",
                        "art":   opt_subs[3] if len(opt_subs) > 3 else "",
                        "optionalSubjects": opt_subs,
                        "status": sstatus,
                    })
                sections_out.append({
                    "id": sec.id,
                    "letter": sec.name,
                    "teacher": "",
                    "students": students_out,
                })
            result.append({
                "id": cls.id,
                "label": cls.name,
                "sections": sections_out,
            })

        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return self._validation_response({"file": ["Please upload a file"]})

        filename = (upload.name or "").lower()
        if not filename.endswith((".csv", ".xlsx")):
            return self._validation_response({"file": ["Only CSV or XLSX files are supported"]})

        rows = []
        if filename.endswith(".csv"):
            decoded = upload.read().decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            try:
                from openpyxl import load_workbook
            except Exception:
                return self._validation_response({"file": ["XLSX import requires openpyxl installation"]})
            workbook = load_workbook(upload, read_only=True)
            worksheet = workbook.active
            header = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            for data_row in worksheet.iter_rows(min_row=2, values_only=True):
                rows.append({header[index]: value for index, value in enumerate(data_row)})

        created = 0
        errors = []
        for index, row in enumerate(rows, start=2):
            normalized_row = self._normalize_import_row(row)
            if not any(str(value or "").strip() for value in normalized_row.values()):
                continue

            serializer = self.get_serializer(data=normalized_row)
            if serializer.is_valid():
                try:
                    self.perform_create(serializer)
                    created += 1
                except IntegrityError:
                    errors.append({"row": index, "field_errors": {"admission_no": "Admission number already exists"}})
            else:
                errors.append({"row": index, "field_errors": self._normalize_field_errors(serializer.errors)})

        return Response(
            {
                "success": len(errors) == 0,
                "message": "Bulk import completed",
                "created": created,
                "failed": len(errors),
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )


class StudentDocumentViewSet(TenantScopedModelViewSet):
    serializer_class = StudentDocumentSerializer
    model = StudentDocument
    permission_codes = {"*": "student_info.student_list.view"}

    def get_queryset(self):
        user = self.request.user
        qs = StudentDocument.objects.select_related("student__school")
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(student__school_id=user.school_id)
        return qs.none()

<<<<<<< HEAD
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def upload_document(self, request):
        """
        Upload a student document.
        
        Expects:
        - student_id (required): Student ID (database ID or UUID)
        - document_type (required): Document type (birth_certificate, aadhaar_card, medical_information)
        - file (required): File to upload
        """
        try:
            student_id_input = request.data.get("student_id")
            document_type = request.data.get("document_type")
            file_obj = request.FILES.get("file")
            
            logger.info(f"Document upload attempt: student_id={student_id_input}, document_type={document_type}, file={file_obj.name if file_obj else None}")
            
            # Validations
            if not student_id_input:
                return Response(
                    {"error": "Student ID is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not document_type:
                return Response(
                    {"error": "Document type is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not file_obj:
                return Response(
                    {"error": "No file selected."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check student exists and user has access.
            # The frontend may send either the numeric DB id or the UUID student_id.
            # We detect which one it is up front — querying a UUIDField with a
            # non-UUID string raises django.core.exceptions.ValidationError, not
            # Student.DoesNotExist, so we must validate the format ourselves.
            import uuid as _uuid
            student = None
            sid_str = str(student_id_input).strip()

            if sid_str.isdigit():
                try:
                    student = Student.objects.get(id=int(sid_str))
                except Student.DoesNotExist:
                    student = None

            if student is None:
                try:
                    _uuid.UUID(sid_str)
                except ValueError:
                    pass
                else:
                    try:
                        student = Student.objects.get(student_id=sid_str)
                    except Student.DoesNotExist:
                        student = None

            if student is None:
                logger.warning(f"Student not found with ID: {student_id_input}")
                return Response(
                    {"error": "Student not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Check permission
            if not request.user.is_superuser and student.school_id != request.user.school_id:
                logger.warning(f"Permission denied: user school {request.user.school_id} != student school {student.school_id}")
                raise PermissionDenied("You don't have permission to upload documents for this student.")
            
            # Validate file type
            allowed_extensions = [".pdf", ".jpg", ".jpeg", ".png"]
            file_name_lower = file_obj.name.lower()
            if not any(file_name_lower.endswith(ext) for ext in allowed_extensions):
                return Response(
                    {"error": "Only PDF, JPG, JPEG, and PNG files are allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate file size (5MB)
            if file_obj.size > 5242880:
                return Response(
                    {"error": "File size must be less than 5MB."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create document record
            document = StudentDocument.objects.create(
                student=student,
                school=student.school,
                document_type=document_type,
                title=document_type.replace("_", " ").title(),
                file=file_obj,
                original_name=file_obj.name,
                file_size=file_obj.size,
                uploaded_by=request.user,
            )
            
            # Serialize and return
            serializer = self.get_serializer(document)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            logger.error(f"Document upload error: {str(e)}", exc_info=True)
            return Response(
                {"error": "An error occurred during file upload. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class StudentTransferHistoryViewSet(TenantScopedModelViewSet):
    serializer_class = StudentTransferHistorySerializer
    model = StudentTransferHistory
    permission_codes = {"*": "student_info.delete_student_record.view"}

    def get_queryset(self):
        user = self.request.user
        qs = StudentTransferHistory.objects.select_related("student__school", "from_school", "to_school")
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(student__school_id=user.school_id)
        return qs.none()


class PromotionBatchViewSet(TenantScopedModelViewSet):
    serializer_class = PromotionBatchSerializer
    model = PromotionBatch
    permission_codes = {"*": "student_info.student_promote.view"}

    def get_queryset(self):
        user = self.request.user
        qs = PromotionBatch.objects.select_related("school", "academic_year", "target_year", "created_by", "confirmed_by").prefetch_related(
            "records",
            "records__student",
            "records__from_class",
            "records__from_section",
            "records__to_class",
            "records__to_section",
            "records__failed_subjects",
        )
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(school_id=user.school_id)
        return qs.none()

    def _recompute_counts(self, batch: PromotionBatch):
        total = batch.records.count()
        promoted = batch.records.filter(status=PromotionRecord.STATUS_PROMOTE).count()
        retained = batch.records.filter(status=PromotionRecord.STATUS_NOT_PROMOTED).count()
        batch.total_students = total
        batch.promoted_count = promoted
        batch.retained_count = retained
        batch.save(update_fields=["total_students", "promoted_count", "retained_count"])

    def _get_client_ip(self, request):
        forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
        if forwarded:
            return forwarded.split(",")[0].strip()
        return request.META.get("REMOTE_ADDR", "")

    def _initialize_records(self, batch: PromotionBatch):
        from apps.core.models import Class

        if batch.records.exists():
            self._recompute_counts(batch)
            return

        classes = list(Class.objects.filter(school_id=batch.school_id).order_by("numeric_order", "id"))
        next_map = {}
        for idx, row in enumerate(classes):
            next_map[row.id] = classes[idx + 1].id if idx + 1 < len(classes) else None

        students = Student.objects.filter(
            school_id=batch.school_id,
            academic_year_id=batch.academic_year_id,
            is_active=True,
            is_deleted=False,
        ).select_related("current_class", "current_section")

        to_create = []
        for student in students:
            to_create.append(
                PromotionRecord(
                    batch=batch,
                    student=student,
                    from_class=student.current_class,
                    from_section=student.current_section,
                    to_class_id=next_map.get(student.current_class_id),
                    status=PromotionRecord.STATUS_PENDING,
                )
            )

        PromotionRecord.objects.bulk_create(to_create, ignore_conflicts=True)
        if to_create:
            batch.status = PromotionBatch.STATUS_IN_PROGRESS
            batch.save(update_fields=["status"])
        self._recompute_counts(batch)

    @action(detail=False, methods=["post"], url_path="create-or-get")
    def create_or_get(self, request):
        serializer = PromotionBatchCreateSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        school = request.user.school
        year = data["academic_year_obj"]
        target = data["target_year_obj"]

        batch, created = PromotionBatch.objects.get_or_create(
            school=school,
            academic_year=year,
            defaults={
                "target_year": target,
                "created_by": request.user,
                "status": PromotionBatch.STATUS_DRAFT,
            },
        )

        if not created and batch.status == PromotionBatch.STATUS_DRAFT and batch.target_year_id != target.id:
            batch.target_year = target
            batch.save(update_fields=["target_year"])

        self._initialize_records(batch)
        return Response(self.get_serializer(batch).data)

    @action(detail=False, methods=["get"], url_path=r"by-year/(?P<year_name>[^/.]+)")
    def by_year(self, request, year_name=None):
        from apps.core.models import AcademicYear

        school_id = request.user.school_id
        year_qs = AcademicYear.objects.filter(name=year_name)
        if not request.user.is_superuser:
            year_qs = year_qs.filter(school_id=school_id)
        year = year_qs.first()
        if not year:
            return Response({"error": "Academic year not found"}, status=status.HTTP_404_NOT_FOUND)

        batch_qs = self.get_queryset().filter(academic_year=year)
        batch = batch_qs.first()
        if not batch:
            return Response({"error": "No promotion batch for this year"}, status=status.HTTP_404_NOT_FOUND)

        return Response(self.get_serializer(batch).data)

    @action(detail=True, methods=["post"], url_path="update-record")
    def update_record(self, request, pk=None):
        from apps.core.models import Section, Subject

        batch = self.get_object()
        serializer = PromotionRecordUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            record = batch.records.get(id=data["record_id"])
        except PromotionRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

        if "to_class" in data:
            record.to_class_id = data.get("to_class")
        if "to_section" in data:
            section_id = data.get("to_section")
            if section_id:
                section = Section.objects.filter(id=section_id).first()
                if not section:
                    raise ValidationError({"to_section": "Invalid section."})
                if record.to_class_id and section.school_class_id != record.to_class_id:
                    raise ValidationError({"to_section": "Section must belong to selected class."})
            record.to_section_id = section_id

        record.status = data["status"]
        if record.status == PromotionRecord.STATUS_NOT_PROMOTED:
            record.retention_reason = data.get("retention_reason", "")
            if not record.retention_reason:
                raise ValidationError({"retention_reason": "Retention reason is required for not promoted status."})
        else:
            record.retention_reason = ""

        if "notes" in data:
            record.notes = data.get("notes", "")

        record.decision_made_at = timezone.now()
        record.decision_made_by = request.user
        record.last_modified_by = request.user
        record.save()

        if "failed_subject_ids" in data:
            subjects = Subject.objects.filter(id__in=data.get("failed_subject_ids", []), school_id=batch.school_id)
            record.failed_subjects.set(subjects)

        PromotionAuditLog.objects.create(
            batch=batch,
            action="record_updated",
            performed_by=request.user,
            record=record,
            details=f"Record updated for {record.student.admission_no}",
            ip_address=self._get_client_ip(request),
        )

        self._recompute_counts(batch)
        return Response(PromotionRecordSerializer(record).data)

    @action(detail=True, methods=["post"], url_path="bulk-update")
    def bulk_update(self, request, pk=None):
        batch = self.get_object()
        serializer = PromotionBulkUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        records = batch.records.all()
        scope = data["scope"]
        if scope == "class":
            records = records.filter(from_class_id=data["class_id"])
        elif scope == "section":
            records = records.filter(from_section_id=data["section_id"])
        else:
            records = records.filter(id__in=data["record_ids"])

        now = timezone.now()
        action_name = data["action"]
        changed = 0
        for record in records:
            if action_name == "promote":
                if not record.to_class_id:
                    continue
                record.status = PromotionRecord.STATUS_PROMOTE
                record.retention_reason = ""
            elif action_name == "skip":
                record.status = PromotionRecord.STATUS_NOT_PROMOTED
                record.retention_reason = record.retention_reason or PromotionRecord.REASON_OTHER
            else:
                record.status = PromotionRecord.STATUS_PENDING
                record.retention_reason = ""
                record.notes = ""
                record.failed_subjects.clear()

            record.decision_made_at = now
            record.decision_made_by = request.user
            record.last_modified_by = request.user
            record.save()
            changed += 1

        PromotionAuditLog.objects.create(
            batch=batch,
            action=f"bulk_{action_name}",
            performed_by=request.user,
            details=f"Bulk action {action_name} applied to {changed} record(s)",
            ip_address=self._get_client_ip(request),
        )

        self._recompute_counts(batch)
        return Response({"updated": changed, "batch": self.get_serializer(batch).data})

    @action(detail=True, methods=["post"], url_path="ai-recommendation")
    def ai_recommendation(self, request, pk=None):
        batch = self.get_object()
        serializer = PromotionAiRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            record = batch.records.select_related("student", "from_class").get(id=data["record_id"])
        except PromotionRecord.DoesNotExist:
            return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)

        reason = data.get("reason") or record.retention_reason or "other"
        failed_subject_names = list(record.failed_subjects.values_list("name", flat=True))
        parts = [
            f"Student {record.student.first_name} {record.student.last_name} requires review.",
            f"Primary reason: {reason.replace('_', ' ')}.",
        ]
        if failed_subject_names:
            parts.append(f"Focus subjects: {', '.join(failed_subject_names)}.")
        parts.append("Recommendation: schedule parent meeting, assign remedial plan, and re-evaluate within 6-8 weeks.")
        recommendation = " ".join(parts)

        record.ai_recommendation = recommendation
        record.last_modified_by = request.user
        record.save(update_fields=["ai_recommendation", "last_modified_by", "last_modified_at"])

        PromotionAuditLog.objects.create(
            batch=batch,
            action="ai_generated",
            performed_by=request.user,
            record=record,
            details=f"AI recommendation generated for {record.student.admission_no}",
            ip_address=self._get_client_ip(request),
        )

        return Response({"recommendation": recommendation})

    @action(detail=True, methods=["post"], url_path="confirm")
    def confirm(self, request, pk=None):
        batch = self.get_object()
        if batch.status == PromotionBatch.STATUS_FINALIZED:
            return Response({"error": "Finalized batch cannot be modified."}, status=status.HTTP_400_BAD_REQUEST)

        promote_records = batch.records.filter(status=PromotionRecord.STATUS_PROMOTE).select_related(
            "student",
            "to_class",
            "to_section",
        )

        promoted = 0
        with transaction.atomic():
            for record in promote_records:
                if not record.to_class_id:
                    continue

                student = record.student
                StudentPromotionHistory.objects.create(
                    student=student,
                    from_class=student.current_class,
                    from_section=student.current_section,
                    to_class=record.to_class,
                    to_section=record.to_section,
                    from_academic_year=batch.academic_year,
                    to_academic_year=batch.target_year,
                    note="Promoted from Promotion Batch",
                    promoted_by=request.user,
                )
                student.current_class = record.to_class
                student.current_section = record.to_section
                student.academic_year = batch.target_year
                student.save(update_fields=["current_class", "current_section", "academic_year", "updated_at"])
                promoted += 1

            batch.status = PromotionBatch.STATUS_CONFIRMED
            batch.confirmed_at = timezone.now()
            batch.confirmed_by = request.user
            batch.save(update_fields=["status", "confirmed_at", "confirmed_by"])

        PromotionAuditLog.objects.create(
            batch=batch,
            action="confirmed",
            performed_by=request.user,
            details=f"Batch confirmed. Promoted {promoted} students.",
            ip_address=self._get_client_ip(request),
        )

        self._recompute_counts(batch)
        return Response({"message": f"Promotion batch confirmed. Promoted {promoted} students.", "batch": self.get_serializer(batch).data})

    @action(detail=True, methods=["post"], url_path="finalize")
    def finalize(self, request, pk=None):
        batch = self.get_object()
        if batch.status != PromotionBatch.STATUS_CONFIRMED:
            return Response({"error": "Only confirmed batches can be finalized."}, status=status.HTTP_400_BAD_REQUEST)

        batch.status = PromotionBatch.STATUS_FINALIZED
        batch.save(update_fields=["status"])
        PromotionAuditLog.objects.create(
            batch=batch,
            action="finalized",
            performed_by=request.user,
            details="Batch finalized",
            ip_address=self._get_client_ip(request),
        )
        return Response(self.get_serializer(batch).data)


class PromotionAuditLogViewSet(TenantScopedModelViewSet):
    serializer_class = PromotionAuditLogSerializer
    model = PromotionAuditLog
    permission_codes = {"*": "student_info.student_promote.view"}

    def get_queryset(self):
        user = self.request.user
        qs = PromotionAuditLog.objects.select_related("batch", "performed_by", "record", "batch__school")
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(batch__school_id=user.school_id)
        return qs.none()


class StudentPromotionHistoryViewSet(TenantScopedModelViewSet):
    serializer_class = StudentPromotionHistorySerializer
    model = StudentPromotionHistory
    permission_codes = {"*": "student_info.student_promote.view"}

    def get_queryset(self):
        user = self.request.user
        qs = StudentPromotionHistory.objects.select_related(
            "student__school",
            "from_class",
            "from_section",
            "to_class",
            "to_section",
            "from_academic_year",
            "to_academic_year",
            "promoted_by",
        )
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(student__school_id=user.school_id)
        return qs.none()


class StudentMultiClassRecordViewSet(TenantScopedModelViewSet):
    serializer_class = StudentMultiClassRecordSerializer
    model = StudentMultiClassRecord
    permission_codes = {
        "*": "student_info.multi_class_student.view",
        "bulk_save": "student_info.multi_class_student.view",
    }

    def get_queryset(self):
        user = self.request.user
        qs = StudentMultiClassRecord.objects.select_related("student", "school_class", "section", "student__school")

        student_id = self.request.query_params.get("student")
        if student_id:
            qs = qs.filter(student_id=student_id)

        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(student__school_id=user.school_id)
        return qs.none()

    def perform_create(self, serializer):
        record = serializer.save()
        if record.is_default:
            StudentMultiClassRecord.objects.filter(student_id=record.student_id).exclude(id=record.id).update(is_default=False)
            Student.objects.filter(id=record.student_id).update(
                current_class_id=record.school_class_id,
                current_section_id=record.section_id,
            )

    def perform_update(self, serializer):
        record = serializer.save()
        if record.is_default:
            StudentMultiClassRecord.objects.filter(student_id=record.student_id).exclude(id=record.id).update(is_default=False)
            Student.objects.filter(id=record.student_id).update(
                current_class_id=record.school_class_id,
                current_section_id=record.section_id,
            )

    @action(detail=False, methods=["post"], url_path="bulk-save")
    def bulk_save(self, request):
        serializer = StudentMultiClassBulkSaveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        student_id = data["student_id"]
        user = request.user
        student_qs = Student.objects.filter(id=student_id)
        if not user.is_superuser:
            student_qs = student_qs.filter(school_id=user.school_id)
        student = student_qs.first()
        if not student:
            raise ValidationError("Student not found in your school.")

        records = data.get("records", [])
        default_seen = False

        with transaction.atomic():
            StudentMultiClassRecord.objects.filter(student=student).delete()
            created = []

            for item in records:
                school_class_id = item["school_class"]
                section_id = item.get("section")
                is_default = bool(item.get("is_default", False))
                if is_default:
                    if default_seen:
                        is_default = False
                    default_seen = True

                from apps.core.models import Class, Section

                class_qs = Class.objects.filter(id=school_class_id)
                if not user.is_superuser:
                    class_qs = class_qs.filter(school_id=student.school_id)
                if not class_qs.exists():
                    raise ValidationError("One or more classes are invalid for this student school.")

                if section_id:
                    section_qs = Section.objects.filter(id=section_id, school_class_id=school_class_id)
                    if not section_qs.exists():
                        raise ValidationError("One or more sections do not belong to selected class.")

                created.append(
                    StudentMultiClassRecord.objects.create(
                        student=student,
                        school_class_id=school_class_id,
                        section_id=section_id,
                        roll_no=item.get("roll_no", ""),
                        is_default=is_default,
                    )
                )

            default_record = next((item for item in created if item.is_default), None)
            if not default_record and created:
                default_record = created[0]
                default_record.is_default = True
                default_record.save(update_fields=["is_default", "updated_at"])

            if default_record:
                student.current_class_id = default_record.school_class_id
                student.current_section_id = default_record.section_id
                student.save(update_fields=["current_class", "current_section", "updated_at"])

        response_data = StudentMultiClassRecordSerializer(
            StudentMultiClassRecord.objects.filter(student=student).order_by("-is_default", "id"),
            many=True,
        ).data
        return Response({"student_id": student.id, "records": response_data}, status=status.HTTP_200_OK)


class StudentSubjectAssignmentViewSet(TenantScopedModelViewSet):
    serializer_class = StudentSubjectAssignmentSerializer
    model = StudentSubjectAssignment
    permission_codes = {
        "*": "student_info.multi_class_student.view",
        "assign_individual": "student_info.multi_class_student.view",
        "assign_bulk": "student_info.multi_class_student.view",
        "upsert_optional": "student_info.multi_class_student.view",
    }

    def get_queryset(self):
        user = self.request.user
        qs = StudentSubjectAssignment.objects.select_related(
            "student",
            "subject",
            "academic_year",
            "school_class",
            "section",
            "assigned_by",
        )
        student_id = self.request.query_params.get("student_id")
        class_id = self.request.query_params.get("class") or self.request.query_params.get("school_class")
        section_id = self.request.query_params.get("section")
        academic_year_id = self.request.query_params.get("academic_year")
        if student_id:
            qs = qs.filter(student_id=student_id)
        if class_id:
            qs = qs.filter(school_class_id=class_id)
        if section_id:
            qs = qs.filter(section_id=section_id)
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)
        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(student__school_id=user.school_id)
        return qs.none()

    def _validation_response(self, field_errors=None, message="Validation failed"):
        normalized = {}
        for key, value in (field_errors or {}).items():
            if isinstance(value, (list, tuple)):
                normalized[key] = str(value[0]) if value else "Validation failed"
            else:
                normalized[key] = str(value)
        return Response(
            {
                "success": False,
                "message": message,
                "field_errors": normalized,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _resolve_students(self, *, student_ids, class_id, section_id, user):
        queryset = Student.objects.filter(id__in=student_ids, current_class_id=class_id, current_section_id=section_id)
        if not user.is_superuser:
            queryset = queryset.filter(school_id=user.school_id)
        return list(queryset)

    def _create_assignments(self, *, students, subject_ids, academic_year_id, class_id, section_id, is_optional, user):
        existing = StudentSubjectAssignment.objects.filter(
            student_id__in=[row.id for row in students],
            subject_id__in=subject_ids,
            academic_year_id=academic_year_id,
        )
        if existing.exists():
            return self._validation_response({"subject_ids": "Subject already assigned to this student"})

        payload = [
            StudentSubjectAssignment(
                student_id=student.id,
                subject_id=subject_id,
                academic_year_id=academic_year_id,
                school_class_id=class_id,
                section_id=section_id,
                is_optional=is_optional,
                assigned_by=user,
            )
            for student in students
            for subject_id in subject_ids
        ]
        StudentSubjectAssignment.objects.bulk_create(payload)
        return None

    @action(detail=False, methods=["post"], url_path="assign-individual")
    def assign_individual(self, request):
        serializer = StudentSubjectAssignmentRequestSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            return self._validation_response(serializer.errors)

        data = serializer.validated_data
        student_id = request.data.get("student_id")
        if not student_id:
            return self._validation_response({"student_id": "Please select a student"})

        students = self._resolve_students(
            student_ids=[int(student_id)],
            class_id=data["school_class"],
            section_id=data["section"],
            user=request.user,
        )
        if not students:
            return self._validation_response({"student_id": "Unable to load students. Please try again"})

        with transaction.atomic():
            error = self._create_assignments(
                students=students,
                subject_ids=data["subject_ids"],
                academic_year_id=data["academic_year"],
                class_id=data["school_class"],
                section_id=data["section"],
                is_optional=data.get("is_optional", False),
                user=request.user,
            )
            if error:
                return error

        return Response({"success": True, "message": "Subjects assigned successfully"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="assign-bulk")
    def assign_bulk(self, request):
        serializer = StudentSubjectAssignmentRequestSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            return self._validation_response(serializer.errors)

        data = serializer.validated_data
        student_ids = data.get("student_ids") or request.data.get("student_ids") or []

        if student_ids:
            students = self._resolve_students(
                student_ids=student_ids,
                class_id=data["school_class"],
                section_id=data["section"],
                user=request.user,
            )
        else:
            students_qs = Student.objects.filter(current_class_id=data["school_class"], current_section_id=data["section"])
            if not request.user.is_superuser:
                students_qs = students_qs.filter(school_id=request.user.school_id)
            students = list(students_qs)

        if not students:
            return self._validation_response({"students": "Unable to load students. Please try again"})

        with transaction.atomic():
            error = self._create_assignments(
                students=students,
                subject_ids=data["subject_ids"],
                academic_year_id=data["academic_year"],
                class_id=data["school_class"],
                section_id=data["section"],
                is_optional=data.get("is_optional", False),
                user=request.user,
            )
            if error:
                return error

        return Response({"success": True, "message": "Subjects assigned successfully"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="upsert-optional")
    def upsert_optional(self, request):
        """
        Upsert optional subjects for one student by subject name.
        Replaces all existing optional assignments.
        Payload: { student_id, subject_names: [...] }
        """
        import traceback
        from apps.core.models import AcademicYear, Subject

        try:
            student_id = request.data.get("student_id")
            subject_names = request.data.get("subject_names", [])

            if not student_id:
                return Response({"success": False, "message": "student_id is required"}, status=status.HTTP_400_BAD_REQUEST)

            user = request.user
            student_qs = Student.objects.filter(id=student_id)
            if not user.is_superuser and getattr(user, "school_id", None):
                student_qs = student_qs.filter(school_id=user.school_id)
            student = student_qs.select_related("current_class", "current_section").first()
            if not student:
                return Response({"success": False, "message": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

            # Determine active academic year for this school
            school_id = student.school_id
            year_qs = AcademicYear.objects.filter(is_current=True, school_id=school_id)
            year = year_qs.first()
            if not year:
                # Fallback: any current academic year
                year = AcademicYear.objects.filter(is_current=True).first()
            if not year:
                # Last fallback: most recent year for this school
                year = AcademicYear.objects.filter(school_id=school_id).order_by("-start_date").first()
            if not year:
                return Response({"success": False, "message": "No active academic year found"}, status=status.HTTP_400_BAD_REQUEST)

            # Resolve subjects by name (bulk lookup + get-or-create for missing)
            names_clean = [n.strip() for n in subject_names if n.strip()]
            subjects = []
            if names_clean:
                existing = {
                    s.name.lower(): s
                    for s in Subject.objects.filter(school_id=school_id, name__in=names_clean)
                }
                # Case-insensitive fallback for names not matched exactly
                lower_map = {k.lower(): v for k, v in existing.items()}
                for name in names_clean:
                    subj = lower_map.get(name.lower())
                    if not subj:
                        try:
                            subj, _ = Subject.objects.get_or_create(
                                school_id=school_id,
                                name=name,
                                defaults={"subject_type": "optional"},
                            )
                        except Exception:
                            subj = Subject.objects.filter(school_id=school_id, name__iexact=name).first()
                        if subj:
                            lower_map[name.lower()] = subj
                    if subj:
                        subjects.append(subj)

            with transaction.atomic():
                # Remove previous optional assignments for this student + academic year
                StudentSubjectAssignment.objects.filter(
                    student=student,
                    academic_year=year,
                    is_optional=True,
                ).delete()

                if subjects and student.current_class and student.current_section:
                    # Exclude subjects already assigned as non-optional (unique constraint guard)
                    non_optional_ids = set(
                        StudentSubjectAssignment.objects.filter(
                            student=student,
                            academic_year=year,
                            is_optional=False,
                        ).values_list("subject_id", flat=True)
                    )
                    new_assignments = [
                        StudentSubjectAssignment(
                            student=student,
                            subject=subj,
                            academic_year=year,
                            school_class=student.current_class,
                            section=student.current_section,
                            is_optional=True,
                            assigned_by=user,
                        )
                        for subj in subjects if subj.id not in non_optional_ids
                    ]
                    if new_assignments:
                        StudentSubjectAssignment.objects.bulk_create(new_assignments, ignore_conflicts=True)

            return Response(
                {"success": True, "message": "Optional subjects updated", "assigned_count": len(subjects)},
                status=status.HTTP_200_OK,
            )

        except Exception as exc:
            logger.error("upsert_optional failed: %s\n%s", exc, traceback.format_exc())
            return Response(
                {"success": False, "message": f"Server error: {exc}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentRecordAuditViewSet(TenantScopedModelViewSet):
    serializer_class = StudentRecordAuditSerializer
    model = StudentRecordAudit
    permission_codes = {"*": "student_info.delete_student_record.view"}
    http_method_names = ["get", "head", "options"]

    def get_queryset(self):
        user = self.request.user
        qs = StudentRecordAudit.objects.select_related("student", "performed_by", "school", "student__current_class", "student__current_section")

        student_id = self.request.query_params.get("student_id")
        action = (self.request.query_params.get("action") or "").strip()
        search = (self.request.query_params.get("search") or "").strip()
        class_id = self.request.query_params.get("class")
        section_id = self.request.query_params.get("section")

        if student_id:
            qs = qs.filter(student_id=student_id)
        if action:
            qs = qs.filter(action=action)
        if class_id:
            qs = qs.filter(student__current_class_id=class_id)
        if section_id:
            qs = qs.filter(student__current_section_id=section_id)
        if search:
            qs = qs.filter(
                Q(student__first_name__icontains=search)
                | Q(student__last_name__icontains=search)
                | Q(student__admission_no__icontains=search)
                | Q(note__icontains=search)
            )

        if user.is_superuser:
            return qs
        if user.school_id:
            return qs.filter(school_id=user.school_id)
        return qs.none()

